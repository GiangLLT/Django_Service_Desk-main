from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from oauth2_provider.views.generic import ProtectedResourceView

from django.core.paginator import Paginator
from django.db.models import Count
from .models import Product, Category, Users, Ticket, Company, TGroup, Comment, Attachment, TImage, Assign_User, Role_Group, Role_Single, Authorization_User, Menu, ReadComment, Ticket_Mapping

from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.urls import reverse
from urllib.parse import urlencode

from django.contrib.sessions.backends.db import SessionStore
from django.contrib.auth import login
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

from django.db.models import Q
from django.db.models import Prefetch
from datetime import datetime, timedelta
from time import strptime
import time

import datetime
import json
import adal
import requests
import msal
import os
import re
from django.utils.text import slugify

from django.core.files.storage import FileSystemStorage
from django.db.models import F
from django.db.models.functions import Extract
from django.db.models import Sum, F, FloatField
from django.db.models.functions import Cast

from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string

from bs4 import BeautifulSoup
import base64

from msal import PublicClientApplication, ConfidentialClientApplication
import base64
import hashlib
import random
import string

import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, Border, Side

import subprocess
from functools import reduce

from django.contrib.auth.decorators import login_required

import subprocess
from subprocess import Popen, PIPE

import git

# Create your views here.
############################################ PAGE TICKET HELPDESK - START ############################################################

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### GITHUB ##################### 
def write_data(file_content,file_name,file_path):
    # Tạo một đối tượng FileSystemStorage để quản lý việc lưu trữ
    fs = FileSystemStorage(location=settings.MEDIA_ROOT)
    date_now = datetime.datetime.now()
    time_now = datetime.datetime.now()

    # Tạo tên tệp
    file_name = date_now.strftime('%d%m%Y') + '_' + time_now.strftime('%H%M') + file_name
    # file_name = date_now.strftime('%d%m%Y') + '_' + time_now.strftime('%H%M') + '_log_file_github.txt'
    # file_content = 'Git fetch data: ' + date_now.strftime('%d/%m/%Y') + '_' + time_now.strftime('%H:%M')

    # Tạo tệp tạm thời và viết nội dung vào nó
    # temp_file_path = os.path.join(settings.MEDIA_ROOT, 'my_project/Logs', file_name)
    temp_file_path = os.path.join(settings.MEDIA_ROOT, file_path , file_name)
    with open(temp_file_path, 'w') as temp_file:
        temp_file.write(file_content)

    # Lưu tệp tạm thời vào FileSystemStorage
    with open(temp_file_path, 'rb') as temp_file:
        fs.save('my_project/Logs/' + file_name, temp_file)

    # Xoá tệp tạm thời
    os.remove(temp_file_path)

    return file_name

def run_cmd_github_test(request):
    try:
        working_directory = 'C:\\inetpub\\wwwroot\\Django_Service_Desk-main'
        cmd = [
            ['cd', working_directory],
            ['git', 'status'],
            ['git', 'fetch', 'origin', 'main'],
            
            ['git', 'merge', 'origin/main']
        ]
        log_contents = ''
        for command in cmd:
            proc = subprocess.Popen(command, shell=True, cwd=working_directory, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            stdout, stderr = proc.communicate()
            if proc.returncode != 0:
                return HttpResponse(f"Command failed with error: {stderr.decode('utf-8')}", status=500)
            log_contents += stdout.decode('utf-8')

        file_name = '_log_file_github.txt'
        file_path = 'my_project/Logs'
        log_file_name = write_data(log_contents,file_name,file_path)
        return HttpResponse(f"Lệnh đã chạy thành công và nội dung đã được ghi vào tệp log: {log_file_name}")

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
    
def run_cmd_github(request):
    try:
        sid = 'S-1-5-21-1716883635-4105501997-2570134539-1000'
        working_directory = 'C:\\inetpub\\wwwroot\\Django_Service_Desk-main'
        commands = [
            'cd /d {}'.format(working_directory),
            'git status',
            'git fetch origin main',
            'git merge origin/main'
        ]

        log_contents = ''
        for command in commands:
            # Chạy lệnh với quyền của người dùng có SID
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            stdout, stderr = process.communicate()

            if process.returncode != 0:
                return HttpResponse(f"Command failed with error: {stderr.decode('utf-8')}", status=500)

            log_contents += stdout.decode('utf-8')

        file_name = '_log_file_github.txt'
        file_path = 'my_project/Logs'
        log_file_name = write_data(log_contents, file_name, file_path)

        return HttpResponse(f"Lệnh đã chạy thành công và nội dung đã được ghi vào tệp log: {log_file_name}")

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

def read_bat_file(request):
    file_path = 'C:\\inetpub\\wwwroot\\Django_Service_Desk-main\\SDP.bat'
    log_contents = ''
    try:
        working_directory = 'C:\\inetpub\\wwwroot\\Django_Service_Desk-main'
        
        # Thực thi lệnh cmd trong file .bat
        proc = subprocess.Popen([file_path], shell=True, cwd=working_directory, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = proc.communicate()
        
        # Kiểm tra kết quả của lệnh cmd
        if proc.returncode == 0:
            log_contents = stdout.decode('utf-8')
            log_file_name = write_data(log_contents)
            return HttpResponse(f"Lệnh đã chạy thành công và nội dung đã được ghi vào tệp log: {log_file_name}")
        else:
            return HttpResponse(f"Command failed with error: {stderr.decode('utf-8')}", status=500)

    except FileNotFoundError:
        return HttpResponse(f"File .bat không tồn tại", status=404)
    except Exception as e:
        return HttpResponse(f"Lỗi: {str(e)}", status=500)

def get_github_data_for_file(request):
        # Thay thế các giá trị sau bằng thông tin GitHub của bạn và token OAuth
        github_username = "GiangLLT"
        repository_name = "Django_Service_Desk"
        branch_name = "Dev"
        local_directory = "C:\test"
        github_token = "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" 

        # Tạo headers với token OAuth để xác thực yêu cầu
        headers = {
            "Authorization": f"token {github_token}"
        }

        # Xác định URL của API GitHub để lấy danh sách tất cả các tệp và thư mục trên nhánh cụ thể
        github_api_url = f"https://api.github.com/repos/{github_username}/{repository_name}/git/trees/{branch_name}?recursive=1"

        # Gửi yêu cầu GET đến API GitHub với headers
        response = requests.get(github_api_url, headers=headers)

        if response.status_code == 200:
            # Lấy dữ liệu JSON từ phản hồi
            github_data = response.json()

            if 'tree' in github_data:
                # Lặp qua danh sách các tệp và thư mục
                for item in github_data['tree']:
                    if item['type'] == 'blob':
                        # Nếu là tệp, tải về và lưu vào máy tính
                        file_path = item['path']
                        file_content_url = f"https://raw.githubusercontent.com/{github_username}/{repository_name}/{branch_name}/{file_path}"
                        local_file_path = os.path.join(local_directory, file_path)
                        os.makedirs(os.path.dirname(local_file_path), exist_ok=True)
                        
                        # Tải về tệp từ GitHub
                        response = requests.get(file_content_url)
                        
                        if response.status_code == 200:
                            content = response.text
                            with open(local_file_path, 'w') as file:
                                file.write(content)
                        else:
                            return JsonResponse({"error": f"Failed to download file '{file_path}' from GitHub. Status code: {response.status_code}"}, status=400)
            
            return JsonResponse({"message": "All GitHub files downloaded successfully"})
        else:
            return JsonResponse({"error": f"Failed to fetch data from GitHub. Status code: {response.status_code}"}, status=400)
        
@csrf_exempt
def push_to_dev(request):
    if request.method == 'POST':
        try:
            isRun = request.POST.get('isRun')
            if isRun:
                now_date = datetime.datetime.now()
                commit_date = now_date.strftime('%d%m%Y')
                commit_time = now_date.strftime('%H%M')
                commit_name = f'Auto commit to Dev branch from Django - {commit_date} - {commit_time}'

                # Đường dẫn tới thư mục chứa repository Git của bạn
                repo_path = 'C:\Data\Document\Another_Code\Python_helpdesk\Python_Django'

                # Khởi tạo đối tượng Repo
                repo = git.Repo(repo_path)

                # Kiểm tra xem bạn đang ở nhánh nào
                current_branch = repo.active_branch.name

                # Nếu không phải là nhánh Dev, chuyển đến nhánh Dev
                if current_branch != 'Dev':
                    dev_branch = repo.branches['Dev']
                    dev_branch.checkout()

                # Commit và push lên nhánh hiện tại (có thể là Dev hoặc đã chuyển về Dev trước đó)
                repo.git.add(all=True)
                repo.git.commit('-m', commit_name)
                origin = repo.remote(name='origin')
                origin.push('Dev')
                return HttpResponse({'success': True, 'message': "Push to {current_branch} branch success!"})
            else:
                return HttpResponse({'success': False, 'message': "Push to {current_branch} branch Failed!"})
            
            # return HttpResponse(f"Push to {current_branch} branch success!", status=200)
        except Exception as e:
            return HttpResponse({'success': False, 'message': {e}})
            # return HttpResponse(f"Error: {e}", status=500)
    else:
        # return HttpResponse("Method not allowed", status=405)
        return HttpResponse({'success': False, 'message': "Method not allowed"})
    
@csrf_exempt
def push_to_main_and_merge(request):
    if request.method == 'POST':
        try:
            # now_date = datetime.datetime.now()
            # commit_date = now_date.strftime('%d%m%Y')
            # commit_time = now_date.strftime('%H%M')
            # commit_name = f'Auto commit to Dev branch from Django - {commit_date} - {commit_time}'
            # Đường dẫn tới thư mục chứa repository Git của bạn
            isRun = request.POST.get('isRun')
            if isRun:
                repo_path = 'C:\Data\Document\Another_Code\Python_helpdesk\Python_Django'

                # Khởi tạo đối tượng Repo
                repo = git.Repo(repo_path)

                # Kiểm tra xem bạn đang ở nhánh nào
                current_branch = repo.active_branch.name

                if current_branch != 'main':
                    # Chuyển đến nhánh Main
                    main_branch = repo.branches['main']
                    main_branch.checkout()

                # Merge từ nhánh Dev
                dev_branch = repo.branches['Dev']
                repo.git.merge(dev_branch)

                # Commit và push lên nhánh Main
                repo.git.add(all=True)
                # repo.git.commit('-m', commit_name)
                origin = repo.remote(name='origin')
                origin.push('main')

                dev_branch = repo.branches['Dev']
                dev_branch.checkout()
                # return HttpResponse("Push to Main and merge with Dev success!", status=200)
                return HttpResponse({'success': True, 'message': "Push to Main and merge with Dev success!"})
            else:
                return HttpResponse({'success': False, 'message': "Push to Main and merge with Dev failed!"})
        except Exception as e:
            # return HttpResponse(f"Error: {e}", status=500)
            return HttpResponse({'success': False, 'message': {e}})
    else:
        # return HttpResponse("Method not allowed", status=405)
        return HttpResponse({'success': False, 'message': "Method not allowed"})

def github_file_list(request):
    github_username = 'GiangLLT'
    repo_name = 'Django_Service_Desk-main'
    token = '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'
    api_url_dev = f'https://api.github.com/repos/{github_username}/{repo_name}/commits?sha=Dev'
    api_url_main = f'https://api.github.com/repos/{github_username}/{repo_name}/commits?sha=main'

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github.v3+json"
    }

    try:
        response_dev = requests.get(api_url_dev, headers=headers)
        response_main = requests.get(api_url_main, headers=headers)

        # Kiểm tra xem response có thành công không (status code 200)
        response_dev.raise_for_status()
        response_main.raise_for_status()
        commits_dev = response_dev.json()
        commits_main = response_main.json()

        commits_dev_info = [{'sha': commit['sha'],
                             'Name': commit['commit']['committer']['name'],
                             'email': commit['commit']['committer']['email'],
                             'avatar': commit['author']['avatar_url'],
                             'message': commit['commit']['message'],
                             'commit_date': datetime.datetime.strptime(commit['commit']['committer']['date'], '%Y-%m-%dT%H:%M:%SZ')}
                            for commit in commits_dev]

        commits_main_info = [{'sha': commit['sha']}
                             for commit in commits_main]
        
        if commits_dev_info and commits_main_info:       
            for commit_info in commits_dev_info:
                commit_info['commit_date_str'] = commit_info['commit_date'].strftime('%d/%m/%Y %H:%M:%S')
                commit_info['main'] = 'False'
                for commit_main in commits_main_info:
                    if commit_info['sha'] == commit_main['sha']:
                        commit_info['main'] = 'True'
                        break
            context = { 
            'success' : True,
            'commits_dev': commits_dev_info}  
        else:
             context = { 
            'success' : False,
            'message': 'No data'}                             

    except requests.exceptions.HTTPError as err:
        if err.response.status_code == 401:
            context = { 
            'success' : False,
            'message': 'Token đã hết hạn hoặc không hợp lệ.'}
        else:
            context = { 
            'success' : False,
            'message': f"Lỗi khi gửi yêu cầu: {err}"}

    except requests.exceptions.RequestException as e:
        commits_dev_info = []
        commits_main_info = []
        context = { 
            'success' : False,
            'message': e}
        # Xử lý lỗi khi không thể lấy thông tin từ GitHub API
    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    # return render(request, 'github.html', {'commits_dev': commits_dev_info, 'commits_main': commits_main_info})

def Load_Github(request):
    try:
        cookie_system_data     = GetCookie(request, 'cookie_system_data')
        cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
        if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
            # status = check_document(request)
            # if status == True:
                return render(request, 'Ticket_Github.html')
            # else:
            #     return redirect('/dashboard/')
        else:
            return redirect('/')
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
################################################### GITHUB ##################### 

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### API ########################
@csrf_exempt
def api_test(request):
    if request.method == 'POST':
        try:
            context=[]
            data = json.loads(request.body)
            if isinstance(data, list) and len(data) > 0:          
                for item in data:
                    Mail_ID = item.get('Mail_ID')
                    Status = item.get('Status')
                    data1 = item.get('api')['data1']
                    data2 = item.get('api')['data2']

                    data_context = {
                        'Mail_ID': Mail_ID,
                        'Status' : Status,
                        'api'    : {
                            'data1' : data1,
                            'data2' : data2
                        }
                    }
                    context.append(data_context)
                return JsonResponse(context, safe=False)
            else:
                return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def api_load_mail(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            if data:          
                Mail_ID = data.get('Mail_ID')
                Status = data.get('Status')
                if Status:
                    request.session['mail_id'] = Mail_ID
                    access_token = get_authorization_code_API()
                    if access_token:
                        request.session['mail_id'] = {
                            'Mail_ID' : Mail_ID,
                            'access_token' : access_token
                        }
                        response = call_graph_api(request)
                        response_data = response.content.decode('utf-8')  # Giải mã dữ liệu từ bytes sang str
                        response_json = json.loads(response_data)  # Chuyển đổi thành đối tượng JSON

                        if 'success' in response_json and response_json['success'] == True:
                            # return JsonResponse({'success': 'Success'}, status=200)
                            # return JsonResponse({'success': True, 'message': 'Create data Success'}, status=200)
                            return JsonResponse({'success': True, 'message': 'Create data Success'})
                        else:
                            return JsonResponse({'success': response_json['success'],'message': response_json['message']})
                            # return JsonResponse({'Error': response_json['error']}, status=400)
                    else:
                        if request.session.get('mail_id'):
                            del request.session['mail_id']
                        return JsonResponse({'error': 'Token is empty'}, status=400)
            else:
                return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def get_authorization_code_API():
    tenant_id = settings.MICROSOFT_TENANT_ID_EMAIL
    client_id = settings.MICROSOFT_CLIENT_ID_EMAIL
    client_secret = settings.MICROSOFT_CLIENT_SECRET_EMAIL
    scope = 'https://graph.microsoft.com/Mail.ReadWrite https://graph.microsoft.com/Mail.Send'
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'

    # Thông tin người dùng cố định
    username = settings.EMAIL_HOST_USER
    password = settings.EMAIL_HOST_PASSWORD

    token_data = {
        'grant_type': 'password',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': scope,
        'username': username,
        'password': password
    }

    token_response = requests.post(token_url, data=token_data)

    if token_response.status_code == 200:
        access_token = token_response.json().get('access_token')
        return access_token
        # Xử lý access token hoặc trả về kết quả tương ứng
    else:
        # Xử lý lỗi hoặc trả về thông báo tương ứng
        access_token = ''
        return access_token
################################################### API #####################  

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### REAL TIME #####################  
def generate_code_verifier(length=32):
    characters = string.ascii_letters + string.digits + '-._~'
    return ''.join(random.choice(characters) for _ in range(length))

def generate_code_challenge(verifier):
    hashed_verifier = hashlib.sha256(verifier.encode()).digest()
    code_challenge = base64.urlsafe_b64encode(hashed_verifier).rstrip(b'=').decode()
    return code_challenge

def get_user_info(access_token):
    access_token = get_authorization_code_API()
    user_info_url = 'https://graph.microsoft.com/v1.0/users'  # Đổi thành URL tương ứng nếu cần thông tin người dùng khác
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    response = requests.get(user_info_url, headers=headers)

    if response.status_code == 200:
        user_info = response.json()
        user_id = user_info['id']
        return user_id
    else:
        return None  # Xử lý lỗi tại đây

def start_realtime_mail_tracking(request):
    access_token = get_authorization_code_API()
    full_host = request.build_absolute_uri('/') 
    redirect_uri = full_host +'api-load-mail/'
    # userid = '6cc6b729-02aa-42cc-a4e2-1aedadfa4c73'

    # Xác thực để theo dõi sự kiện email mới
    url = 'https://graph.microsoft.com/v1.0/subscriptions'
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    data = {
        "changeType": "created",
        "notificationUrl": redirect_uri,  # Thay bằng URL của view xử lý sự kiện
        "resource": "me/mailFolders('SDP')/messages",
        # "resource": f"users/{userid}/mailFolders('SDP')/messages",
        "filter": {
            "subject": "[SDP]",
            "isRead": False
        },
        "expirationDateTime": "9999-12-31T23:59:59.0000000Z",
        "clientState": "secretClientValue"
    }

    response = requests.post(url, json=data, headers=headers)

    if response.status_code == 201:
        return JsonResponse({"message": "Real-time mail tracking started successfully."})
    else:
        return JsonResponse({"message": "Failed to start real-time mail tracking."})

# View để xử lý thông báo từ Microsoft Graph API
def handle_incoming_email(request):
    if request.method == 'POST':
        data = request.body.decode('utf-8')
        # Xử lý dữ liệu từ Microsoft Graph API khi có sự kiện email mới
        return JsonResponse({"message": "Email event handled successfully."})
    return JsonResponse({"message": "Invalid request."})
################################################### REAL TIME #####################  

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### Login O365, system #####################  
def Login_function(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    if(cookie_microsoft_data):
        response_data = json.loads(cookie_microsoft_data)
        # Lưu dữ liệu vào session
        # a = request.session.get('User_Info').get('UserID')
        # if 'UserInfo' in request.session:
        if 'UserInfo' not in request.session or request.session['UserInfo']['ID_user'] != response_data.get('ID_user'):
            request.session['UserInfo'] = {
                'ID_user': response_data.get('ID_user'),
                'Mail': response_data.get('Mail'),
                'FullName': response_data.get('FullName'),
                'displayName': response_data.get('displayName'),
                'Avatar': response_data.get('Avatar'),
                'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
        }
        return redirect('/dashboard/')
    elif(cookie_system_data):
        response_data = json.loads(cookie_system_data)
        # if 'UserInfo' in request.session:
        if 'UserInfo' not in request.session or request.session['UserInfo']['ID_user'] != response_data.get('ID_user'):
            request.session['UserInfo'] = {
                'ID_user': response_data.get('ID_user'),
                'Mail': response_data.get('Mail'),
                'FullName': response_data.get('FullName'),
                'displayName': response_data.get('displayName'),
                'Avatar': response_data.get('Avatar'),
                'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
        }
        return redirect('/dashboard/')
    else:
        return render(request, 'Login.html')
        # template =  loader.get_template('Login.html')
        # return HttpResponse(template.render())
    
    # return redirect('/')

#lOGIN SYSTEM
@csrf_exempt
def login_system(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('pass')
            remember = request.POST.get('checkbox')
            user = Users.objects.filter(Q(User_Type__lt= 3), Mail = email, Password = password, User_Status = True,).first() # 0 - administrator , 1 - Mod
            # user = Users.objects.filter(Mail = email, Password = password, User_Status = True,).first() # 0 - administrator , 1 - Mod
            if(user):
                 # Lưu dữ liệu vào session
                #  for user in user:
                request.session['UserInfo'] = {
                    'ID_user': user.ID_user,
                    'Mail': user.Mail,
                    'FullName': user.FullName,
                    'displayName': user.displayName,
                    'Avatar': user.Avatar,
                    'Photo': user.displayName[0] +user.FullName[0],
                    'Acc_type': user.User_Type,
                }
                session_data = json.dumps(request.session['UserInfo'])

                # Lưu chuỗi JSON vào cookie
                # Lưu access_token vào cookie (lưu ý: cần kiểm tra tính bảo mật của cookie)
                if(remember == 'true'):
                    # url_redirect = '/dashboard/'
                    # response = HttpResponse('cookie success')
                    # response = SetCookie(response , 'cookie_system_data', session_data, url_redirect)
                    # response = DeleteCookie(response , 'cookie_microsoft_data')
                    # return response
                    return JsonResponse({
                        'success': True,
                        'message': 'Login Successed',
                        'remember': True,
                        'cookie_name': 'cookie_system_data',
                        'cookie_data': session_data,
                        'cookie_day': 90,
                        })
                else:  
                # return HttpResponseRedirect('/danh-sach-yeu-cau/')  
                # return redirect('/dashboard/')                       
                    return JsonResponse({
                        'success': True,
                        'message': 'Login Successed',}) 
            else:
                return JsonResponse({
                'success': False,
                'message': 'Login Failed',})
    except Users.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'User does not exist',
        })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Login Failed: {str(ex)}',
        })
#lOGIN SYSTEM

# Page Login account office 365
def microsoft_login_token(request):
    code = request.GET.get('code')
    full_host = request.build_absolute_uri('/')
    url_host = full_host + 'login/callback/'
    if code is None:
        # Redirect to the Microsoft login page
        url = 'https://login.microsoftonline.com/common/oauth2/v2.0/authorize'
        params = {
            'client_id': '5c17ff26-50a1-4003-bc31-f0545709c2f7',
            'response_type': 'code',
            # 'redirect_uri': 'https://localhost:8000/login/callback/',
            'redirect_uri': url_host,
            'scope': 'https://graph.microsoft.com/.default',
        }
        auth_url = url + '?' + urlencode(params)
        return redirect(auth_url)

    # Exchange the authorization code for an access token and a refresh token
    token_url = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'
    MICROSOFT_CLIENT_SECRET = settings.MICROSOFT_CLIENT_SECRET
    data = {
        'grant_type': 'authorization_code',
        'code': code,
        # 'redirect_uri': 'https://localhost:8000/login/callback/',
        'redirect_uri': url_host,
        'client_id': '5c17ff26-50a1-4003-bc31-f0545709c2f7',
        # 'client_secret': 'EeJ8Q~ip-6TA~p1C7Y9t24l81qig0lFv1t5CPdwO', 
        'client_secret': MICROSOFT_CLIENT_SECRET, 
    }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
    }
    response = requests.post(token_url, data=data, headers=headers)
    response_data = response.json()

    # Save the access token and refresh token in the session
    access_token = response_data['access_token']
    request.session['access_token'] = access_token
    if 'refresh_token' in response_data:
        refresh_token = response_data['refresh_token']
        request.session['refresh_token'] = refresh_token
    else:
        refresh_token = None  # or handle the error in some other way


    # Use the access token to get the user's profile
    user_url = 'https://graph.microsoft.com/v1.0/me'
    headers = {
        'Authorization': 'Bearer ' + access_token,
        'Accept': 'application/json',
    }
    response = requests.get(user_url, headers=headers)
    response_data = response.json()
    #get avatar
    photo_data = response_data.get('photo', {})
    photo_url = ''
    if 'small' in   photo_data:
        photo_url = photo_data['small']
    elif 'large' in photo_data:
            photo_url = photo_data['large']

    #check user exit
    email = response_data.get('mail')
    exit_IDuser = Users.objects.filter(Mail = email)
    if(exit_IDuser):
        for user in exit_IDuser:
            if user.User_Status:
                # Cập nhật các thuộc tính của User
                user.Mail = email
                user.FullName = response_data.get('surname') + " " + response_data.get('givenName')
                user.displayName = response_data.get('displayName')
                if(response_data.get('jobTitle')):
                    user.Jobtitle = response_data.get('jobTitle')
                if(response_data.get('mobilePhone')):
                    user.Phone = response_data.get('mobilePhone')
                user.Avatar = photo_url
                user.save()
                comp = Company.objects.get(Company_ID = user.Company_ID)
                if 'No Data' in comp.Company_Name:
                    url_redirect = '/cap-nhat-thong-tin-ca-nhan/'
                else:
                    url_redirect = '/dashboard/'            
            else:
                return redirect('/page-404/')
    else:
        #check data ID User -get new ID
        UserID = Check_IDUser()
        comID = Company.objects.get(Company_Name = 'No Data')
        # Insert dữ liệu vào cơ sở dữ liệu
        user = Users(
            ID_user     =UserID,
            Mail        =email,
            Password    ='',
            FullName    =response_data.get('surname') + " " + response_data.get('givenName'),
            Company_ID  =comID.Company_ID,
            displayName =response_data.get('displayName'),
            Birthday    = response_data.get('birthday') if response_data.get('birthday') else datetime.date(1990, 1, 1),  # Chưa có thông tin về ngày sinh trong Microsoft Graph API
            Acc_Type    ='Microsoft',
            Address     ='',
            Jobtitle    =response_data.get('jobTitle') if response_data.get('jobTitle') else "",
            Phone       =response_data.get('mobilePhone') if response_data.get('jobTitle') else 0,
            Avatar      =photo_url,
            User_Type   = 2, # 2 - EndUser
            Date_Create   = datetime.datetime.now().date(),
            Time_Create   = datetime.datetime.now().time(),
            User_Status =True  #User_Status là True khi mới đăng ký        
        )
        user.save()
        login_role(UserID)
        url_redirect = '/cap-nhat-thong-tin-ca-nhan/'


    # Lưu dữ liệu vào session
    request.session['UserInfo'] = {
        'ID_user': user.ID_user,
        'Mail': user.Mail,
        'FullName': user.FullName,
        'displayName': user.displayName,
        'Avatar': user.Avatar,
        'Photo': user.displayName[0] +user.FullName[0],
        'Acc_type': user.User_Type,
    }

    # Save the user's profile in a cookie
   # Chuyển đổi session data thành chuỗi JSON
    session_data = json.dumps(request.session['UserInfo'])

    # Lưu chuỗi JSON vào cookie
     # Lưu access_token vào cookie (lưu ý: cần kiểm tra tính bảo mật của cookie)
    # url_redirect = '/dashboard/'
    response = SetCookie(response , 'cookie_microsoft_data', session_data, url_redirect)
    # response = DeleteCookie(response , 'cookie_google_data')
    # response = DeleteCookie(response , 'cookie_facebook_data')  
    response = DeleteCookie(response , 'cookie_system_data')  
    return response
# Page Login account office 365

def login_role(ID_user):
    user_create = Users.objects.get(ID_user = 'U000000001')
    assign = Users.objects.get(ID_user = ID_user)
    role_include = [
        'ZTC_View',
        'ZTC_Add',
        'ZDC_View',
    ]
    # Tạo điều kiện Q động cho mỗi giá trị trong role_mod_exclude
    conditions = [Q(Role_Name__icontains=value) for value in role_include]
    # Kết hợp các điều kiện bằng toán tử OR
    combined_condition = reduce(lambda x, y: x | y, conditions)
    # Loại bỏ các bản ghi thỏa mãn điều kiện
    list_role = Role_Single.objects.filter(combined_condition)
    b = list(list_role)
    for role in list_role:
        auth_role = Authorization_User(
            ID_user = assign,
            Role_ID = role,
            Authorization_From = datetime.datetime.now().date(),
            Authorization_To = datetime.date(9999, 12, 31),
            Authorization_CreateID = user_create.ID_user,
            Authorization_CreateBy = user_create.FullName,
            Authorization_Date = datetime.datetime.now().date(),
            Authorization_Time = datetime.datetime.now().time(),
            Authorization_Status = 1
        )
        auth_role.save()

#CHEKC MAIL OFFICE 365
def check_emails(request):
    MICROSOFT_CLIENT_SECRET = settings.MICROSOFT_CLIENT_SECRET
    client_id = '5c17ff26-50a1-4003-bc31-f0545709c2f7'
    client_secret = MICROSOFT_CLIENT_SECRET
    # scope = 'https://graph.microsoft.com/.default'
    scope = 'https://graph.microsoft.com/Mail.Read'
    tenant_id = 'c43d3f81-f57a-48cc-8b07-74b39935d876'

    # Đăng nhập để lấy access token
    token_url = 'https://login.microsoftonline.com/'+tenant_id+'/oauth2/v2.0/token'
    token_data = {
        'grant_type': 'authorization_code',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': scope,
    }
    token_response = requests.post(token_url, data=token_data)
    access_token = token_response.json().get('access_token')

    while True:
        # Gọi Microsoft Graph API để lấy danh sách email từ hộp thư
        response = requests.get(
            'https://graph.microsoft.com/v1.0/me/messages',
            headers={'Authorization': f'Bearer {access_token}'}
        )
        emails = response.json().get('value', [])
        for email in emails:
            if email['subject'] == 'Ticket Request':
                # Lấy email và pass từ email được trigger
                email_address = email['from']
                email_password = email['body']

        # Xử lý dữ liệu từ emails để tạo ticket
        # Ví dụ: lặp qua danh sách emails và tạo ticket

        # Đợi một khoảng thời gian trước khi kiểm tra lại
        time.sleep(5)  # Đợi 5 giây trước khi kiểm tra lại

def get_authorization_code(request):
    full_host = request.build_absolute_uri('/') 
    # Redirect người dùng đến URL xác thực để đăng nhập
    redirect_uri = full_host +'load-mail-unread/'
    tenant_id = settings.MICROSOFT_TENANT_ID_EMAIL
    client_id = settings.MICROSOFT_CLIENT_ID_EMAIL
    scope = 'https://graph.microsoft.com/Mail.ReadWrite+Mail.Send'
    # scope = 'https://graph.microsoft.com/Mail.Read'
    authorization_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/authorize?client_id={client_id}&response_type=code&redirect_uri={redirect_uri}&scope={scope}'
    return redirect(authorization_url)

def get_access_token(authorization_code,full_host):
    redirect_uri = full_host +'load-mail-unread/'
    tenant_id = settings.MICROSOFT_TENANT_ID_EMAIL
    client_id = settings.MICROSOFT_CLIENT_ID_EMAIL
    client_secret = settings.MICROSOFT_CLIENT_SECRET_EMAIL
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    token_data = {
        # 'grant_type': 'client_credentials', #Application Permissions  (Quyền Ứng dụng)
        'grant_type': 'authorization_code', #Delegated Permissions (Quyền Ủy quyền)
        'client_id': client_id,
        'client_secret': client_secret,
        'code': authorization_code,
        'redirect_uri': redirect_uri,
        'scope': 'https://graph.microsoft.com/Mail.ReadWrite https://graph.microsoft.com/Mail.Send',
    }
    token_response = requests.post(token_url, data=token_data)
    access_token = token_response.json().get('access_token')
    return access_token

# Sử dụng mã truy cập để gọi Microsoft Graph API
def call_graph_api(request):    
    mailid = ''
    emails = []
    authorization_code = request.GET.get('code')
    full_host = request.build_absolute_uri('/') 

    session_data = request.session.get('mail_id', {})
    if session_data:
        mailid = session_data.get('Mail_ID')
        access_token = session_data.get('access_token')
    else:
        # mailid = request.session.get('mail_id')  
        access_token = get_access_token(authorization_code,full_host)
    
    if access_token == None :
        return JsonResponse({'error': 'token is empty'}, status=400)
    
    # Tính thời gian ngày hôm nay và ngày hôm trước
    today = datetime.datetime.utcnow().date()
    yesterday = today - timedelta(days=1)

    # Đường dẫn thư mục trên máy chủ để lưu các tệp đính kèm
    # attachment_dir = '/static/Asset/Attachment-Upload/'
    attachment_dir = os.path.join(settings.BASE_DIR , 'static', 'Asset', 'Attachment-Upload')
    attachment_dir_img = os.path.join(settings.BASE_DIR , 'static', 'Asset', 'Attachment-Image')
    if(mailid):
        response = requests.get(
            # f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and subject eq \'{prefix}\' and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
            f'https://graph.microsoft.com/v1.0/me/messages/{mailid}',
            headers={'Authorization': f'Bearer {access_token}'}
        )
        if response.status_code == 200:
         data_email = response.json()
         emails.append(data_email)
    else:
        prefix = '[SDP]'
        folder_mail = 'Inbox'
        folder_mail_sub = 'SDP'
        folderID = folder_email(access_token,folder_mail,folder_mail_sub)
        if folderID:
            response = requests.get(
                # f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and subject eq \'{prefix}\' and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
                # f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and contains(subject, \'{prefix}\') and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
                f'https://graph.microsoft.com/v1.0/me/mailFolders/{folderID}/messages?$filter=isRead eq false and contains(subject, \'{prefix}\') and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
                headers={'Authorization': f'Bearer {access_token}'}
            )
            emails = response.json().get('value', [])   
    if emails:
        folder_mail_sub_move = 'DONE'
        folderID_move = folder_email(access_token,folder_mail,folder_mail_sub_move)
        for email in emails:
            Mail_status = ''
            email_from = email['from']['emailAddress']['address']  # Địa chỉ email người gửi
            user = Users.objects.filter(Mail = email_from, User_Status = True).first()
            if user:
                email_id = email['id']  # ID của email
                email_subject = email['subject']  # Tiêu đề email          
                # Lấy danh sách người nhận
                recipients = email['toRecipients']
                recipient_addresses = [recipient['emailAddress']['address'] for recipient in recipients]

                # Lấy danh sách người được CC
                cc_recipients = email.get('ccRecipients', [])
                cc_addresses = [cc['emailAddress']['address'] for cc in cc_recipients]

                email_body = email['body']['content']  # Nội dung email

                #function create Ticket by mail 
                email_group_ticket = check_group_ticket(email_subject,email_body)           
                email_type_ticket = 1 #Type Hỗ Trợ
                email_company_ticket = user.Company_ID  
                
                # if isinstance(email['conversationId'], dict):
                    # Lấy email ID của email gốc                   
                    # conversation_Id = email['conversationId']['id'] 
                conversation_Id = email['conversationId']
                Context = Ticket_Mapping.objects.filter(Email_ID = conversation_Id).first()
                if Context:
                    Mail_status = 'Reply'
                else:
                    Context = Create_Ticket_Mail(request,email_id, email_subject, email_body,email_type_ticket,email_company_ticket,email_group_ticket,email_from,recipients)
                # else:
                #     Context = Create_Ticket_Mail(request,email_id, email_subject, email_body,email_type_ticket,email_company_ticket,email_group_ticket,email_from,recipients)

                if Context:
                    email_attachments = email['hasAttachments']  # Danh sách đính kèm
                    # Xử lý danh sách đính kèm
                    # if email_attachments:
                    # attachment_names = []
                    attachment_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/attachments'
                    attachment_response = requests.get(
                        attachment_url,
                        headers={'Authorization': f'Bearer {access_token}'}
                    )               
                    attachments = attachment_response.json().get('value', [])
                    if attachments:
                        #image
                        soup = BeautifulSoup(email_body, 'html.parser')
                        img_tags = soup.find_all('img')
                        for img_tag in img_tags:
                            src = img_tag.get('src', '')
                            if src.startswith('cid:'):
                                content_id = src[4:]

                                for attachment in attachments:
                                    if attachment['contentType'].startswith('image/') and attachment['contentId'] == content_id:
                                        img_data = attachment['contentBytes']

                                        img_filename = f'{content_id}.png'  # Tên tệp hình ảnh
                                        img_path = os.path.join(attachment_dir_img, img_filename)  # Đường dẫn tới tệp hình ảnh trên máy chủ

                                        with open(img_path, 'wb') as img_file:
                                            img_file.write(base64.b64decode(img_data))
                                        
                                        with open(img_path, 'rb') as img_file:
                                            img_base64 = base64.b64encode(img_file.read()).decode('utf-8')
                                            # img_src = f'data:{attachment["contentType"]};base64,{img_base64}'
                                            img_src = f'/static/Asset/Attachment-Image/{content_id}.png'
                                            img_tag['src'] = img_src
                                        break

                        if Mail_status: #emailID original is exit
                            body_email = str(soup)
                            ticketid = Context.Ticket_ID.Ticket_ID
                            commentID = Create_Comment_Mail(ticketid,user,body_email)
                            if commentID:
                                Update_Unread_Comment(commentID,ticketid,user.ID_user)
                            # if commentID:
                            #     Create_Mapping_Mail(Context,original_email_id, commentID)
                        else: #emailID original is not exit
                            ticket = Ticket.objects.get(Ticket_ID = Context['Ticket_ID']) 
                            if ticket:
                                ticketid = ticket.Ticket_ID
                                ticket.Ticket_Desc = str(soup)    
                                ticket.save() 
                                Create_Mapping_Mail(ticket.Ticket_ID,conversation_Id,'0')                                  
                        #attachment file
                        for attachment in attachments:
                            if not attachment['contentType'].startswith('image/'):
                                ticketID = ticketid
                                # ticketID = Context['Ticket_ID']
                                attachment_id   = attachment['id']
                                attachment_name = attachment['name']
                                # attachment_type = attachment['contentType']
                                # attachment_size = attachment['size']
                                # attachment_names.append(attachment_name)

                                attachment_item = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/attachments/{attachment_id}'
                                attachment_item_response = requests.get(
                                    attachment_item,
                                    headers={'Authorization': f'Bearer {access_token}'}
                                )
                                attachment_data = attachment_item_response.content
                                # Xây dựng đường dẫn lưu tệp trên máy chủ
                                current_datetime = datetime.datetime.now()
                                numeric_date = current_datetime.strftime('%d%m%Y')
                                numeric_time = current_datetime.strftime('%H%M')
                                attachment_name_full = str(ticketID) + '_' + numeric_date + '_' + numeric_time + '_'+ user.ID_user +'_'+ attachment_name
                                attachment_path = os.path.join(attachment_dir, attachment_name_full)

                                # Lưu tệp đính kèm vào máy chủ
                                with open(attachment_path, 'wb') as attachment_file:
                                    attachment_file.write(attachment_data)
                                
                                #create attachment data
                                create_attachment_mail(ticketID,attachment_name_full,user.ID_user)

                    # reply_email(request, access_token, email_id, email_from, Context['Ticket_ID'], Context['Slug_Title'])             

                    # else:
                    #     #gửi mail thống báo lỗi không thành công.
                    #     mail = "Gửi mail thông báo"
                    if(mailid):
                        del request.session['mail_id']
                    if Mail_status == '':
                        reply_email(request, access_token, email_id, email_from, Context['Ticket_ID'], Context['Slug_Title'])
                        send_email_new_ticket(request,email_id, email_from,Context['Ticket_ID'],Context['Slug_Title'],Context['Email_Assign'],email_body)
                    readTrue = read_email(access_token,email_id)  
                    if readTrue is not True:
                        folder_mail_move = 'DONE'
                        folderID_move = folder_email(access_token,folder_mail,folder_mail_move)
                        if folderID_move:
                            move_email(access_token,email_id,folderID_move)
            else:
                if(mailid):
                    del request.session['mail_id']
                email_id = email['id']  # ID của email
                subject = "[SDP] Thông báo lỗi - User Chưa có trên hệ thống"    
                reply_email_NoExitUser(request, access_token, email_id, email_from, subject)           
        return JsonResponse({'success': True,'message': 'Create data Success'}, status=200)
    else:
        # return JsonResponse({'error': 'No email data'}, status=400)
        return JsonResponse({'success': False ,'message': 'No email data'}, status=400)
    # time.sleep(5)
    return JsonResponse({'success': 'Success'}, status=200)
    # return redirect('/get-code/')

def check_group_ticket(title_ticket,email_body):
    try:
        ticket_type = TGroup.objects.filter(~Q(TGroup_Name = 'Khác') or ~Q(TGroup_Name = 'Other')).values('TGroup_ID','TGroup_Name')
        if ticket_type:
            for t in ticket_type:
                if t['TGroup_Name'].lower() in title_ticket.lower():
                    return t['TGroup_ID']
                if t['TGroup_Name'].lower() in email_body.lower():
                    return t['TGroup_ID']
        ticket_other = TGroup.objects.filter(Q(TGroup_Name = 'Khác') or Q(TGroup_Name = 'Other')).values('TGroup_ID','TGroup_Name')
        return ticket_other[0]['TGroup_ID']
    except Exception as ex:
        return ''

def read_email(access_token,email_id):
    # Cập nhật trạng thái của email sang đã đọc
    email_read_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}'
    update_data = {
        'isRead': True
    }
    update_headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    update_response = requests.patch(email_read_url, json=update_data, headers=update_headers)

    if update_response.status_code == 204:
        return True
    else:
        return False    

def move_email(access_token,email_id,folderID):
    # Di chuyển email dựa trên ID
    move_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/move'
    move_data = {'destinationId': folderID}  # Thay 'Done' bằng ID thư mục Done thực tế
    move_headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    move_response = requests.post(move_url, json=move_data, headers=move_headers)

    if move_response.status_code == 201:
        return True
        # print(f'Successfully moved message {email_id_to_move} to Done folder')
    else:
        return False
        # print(f'Failed to move message {email_id_to_move}: {move_response.text}')

def folder_email(access_token,folder_name,folder_mail_sub):
    folder_url = 'https://graph.microsoft.com/v1.0/me/mailFolders'
    response = requests.get(folder_url, headers={'Authorization': f'Bearer {access_token}'})
    folders = response.json().get('value', [])

    sdp_folder_id = None
    for folder in folders:
        if folder.get('displayName') == folder_name:
            sdp_folder_id = folder.get('id')
            break

    if sdp_folder_id is None:
        return ''
    else:
        if folder_mail_sub:
            sub_folder_id = get_inbox_subfolders(access_token, sdp_folder_id,folder_mail_sub)
            return sub_folder_id
        else:
            return sdp_folder_id

def get_inbox_subfolders(access_token, inbox_folder_id,folder_mail_sub):
    url = f'https://graph.microsoft.com/v1.0/me/mailFolders/{inbox_folder_id}/childFolders'
    headers = {'Authorization': f'Bearer {access_token}'}

    response = requests.get(url, headers=headers)
    folders_sub = response.json().get('value', [])
    sub_folder_id = None
    if folders_sub:
        for folder in folders_sub:
            if folder.get('displayName') == folder_mail_sub:
                sub_folder_id = folder.get('id')
                break
        return sub_folder_id
    else:
        return ''
#CHEKC MAIL OFFICE 365

#FUNCTION CREATE TICKET 
def Create_Ticket_Mail(request,email_id, email_subject, email_body,email_type_ticket,email_company_ticket,email_group_ticket,email_from,recipients):
    try:
            if recipients:
                recipient_addresses = [recipient['emailAddress']['address'] for recipient in recipients]
                # email = Assign_User.objects.filter(TGroup_ID = email_group_ticket, ID_user__Mail__in = recipient_addresses).values('Assign_User_ID','Assign_User_Name')
                email = Assign_User.objects.filter(TGroup_ID = email_group_ticket, ID_user__Mail__in = recipient_addresses)
                if email:
                    sorted_emails = sorted(email, key=lambda x: recipient_addresses.index(x.ID_user.Mail))
                    mail = sorted_emails[0]
                    assign_User = mail.ID_user.ID_user
                    assign_Name = mail.ID_user.FullName
                else: 
                    suggest_data = suggest_ticket(request,email_subject, email_group_ticket)
                    assign_User = suggest_data['ticket_assigned_user_id']
                    assign_Name = suggest_data['ticket_assigned_username']
            else:    
                suggest_data = suggest_ticket(request,email_subject, email_group_ticket)
                assign_User = suggest_data['ticket_assigned_user_id']
                assign_Name = suggest_data['ticket_assigned_username']
            
            company = Company.objects.get( Company_ID = email_company_ticket)
            group = TGroup.objects.get( TGroup_ID = email_group_ticket)
            user = Users.objects.get( Mail = email_from)
            if user:
                slug_title =  convert_title_to_slug(email_subject)
                #create Ticket
                ticket = Ticket.objects.create(
                    Ticket_Title = email_subject, 
                    Ticket_Desc  = email_body, 
                    Ticket_Type  = email_type_ticket, 
                    Company_ID   = company,
                    TGroup_ID    = group, 
                    ID_User      = user,
                    Ticket_User_Name = user.FullName,
                    Ticket_Date  = datetime.datetime.now().date(),
                    Ticket_Time  = datetime.datetime.now().time(),
                    Ticket_User_Asign = assign_User,
                    Ticket_Name_Asign = assign_Name,
                    Ticket_Status = 1,
                )
                ticket.save()
                TicketID = ticket.Ticket_ID
                email_assign = Users.objects.get(ID_user = assign_User)
                context = {
                    'Ticket_ID': TicketID,
                    'Slug_Title': slug_title,
                    'Email_Assign': email_assign
                }
                return context
    except Assign_User.DoesNotExist:
                #Sent Email thông báo chưa có tài khoản
                pass
    except Exception as ex:
        return ''
        # return JsonResponse({
        #     'success': False,
        #     'message': f'Lỗi: {str(ex)}',
        # })
#FUNCTION CREATE TICKET 

#FUNCTION CREATE COMMENT 
def Create_Comment_Mail(TicketID,User, email_body):
    try:
        if User:
            ticket = Ticket.objects.get(Ticket_ID = TicketID)
            #create Comment
            comment = Comment.objects.create(
                Ticket_ID = ticket, 
                ID_user  = User, 
                Comment_Desc  = email_body, 
                Comment_level   = 0,
                Comment_User_Name    = User.FullName, 
                Comment_Date  = datetime.datetime.now().date(),
                Comment_Time  = datetime.datetime.now().time(),
                Comment_Status = 1,
            )
            comment.save()
            CommentID = comment.Comment_ID
            return CommentID
    except Exception as ex:
        return ''
        # return JsonResponse({
        #     'success': False,
        #     'message': f'Lỗi: {str(ex)}',
        # })
#FUNCTION CREATE COMMENT 

#FUNCTION CREATE MAPPING 
def Create_Mapping_Mail(TicketID,EmailID, CommentID):
    try:
        if TicketID:
            ticketid = Ticket.objects.get(Ticket_ID = TicketID)
            mapping = Ticket_Mapping.objects.create(
                Ticket_ID = ticketid, 
                Email_ID  = EmailID, 
                Comment_ID  = CommentID, 
                Ticket_Mapping_Status = 1,
            )
            mapping.save()
            mappingID = mapping.Ticket_Mapping_ID
            data = {
                'Mapping_ID': mappingID,
            }
            return data
    except Exception as ex:
        return ''
#FUNCTION CREATE MAPPING 

def create_attachment_mail(TicketID, AttachmentUrl,UserID):
    try:
        user = Users.objects.get(ID_user = UserID)
        ticket = Ticket.objects.get(Ticket_ID = TicketID)
        if ticket and user:
            attach = Attachment.objects.create(
                Ticket_ID =  ticket,
                ID_user   = user,
                Attachment_Url = AttachmentUrl,
                Attachment_User_Name = user.FullName,
                Attachment_Date = datetime.datetime.now().date(),
                Attachment_Time = datetime.datetime.now().time(),
                Attachment_Status = 1
            )
            attach.save()
            return True
    except Exception as ex:
        return False

#FUNCTION SEND EMAIL
def send_email(Email,TicketID,Ticket_Title,slug_title,Ticket_User_Name,Ticket_Date,Ticket_Time,Ticket_Name_Asign, request):
    try:
        # full_host = request.get_full_host()   
        full_host = request.build_absolute_uri('/')   
        subject = str(TicketID) + ' - ' + Ticket_Title
        data = {
        'TicketID': TicketID,
        'AssignName': Ticket_Name_Asign,
        'Title': Ticket_Title,
        'CreateDate': Ticket_Date + ' ' + Ticket_Time,
        'CreateName': Ticket_User_Name,
        'Link': full_host+'chi-tiet-yeu-cau/'+str(TicketID)+'/'+slug_title+'/',
        }
        message = render_to_string('Emaiil_Ticket.html', context=data)
        
        # message = 'This is a test email.'
        # from_email = 'giang.llt@bamboocap.com.vn'
        from_email = settings.MICROSOFT_EMAIL
        recipient_list = [Email]
        # cc_list = ['cc@example.com']
        send_mail(
            subject=subject,
            message=message,
            from_email=from_email,
            recipient_list=recipient_list,
            # cc=cc_list,
            fail_silently=False,  # Thông báo lỗi nếu có vấn đề khi gửi email
            html_message=message,  # Sử dụng nội dung HTML
        )
        # return HttpResponse('Email sent successfully')
        return True
    except Exception as ex:
        return False
    
def send_email_new_ticket(request, email_id, email_from, TicketID, slug_title,email_assign,email_body):
    try:
        ticket = Ticket.objects.get(Ticket_ID = TicketID)
        if ticket:
            Ticket_Title = ticket.Ticket_Title
            # Ticket_Desc = ticket.Ticket_Desc
            Ticket_Desc = email_body
            Ticket_Name_Asign = ticket.Ticket_Name_Asign
            Ticket_Date = ticket.Ticket_Date.strftime('%d/%m/%Y')
            Ticket_Time = ticket.Ticket_Time.strftime('%H:%M')
            Ticket_User_Name= ticket.Ticket_User_Name 
        full_host = request.build_absolute_uri('/')   
        subject = str(TicketID) + ' - ' + Ticket_Title
        data = {
        'TicketID': TicketID,
        'TicketDesc': Ticket_Desc,
        'AssignName': Ticket_Name_Asign,
        'Title': Ticket_Title,
        'CreateDate': Ticket_Date + ' ' + Ticket_Time,
        'CreateName': Ticket_User_Name,
        'Link': full_host+'chi-tiet-yeu-cau/'+str(TicketID)+'/'+slug_title+'/',
        }
        # message = render_to_string('Emaiil_Ticket_Rep.html', context=data)
        
        # message = 'This is a test email.'
        # from_email = 'giang.llt@bamboocap.com.vn'
        from_email = settings.MICROSOFT_EMAIL
        # recipient_list = [email_from]
        # # cc_list = ['cc@example.com']
        # send_mail(
        #     subject=subject,
        #     message=message,
        #     from_email=from_email,
        #     recipient_list=recipient_list,
        #     # cc=cc_list,
        #     fail_silently=False,  # Thông báo lỗi nếu có vấn đề khi gửi email
        #     html_message=message,  # Sử dụng nội dung HTML
        # )

        #Assign Email
        recipient_list_assign = [email_assign.Mail]
        message_asign = render_to_string('Email_Ticket.html', context=data)
        # message_asign = render_to_string('Email_Ticket_Custom.html', context=data)
        send_mail(
            subject=subject,
            message=message_asign,
            from_email=from_email,
            recipient_list=recipient_list_assign,
            # cc=cc_list,
            fail_silently=False,  # Thông báo lỗi nếu có vấn đề khi gửi email
            html_message=message_asign,  # Sử dụng nội dung HTML
        )
        # return HttpResponse('Email sent successfully')
        return True
    except Exception as ex:
        return False

def reply_email(request, access_token, email_id, email_from, TicketID, slug_title):
    try:
        ticket = Ticket.objects.get(Ticket_ID = TicketID)
        if ticket:
            Ticket_Title = ticket.Ticket_Title
            Ticket_Name_Asign = ticket.Ticket_Name_Asign
            Ticket_Date = ticket.Ticket_Date
            Ticket_Time = ticket.Ticket_Time
            Ticket_User_Name= ticket.Ticket_User_Name
            # for e in Email:
            #     list_mail.append(e['emailAddress']['address'])

        full_host = request.build_absolute_uri('/')   
        subject = str(TicketID) + ' - ' + Ticket_Title
        data = {
            'TicketID': TicketID,
            'AssignName': Ticket_Name_Asign,
            'Title': Ticket_Title,
            'CreateDate': Ticket_Date.strftime('%d/%m/%Y') + ' ' + Ticket_Time.strftime('%H:%M'),
            'CreateName': Ticket_User_Name,
            'Link': full_host + 'chi-tiet-yeu-cau/' + str(TicketID) + '/' + slug_title + '/',
        }
        message = render_to_string('Emaiil_Ticket_Rep.html', context=data)

        # URL API của Microsoft Graph
        reply_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/createReply'

        # Dữ liệu để tạo nội dung reply
        reply_data = {
            "message": {
                "subject":' Re: ' + subject,
                "body": {
                    "contentType": "html",
                    "content": message
                },
                "toRecipients": [
                    {
                        "emailAddress": {
                            "address": email_from
                        }
                    }
                ]
            }
        }

        # Gửi yêu cầu POST để tạo reply email
        response = requests.post(
            reply_url,
            headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
            json=reply_data
        )
        # Xử lý phản hồi từ Microsoft Graph API
        if response.status_code == 201:
            # Sau khi tạo reply thành công, gửi email bằng cách cập nhật threadId và gửi yêu cầu POST tới /send
            response_data = response.json()
            thread_id = response_data.get('id')
            send_url = f'https://graph.microsoft.com/v1.0/me/messages/{thread_id}/send'

            send_response = requests.post(
                send_url,
                headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
            )

            if send_response.status_code == 202:
                return True
            else:
                print("Failed to send email:", send_response.content)
                return False
        else:
            print("Failed to create reply email:", response.content)
            return False

    except Exception as ex:
        return False

def reply_email_NoExitUser(request, access_token, email_id, email_from, subject):
    try:
        message = render_to_string('Email_NoExit_User.html')
        # URL API của Microsoft Graph
        reply_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/createReply'

        # Dữ liệu để tạo nội dung reply
        reply_data = {
            "message": {
                "subject":' Re: ' + subject,
                "body": {
                    "contentType": "html",
                    "content": message
                },
                "toRecipients": [
                    {
                        "emailAddress": {
                            "address": email_from
                        }
                    }
                ]
            }
        }

        # Gửi yêu cầu POST để tạo reply email
        response = requests.post(
            reply_url,
            headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
            json=reply_data
        )
        # Xử lý phản hồi từ Microsoft Graph API
        if response.status_code == 201:
            # Sau khi tạo reply thành công, gửi email bằng cách cập nhật threadId và gửi yêu cầu POST tới /send
            response_data = response.json()
            thread_id = response_data.get('id')
            send_url = f'https://graph.microsoft.com/v1.0/me/messages/{thread_id}/send'

            send_response = requests.post(
                send_url,
                headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
            )

            if send_response.status_code == 202:
                return True
            else:
                print("Failed to send email:", send_response.content)
                return False
        else:
            print("Failed to create reply email:", response.content)
            return False

    except Exception as ex:
        return False
#FUNCTION SEND EMAIL

#LOGOUT 
@csrf_exempt
def logout(request):
    # Kiểm tra xem cookie 'cookie_microsoft_data' có tồn tại không trước khi xóa nó
    if 'cookie_microsoft_data' in request.COOKIES:
        response = HttpResponseRedirect('/')  # Chuyển hướng đến trang home (thay đổi tên view tương ứng)
        response.delete_cookie('cookie_microsoft_data')  # Xóa cookie 'cookie_microsoft_data'
        request.session.flush()  # Xóa tất cả dữ liệu trong session
        return response
    elif 'cookie_system_data' in request.COOKIES:
        response = HttpResponseRedirect('/')  # Chuyển hướng đến trang home (thay đổi tên view tương ứng)
        response.delete_cookie('cookie_system_data')  # Xóa cookie 'cookie_microsoft_data'
        request.session.flush()  # Xóa tất cả dữ liệu trong session
        return response
    else:
        # return HttpResponse("Không tìm thấy cookie để xóa.")
        return HttpResponseRedirect('/')
#LOGOUT 

#FORGOT PASSWORD
def forgot_pass(request):
    # Load template in folder templates
    template = loader.get_template('Login_Reset_pass.html')
    return HttpResponse(template.render(request=request))

@csrf_exempt
def Reset_Pass(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            user = Users.objects.filter(Mail = email)
            if user:
                NewPass = generate_random_string()
                user.Password = NewPass
                user.save()                               
                send_email_reset_pass(user, NewPass)
                return JsonResponse({
                    'success': True,
                    'message': 'Cập nhật mật khẩu thành công.'
                }) 
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Email Không tồn tại trên hệ thống.'
                })          
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FORGOT PASSWORD

################################################### Login O365, system #####################  

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### FUNCTION , CHECK DATA, OTHER ########### 
#CHECK AND GET DATA USERID
def Check_IDUser():
     User_data =  Users.objects.using('default').all()
     if(User_data):
         count = User_data.count()
         if count > 0:
             last_user = User_data[count - 1]
             id = last_user.ID_user
             UserID = 'U' + str(int(id[1:]) + 1).zfill(9)
         else:
             UserID = "U000000001"
     else:
            UserID = "U000000001"
     return UserID
#CHECK AND GET DATA USERID

#CHECK USER EXIT
def check_User_exit(email):
    exit_IDuser = Users.objects.filter(Mail = email)
    if(exit_IDuser):
        return exit_IDuser
    else:
        return None
#CHECK USER EXIT

#FUCNTION SET SESTION USER
def set_session_user(request, data_user):
    request.session['UserInfo'] = {
         'ID_user': data_user['ID_user'],
         'Mail': data_user['Mail'],
         'FullName': data_user['FullName'],
         'displayName': data_user['displayName'],
         'Avatar': data_user['Avatar'],
         'Photo': data_user['displayName'][0] + data_user['FullName'][0],
    }
#FUCNTION SET SESTION USER

#FUNCTION SET , GET , DELETE COOKIES
def SetCookie(response , name_data, cookie_data, url_redirect):
    age =  90 * 24 * 60 * 60 #90 day, cookie lifes
    response = HttpResponseRedirect(url_redirect)
    response.set_cookie(name_data,cookie_data,age)
    return response

def GetCookie(request, name_data):
    response = request.COOKIES.get(name_data)
    return response

def DeleteCookie(response, name_data):
    response.delete_cookie(name_data)
    return response
#FUNCTION SET , GET , DELETE COOKIES

def date_handler(obj):
    if isinstance(obj, datetime.date):
        return obj.strftime('%Y-%m-%d')
    return None

def page_404(request):
    return render(request, 'Ticket_404.html')
################################################### FUNCTION , CHECK DATA, OTHER ########### 
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### AUTHORIZE MENU MASTER PAGE ###########
#LOAD DATA ROLE MENU
@csrf_exempt
def role_menu(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            IDuser = userinfo['ID_user']
            AccType = userinfo['Acc_type']
            list_group_data = []
            list_menu_data = []
            if AccType == 0:
                menus =  Menu.objects.filter(Menu_Status = True).order_by('Menu_Level')
                for menu in menus:
                    menu_data = {
                        'Menu_ID' : menu.Menu_ID,
                        'Menu_Name' : menu.Menu_Name,
                        'Menu_Icon' : menu.Menu_Icon,
                        'Menu_Adress' : menu.Menu_Adress,
                    }
                    list_menu_data.append(menu_data)

                list_group = Role_Group.objects.filter(Role_Group_Status = True).order_by('Role_Group_ID','Menu_ID')
                for gr in list_group:
                    group_data = {
                        'Menu_ID' : gr.Menu_ID.Menu_ID,
                        'Role_Group_ID' : gr.Role_Group_ID,
                        'Role_Group_Name' : gr.Role_Group_Name,
                        'Role_Group_Address' : gr.Role_Group_Address,
                    }
                    list_group_data.append(group_data)
            else:
                current = datetime.datetime.now().date()
                auth = Authorization_User.objects.filter(
                    ID_user = IDuser,
                    Authorization_From__lte=current,
                    Authorization_To__gte=current,
                    Authorization_Status =  True
                )
                distinct_auth = auth.values_list('Role_ID', flat=True).distinct()
                group = Role_Single.objects.filter(Role_ID__in = distinct_auth).values('Role_Group_ID')
                distinct_group = group.values_list('Role_Group_ID', flat=True).distinct()
                list_group = Role_Group.objects.filter(Role_Group_ID__in =  distinct_group).order_by('Role_Group_ID','Menu_ID')
                for gr in list_group:
                    group_data = {
                        'Menu_ID' : gr.Menu_ID.Menu_ID,
                        'Role_Group_ID' : gr.Role_Group_ID,
                        'Role_Group_Name' : gr.Role_Group_Name,
                        'Role_Group_Address' : gr.Role_Group_Address,
                    }
                    list_group_data.append(group_data)

                menus = Menu.objects.filter(Menu_Status = True).order_by('Menu_Level')
                for menu in menus:
                    menu_data = {
                        'Menu_ID' : menu.Menu_ID,
                        'Menu_Name' : menu.Menu_Name,
                        'Menu_Icon' : menu.Menu_Icon,
                        'Menu_Adress' : menu.Menu_Adress,
                    }
                    list_menu_data.append(menu_data)

            context = {
                    'success' : True,
                    'isData' : True,
                    'Group_Roles'  : list_group_data,
                    'Menus'  : list_menu_data,
                }   
        else:
             context = {
                    'success' : True,
                    'isData' : False,
                }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#LOAD DATA ROLE MENU

#LOAD DATA READ COMMENT
@csrf_exempt
def read_comment(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            IDuser = userinfo['ID_user']
            # IsRead = Comment.objects.filter(
            #     Q(Ticket_ID__ID_User=IDuser) | Q(Ticket_ID__Ticket_User_Asign=IDuser) # Điều kiện OR
            #     & Q(ReadComment__ReadComment_Isread=True)  # Lọc theo trạng thái isread               
            # )
            IsRead = ReadComment.objects.filter(
                # Q(Comment_ID__Ticket_ID__ID_User=IDuser) | Q(Comment_ID__Ticket_ID__Ticket_User_Asign=IDuser) & Q(Comment_ID__Comment_Status = True),
                Comment_ID__Comment_Status = True,
                ReadComment_Isread=False, 
                ID_user = IDuser       
            ).order_by('-ReadComment_ID')
            a = list(IsRead)
            if IsRead:
                #  user = Users.objects.get(ID_user = IsRead.Id_user.ID_create)                  
                 list_comment = [{
                     'ReadComment_ID': r.ReadComment_ID, 
                     'TicketID': r.Comment_ID.Ticket_ID.Ticket_ID, 
                     'Ticket_Title': r.Comment_ID.Ticket_ID.Ticket_Title,
                     'CommentID': r.Comment_ID.Comment_ID,
                     'Comment_Date': r.Comment_ID.Comment_Date.strftime('%d/%m/%Y'),
                     'Comment_Time': r.Comment_ID.Comment_Time.strftime('%H:%M'),
                     'Ticket_Slug': convert_title_to_slug(r.Comment_ID.Ticket_ID.Ticket_Title),
                     'Comment_User': r.Comment_ID.ID_user.ID_user,
                     'Comment_UserName': r.Comment_ID.ID_user.FullName,
                     'Isread': r.ReadComment_Isread,
                     'Avatar':   r.Comment_ID.ID_user.Avatar if r.Comment_ID.ID_user.Avatar else 'Null',
                     'img'   :   r.Comment_ID.ID_user.displayName[0] + r.Comment_ID.ID_user.FullName[0],
                     } for r in IsRead]
                 context = { 
                'success': True,
                'data': list_comment,
                'Count': len(list_comment),
                }
            else:
                context = { 
                'success': True,
                'data': '',
                }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#LOAD DATA READ COMMENT
################################################### AUTHORIZE MENU MASTER PAGE ###########
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE DASHBOARD ########### 
def load_dashboard_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        current_date = datetime.datetime.now()
        #AUTHORIZATION DASHBOARD
        Dash_Role_Data = [
            {'TCode': 'ZDB_Ticket','Role': 'Ticket','Status': 'False'},
            {'TCode': 'ZDB_User','Role': 'User','Status': 'False'},
            {'TCode': 'ZDB_Per','Role': 'Performance','Status': 'False'},
        ]
        Sing_Role = Role_Single.objects.filter(Role_Group_ID = 1, Role_Status = True)
        for s in Sing_Role:
            auth = Authorization_User.objects.filter(
                ID_user =  userinfo['ID_user'], 
                Role_ID = s.Role_ID,
                Authorization_From__lte=current_date,
                Authorization_To__gte=current_date, 
                Authorization_Status = True).order_by('-Authorization_To').first() 
            if auth:
                for item_role in Dash_Role_Data:
                    tcode = item_role['TCode']
                    # if int(item_role['ID']) ==  auth.Role_ID.Role_ID:
                    #     item_role['Status'] = 'True'
                    if tcode in auth.Role_ID.Role_Name:
                         item_role['Status'] = 'True'

        #Dashboard for Admn
        Role_view_Data_Tickets = Dash_Role_Data[0]['Status']
        if int(userinfo['Acc_type']) < 2:
            IsAdmin = True
        else:
            IsAdmin = False

        if IsAdmin == True or Role_view_Data_Tickets == 'True':
            Role_view_Data_Users = Dash_Role_Data[1]['Status']
            Role_view_Data_Performance = Dash_Role_Data[2]['Status']

            ########################### TICKET ON YEAR DATA###########################    
            # total_tickets_in_year = Ticket.objects.filter(Ticket_Date__year=current_date.year).count()
            total_tickets_in_year = Ticket.objects.all().count()
            ########################### TICKET ON MONTH DATA###########################
            total_tickets_in_month = Ticket.objects.filter(Ticket_Date__month=current_date.now().month).count()
            total_tickets_in_last_month = Ticket.objects.filter(Ticket_Date__month=(current_date.now().month - 1)).count()
            ########################### TICKET ON WEEK DATA###########################
            total_tickets_in_week = Ticket.objects.filter(Ticket_Date__week=current_date.now().isocalendar()[1]).count()
            last_week = current_date.now().isocalendar()[1] - 1  
            total_tickets_in_last_week = Ticket.objects.filter(Ticket_Date__week=last_week).count()   
            ########################### TICKET ON DAY DATA###########################
            total_tickets_in_day = Ticket.objects.filter(Ticket_Date=current_date).count()

            ########################### TICKET BY MONTH ########################### 
            current_year = current_date.year
            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            #TICKET BY MONTH
            ticket_count_by_month = Ticket.objects.filter(Ticket_Date__year=current_year) \
                                                .annotate(month=Extract('Ticket_Date', 'month')) \
                                                .values('month') \
                                                .annotate(count=Count('Ticket_ID')) \
                                                .values('month', 'count') \
                                                .order_by('month')
            ticket_counts_month = [{'month': month_names[month['month']-1] , 'count':month['count']} for month in ticket_count_by_month]
           
            #TICKET BY MONTH AND STATUS
            ticket_counts_by_month_status = Ticket.objects.filter(Ticket_Date__year=current_year) \
                                       .annotate(month=Extract('Ticket_Date', 'month')) \
                                       .values('month', 'Ticket_Status') \
                                       .annotate(count=Count('Ticket_ID')) \
                                       .order_by('month', 'Ticket_Status')      
            ticket_counts_month_status =[ 
                {'month': 'January','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'February','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'March','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'April','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'May','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'June','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'July','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'August','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'September','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'October','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'November','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'December','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
            ]
            for row in ticket_counts_by_month_status:
                month = row['month']
                month_name = month_names[month - 1]
                status = row['Ticket_Status']
                count = row['count']
                
                for item in ticket_counts_month_status:
                    if item['month'] == month_name:
                        if status == 0:
                            item['Done'] += count
                            break
                        elif status == 1:
                            item['Inprogress'] += count
                            break
                        elif status == 2:
                            item['Pendding'] += count
                            break
                        elif status == 3:
                            item['Cancel'] += count
                            break
            #TICKET BY WEEK AND STATUS
            ticket_counts_by_week_status = Ticket.objects.filter(Ticket_Date__week=current_date.now().isocalendar()[1]) \
                                       .values('Ticket_Status') \
                                       .annotate(count=Count('Ticket_ID')) \
                                       .order_by('Ticket_Status')
            Ticket_by_week = []
            for week in ticket_counts_by_week_status:
                status = week['Ticket_Status']
                count = week['count']
                if status == 0:
                    Ticket_by_week.append({'Status': 'Hoàn Thành', 'count': count})
                elif status == 1:
                    Ticket_by_week.append({'Status': 'Đang Làm', 'count': count})
                elif status == 2:
                    Ticket_by_week.append({'Status': 'Đang Treo', 'count': count})
                elif status == 3:
                    Ticket_by_week.append({'Status': 'Hủy', 'count': count})

            #View Data User
            if Role_view_Data_Users == 'True' or IsAdmin == True:
                ########################### USERS ON YEAR DATA###########################    
                # total_users_in_year = Users.objects.filter(Date_Create__year=current_date.year).count()
                total_users_in_year = Users.objects.all().count()
                ########################### USERS ON MONTH DATA###########################
                total_users_in_month = Users.objects.filter(Date_Create__month=current_date.now().month).count()
                total_users_in_last_month = Users.objects.filter(Date_Create__month=(current_date.now().month) -1 ).count()
                ########################### USERS ON WEEK DATA###########################
                total_users_in_week = Users.objects.filter(Date_Create__week=current_date.now().isocalendar()[1]).count()  
                user_last_month =  current_date.now().isocalendar()[1] - 1    
                total_users_in_last_week = Users.objects.filter(Date_Create__week=user_last_month).count()       
                ########################### USERS ON DAY DATA###########################
                total_users_in_day = Users.objects.filter(Date_Create=current_date).count()
                #USERS BY MONTH AND STATUS
                user_counts_by_month_status = Users.objects.filter(Date_Create__year=current_year) \
                                        .annotate(month=Extract('Date_Create', 'month')) \
                                        .values('month') \
                                        .annotate(count=Count('ID_user')) \
                                        .values('month','User_Status','count') \
                                        .order_by('month','User_Status')
                user_counts_month_status =[ 
                    {'month': 'January','Active': 0, 'Unactive': 0}   , 
                    {'month': 'February','Active': 0, 'Unactive': 0}  , 
                    {'month': 'March','Active': 0, 'Unactive': 0}     , 
                    {'month': 'April','Active': 0, 'Unactive': 0}     , 
                    {'month': 'May','Active': 0, 'Unactive': 0}       , 
                    {'month': 'June','Active': 0, 'Unactive': 0}      ,    
                    {'month': 'July','Active': 0, 'Unactive': 0}      ,     
                    {'month': 'August','Active': 0, 'Unactive': 0}    ,     
                    {'month': 'September','Active': 0, 'Unactive':0}  ,     
                    {'month': 'October','Active': 0, 'Unactive': 0}   ,     
                    {'month': 'November','Active': 0, 'Unactive': 0}  ,     
                    {'month': 'December','Active': 0, 'Unactive': 0}  ,     
                ]
                for row in user_counts_by_month_status:
                    month = row['month']
                    month_name = month_names[month - 1]
                    type = row['User_Status']
                    count = row['count']
                    
                    for item in user_counts_month_status:
                        if item['month'] == month_name:
                            if type == True:
                                item['Active'] += count
                                break
                            elif type == False:
                                item['Unactive'] += count
                                break
                #USERS BY ROLE
                user_counts_by_type = Users.objects.all() \
                                        .annotate(count=Count('ID_user')) \
                                        .values('User_Type','count') \
                                        .order_by('User_Type')
                user_counts_acc_type =[ 
                    {'type': 'Administrator','count': 0}  , 
                    {'type': 'Mod',          'count': 0}  , 
                    {'type': 'User Member',  'count': 0}  ,     
                ]
                for row in user_counts_by_type:
                    type = row['User_Type']
                    count = row['count']
                    
                    for item in user_counts_acc_type:
                        if type == 0:
                            if item['type'] == 'Administrator':
                                item['count'] += count
                                break
                        elif type == 1:
                            if item['type'] == 'Mod':
                                item['count'] += count
                                break
                        elif type == 2:
                            if item['type'] == 'User Member':
                                item['count'] += count
                                break
                #USERS BY TYPE
                user_counts_by_acc = Users.objects.all() \
                                        .values('Acc_Type') \
                                        .annotate(count=Count('ID_user')) \
                                        .values('Acc_Type','count') \
                                        .order_by('Acc_Type')
                user_counts_acc = [{'Acc': acc['Acc_Type'] , 'count': acc['count']} for acc in user_counts_by_acc]
            else:
                total_users_in_year = ""
                total_users_in_month  = ""
                total_users_in_last_month  = ""
                total_users_in_week  = ""
                total_users_in_last_week = ""
                total_users_in_day = ""
                user_counts_month_status = ""
                user_counts_acc_type = ""
                user_counts_acc = ""
            

            #View Data Performance
            if Role_view_Data_Performance == 'True' or IsAdmin == True:
            #USERS PERCENT BY YEAR
                users = Users.objects.filter(User_Type__lt=2, User_Status =  True).values('ID_user','Avatar','FullName','displayName')
                ticket_counts_by_user_year = Ticket.objects.filter(Ticket_Date__year=current_year, Ticket_User_Asign__in= users.values('ID_user'), Ticket_Status = 0) \
                    .values('Ticket_User_Asign','Ticket_Name_Asign') \
                    .annotate(count=Count('Ticket_ID')) \
                    .order_by('Ticket_User_Asign')
                total_tickets_year = ticket_counts_by_user_year.aggregate(total=Sum('count'))['total']
                ticket_counts_by_user_year = ticket_counts_by_user_year.annotate(
                    percentage=Cast(F('count') * 100.0 / total_tickets_year, output_field=FloatField())
                )
                ticket_per_by_user_year = []
                for item in ticket_counts_by_user_year:
                    user = users.filter(ID_user=item['Ticket_User_Asign']).first()
                    ticket_per_by_user_year.append({
                        'name': item['Ticket_Name_Asign'],
                        'avatar': user['Avatar'],
                        'fullname': user['FullName'],
                        'displayname': user['displayName'],
                        'count': item['count'],
                        'per': item['percentage'],
                    })
                ticket_per_by_user_year = sorted(ticket_per_by_user_year, key=lambda x: x['per'], reverse=True)
                #USERS PERCENT BY MONTH
                ticket_counts_by_user_month = Ticket.objects.filter(Ticket_Date__month=current_date.month, Ticket_User_Asign__in= users.values('ID_user'),Ticket_Status = 0) \
                    .values('Ticket_User_Asign','Ticket_Name_Asign') \
                    .annotate(count=Count('Ticket_ID')) \
                    .order_by('Ticket_User_Asign')
                total_tickets_month = ticket_counts_by_user_month.aggregate(total=Sum('count'))['total']
                ticket_counts_by_user_month = ticket_counts_by_user_month.annotate(
                    percentage=Cast(F('count') * 100.0 / total_tickets_month, output_field=FloatField())
                )
                ticket_per_by_user_month = []
                for item in ticket_counts_by_user_month:
                    user = users.filter(ID_user=item['Ticket_User_Asign']).first()
                    ticket_per_by_user_month.append({
                        'name': item['Ticket_Name_Asign'],
                        'avatar': user['Avatar'],
                        'fullname': user['FullName'],
                        'displayname': user['displayName'],
                        'count': item['count'],
                        'per': item['percentage'],
                    })
                ticket_per_by_user_month = sorted(ticket_per_by_user_month, key=lambda x: x['per'], reverse=True)
                #USERS BY TYPE
                group_counts_by_ticket = Ticket.objects.filter(Ticket_Date__year=current_year) \
                                        .values('TGroup_ID__TGroup_Name') \
                                        .annotate(count=Count('Ticket_ID')) \
                                        .values('TGroup_ID__TGroup_Name','count') \
                                        .order_by('TGroup_ID')
                group_counts_ticket = [{'Groups': acc['TGroup_ID__TGroup_Name'] , 'count': acc['count']} for acc in group_counts_by_ticket]
            else:
                ticket_per_by_user_year  = ""
                ticket_per_by_user_month = "" 
                group_counts_ticket      = ""

            context = { 
            # 'IsAdmin'                 : True,   
            'IsAdmin'                 : IsAdmin,   
            'Total_year'              : total_tickets_in_year,
            'Total_month'             : total_tickets_in_month,
            'Total_last_month'        : total_tickets_in_last_month,
            'Total_week'              : total_tickets_in_week,
            'Total_last_week'         : total_tickets_in_last_week,
            'Total_day'               : total_tickets_in_day,
            'Ticket_by_month'         : ticket_counts_month,
            'Ticket_by_status'        : ticket_counts_month_status,
            'Ticket_by_week'          : Ticket_by_week,
            'Total_users_year'        : total_users_in_year,
            'Total_users_month'       : total_users_in_month,
            'Total_users_last_month'  : total_users_in_last_month,
            'Total_users_week'        : total_users_in_week,
            'Total_users_last_week'   : total_users_in_last_week,
            'Total_users_day'         : total_users_in_day,
            'User_by_month'           : user_counts_month_status,
            'User_by_type'            : user_counts_acc_type,
            'User_by_acc'             : user_counts_acc,
            'User_by_year_per'        : ticket_per_by_user_year,
            'User_by_month_per'       : ticket_per_by_user_month,
            'Dash_Role_Data'          : Dash_Role_Data,
            'Group_Ticket_Data'       : group_counts_ticket
            }               
        

         #Dashboard for User           
        else:
            userID = userinfo['ID_user']
            ########################### TICKET ON YEAR DATA###########################    
            total_tickets_in_year = Ticket.objects.filter(Ticket_Date__year=current_date.year, ID_User = userID).count()
            ########################### TICKET ON MONTH DATA###########################
            total_tickets_in_month = Ticket.objects.filter(Ticket_Date__month=current_date.now().month, ID_User = userID).count()
            total_tickets_in_last_month = Ticket.objects.filter(Ticket_Date__month=(current_date.now().month -1 ), ID_User = userID).count()
            ########################### TICKET ON WEEK DATA###########################
            total_tickets_in_week = Ticket.objects.filter(Ticket_Date__week=current_date.now().isocalendar()[1], ID_User = userID).count()   
            ticket_last_week =  current_date.now().isocalendar()[1] - 1    
            total_tickets_in_last_week = Ticket.objects.filter(Ticket_Date__week=ticket_last_week, ID_User = userID).count()       
            ########################### TICKET ON DAY DATA###########################
            total_tickets_in_day = Ticket.objects.filter(Ticket_Date=current_date, ID_User = userID).count()

             ########################### TICKET BY MONTH ########################### 
            current_year = current_date.year
            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            #TICKET BY MONTH
            ticket_count_by_month = Ticket.objects.filter(Ticket_Date__year=current_year, ID_User = userID) \
                                                .annotate(month=Extract('Ticket_Date', 'month')) \
                                                .values('month') \
                                                .annotate(count=Count('Ticket_ID')) \
                                                .values('month', 'count') \
                                                .order_by('month')
            ticket_counts_month = [{'month': month_names[month['month']-1] , 'count':month['count']} for month in ticket_count_by_month]
           
            #TICKET BY MONTH AND STATUS
            ticket_counts_by_month_status = Ticket.objects.filter(Ticket_Date__year=current_year,ID_User = userID) \
                                       .annotate(month=Extract('Ticket_Date', 'month')) \
                                       .values('month', 'Ticket_Status') \
                                       .annotate(count=Count('Ticket_ID')) \
                                       .order_by('month', 'Ticket_Status')      
            ticket_counts_month_status =[ 
                {'month': 'January','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'February','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'March','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'April','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'May','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'June','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'July','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'August','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'September','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'October','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'November','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
                {'month': 'December','Done': 0, 'Inprogress': 0, 'Pendding': 0, 'Cancel': 0},
            ]
            for row in ticket_counts_by_month_status:
                month = row['month']
                month_name = month_names[month - 1]
                status = row['Ticket_Status']
                count = row['count']
                
                for item in ticket_counts_month_status:
                    if item['month'] == month_name:
                        if status == 0:
                            item['Done'] += count
                            break
                        elif status == 1:
                            item['Inprogress'] += count
                            break
                        elif status == 2:
                            item['Pendding'] += count
                            break
                        elif status == 3:
                            item['Cancel'] += count
                            break
            #TICKET BY WEEK AND STATUS
            ticket_counts_by_week_status = Ticket.objects.filter(Ticket_Date__week=current_date.now().isocalendar()[1], ID_User = userID) \
                                       .values('Ticket_Status') \
                                       .annotate(count=Count('Ticket_ID')) \
                                       .order_by('Ticket_Status')
            Ticket_by_week = []
            for week in ticket_counts_by_week_status:
                status = week['Ticket_Status']
                count = week['count']
                if status == 0:
                    Ticket_by_week.append({'Status': 'Hoàn Thành', 'count': count})
                elif status == 1:
                    Ticket_by_week.append({'Status': 'Đang Làm', 'count': count})
                elif status == 2:
                    Ticket_by_week.append({'Status': 'Đang Treo', 'count': count})
                elif status == 3:
                    Ticket_by_week.append({'Status': 'Hủy', 'count': count})

            

            context = { 
                'Total_year'        : total_tickets_in_year,
                'Total_month'       : total_tickets_in_month,
                'Total_last_month'  : total_tickets_in_last_month,
                'Total_week'        : total_tickets_in_week,
                'Total_last_week'   : total_tickets_in_last_week,
                'Total_day'         : total_tickets_in_day,
                'Ticket_by_month'   : ticket_counts_month,
                'Ticket_by_status'  : ticket_counts_month_status,
                'Ticket_by_week'    : Ticket_by_week,
                'Dash_Role_Data'    : Dash_Role_Data
            }

        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('')

def Dashboard(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if(request.session['UserInfo']):
        return render(request, 'Ticket_Dashboard.html')
    else:
        return redirect('/')
    # template =  loader.get_template('Ticket_Dashboard.html')
    # return HttpResponse(template.render())
################################################### PAGE DASHBOARD ########### 
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE TICKET DATA #######################  
#FUNCTION LOAD AND PROCESS DATA TICKET
@csrf_exempt
def load_Ticket_Json_Test(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### TICKET DATA###########################   
        isAdmin = request.POST.get('isAdmin') 
        if(isAdmin == 'true'):
            tickets = Ticket.objects.all().order_by('-Ticket_ID')
        else:
            tickets = Ticket.objects.filter(ID_User = userinfo['ID_user']).order_by('-Ticket_ID')

        a= list(tickets)

        # list_ticket = []
        # for ticket in tickets:
        #     slug_title =  convert_title_to_slug(ticket.Ticket_Title)
        #     ticket_data = {
        #         'Ticket_ID': ticket.Ticket_ID,
        #         'Ticket_Title': ticket.Ticket_Title,
        #         'Ticket_Title_Slug': slug_title,
        #         # 'Ticket_Desc': ticket.Ticket_Desc[:30] if len(ticket.Ticket_Desc) > 30 else ticket.Ticket_Desc,
        #         'Ticket_Desc': '',
        #         'Ticket_Type': ticket.Ticket_Type,
        #         'Company_ID': ticket.Company_ID.Company_ID,
        #         'Company_Name': ticket.Company_ID.Company_Name,
        #         'Group_ID': ticket.TGroup_ID.TGroup_ID,
        #         'Group_Name': ticket.TGroup_ID.TGroup_Name,
        #         'ID_user': ticket.ID_User.ID_user,
        #         'Ticket_User_Name': ticket.Ticket_User_Name,
        #         # 'Ticket_Date': ticket.Ticket_Date.strftime('%Y-%m-%d'),
        #         'Ticket_Date': ticket.Ticket_Date.strftime('%d/%m/%Y'),
        #         'Ticket_Time': ticket.Ticket_Time.strftime('%H:%M'),
        #         'Ticket_User_Asign': ticket.Ticket_User_Asign,
        #         'Ticket_Name_Asign': ticket.Ticket_Name_Asign,
        #         'Ticket_Status': ticket.Ticket_Status,
        #     }
        #     list_ticket.append(ticket_data)       
        # data = json.dumps(list_ticket, default=date_handler, ensure_ascii=False)

         
        list_ticket = [{
            'Ticket_ID': ticket.Ticket_ID,
            'Ticket_Title': ticket.Ticket_Title,
            'Ticket_Title_Slug': convert_title_to_slug(ticket.Ticket_Title),
            'Ticket_Desc': '',
            'Ticket_Type': ticket.Ticket_Type,
            'Company_ID': ticket.Company_ID.Company_ID,
            'Company_Name': ticket.Company_ID.Company_Name,
            'Group_ID': ticket.TGroup_ID.TGroup_ID,
            'Group_Name': ticket.TGroup_ID.TGroup_Name,
            'ID_user': ticket.ID_User.ID_user,
            'Ticket_User_Name': ticket.Ticket_User_Name,
            'Ticket_Date': ticket.Ticket_Date.strftime('%d/%m/%Y'),
            'Ticket_Time': ticket.Ticket_Time.strftime('%H:%M'),
            'Ticket_User_Asign': ticket.Ticket_User_Asign,
            'Ticket_Name_Asign': ticket.Ticket_Name_Asign,
            'Ticket_Status': ticket.Ticket_Status,
        }for ticket in tickets]
        ########################### COMPANY DATA###########################
        companys = Company.objects.filter(Company_Status = True)
        list_companys = [{'Company_ID': company.Company_ID, 'Company_Name': company.Company_Name} for company in companys]
        ########################### TGROUP DATA###########################
        tgroups = TGroup.objects.filter(TGroup_Status = True)
        list_tgroups = [{'TGroup_ID': tgroup.TGroup_ID, 'TGroup_Name': tgroup.TGroup_Name} for tgroup in tgroups]
        ########################### SUPPORT DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]
        ########################### COMPANY USER DATA###########################
        # id_user = userinfo['ID_user']
        comp_user = Users.objects.filter(ID_user = userinfo['ID_user'])
        list_user_company = [{'Company_ID': uc.Company_ID} for uc in comp_user]
        context = { 
            'data': list_ticket,
            'companys': list_companys,
            'tgroups': list_tgroups,
            'users': list_users,
            'users_company': list_user_company}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
        # return HttpResponse(data, content_type='application/json')
    else:
        return redirect('/')

@csrf_exempt
def load_Ticket_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### TICKET DATA###########################   
        isAdmin = request.POST.get('isAdmin') 
        if(isAdmin == 'true'):
            # tickets = Ticket.objects.all().order_by('-Ticket_ID')
            tickets = Ticket.objects.values(
                'Ticket_ID', 'Ticket_Title', 'Ticket_Type', 'Company_ID__Company_ID', 'Company_ID__Company_Name', 
                'TGroup_ID__TGroup_ID','TGroup_ID__TGroup_Name', 'ID_User__ID_user', 'Ticket_User_Name', 
                'Ticket_Date', 'Ticket_Time', 'Ticket_User_Asign', 'Ticket_Name_Asign', 'Ticket_Status'
            ).all().order_by('-Ticket_ID')
        else:
            tickets = Ticket.objects.filter(ID_User = userinfo['ID_user']).values(
                'Ticket_ID', 'Ticket_Title', 'Ticket_Type', 'Company_ID__Company_ID', 'Company_ID__Company_Name', 
                'TGroup_ID__TGroup_ID','TGroup_ID__TGroup_Name', 'ID_User__ID_user', 'Ticket_User_Name', 
                'Ticket_Date', 'Ticket_Time', 'Ticket_User_Asign', 'Ticket_Name_Asign', 'Ticket_Status'
            ).order_by('-Ticket_ID')

        list_ticket = []
        for ticket in tickets:
            if isinstance(ticket, dict):  # Kiểm tra nếu ticket là một từ điển
                ticket_data = {
                    'Ticket_ID': ticket.get('Ticket_ID'),
                    'Ticket_Title': ticket.get('Ticket_Title'),
                    'Ticket_Title_Slug': convert_title_to_slug(ticket.get('Ticket_Title')),
                    'Ticket_Desc': '',
                    'Ticket_Type': ticket.get('Ticket_Type'),
                    'Company_ID': ticket.get('Company_ID__Company_ID'),
                    'Company_Name': ticket.get('Company_ID__Company_Name'),
                    'Group_ID': ticket.get('TGroup_ID__TGroup_ID'),
                    'Group_Name': ticket.get('TGroup_ID__TGroup_Name'),
                    'ID_user': ticket.get('ID_User__ID_user'),
                    'Ticket_User_Name': ticket.get('Ticket_User_Name'),
                    'Ticket_Date': ticket.get('Ticket_Date').strftime('%d/%m/%Y') if ticket.get('Ticket_Date') else None,
                    'Ticket_Time': ticket.get('Ticket_Time').strftime('%H:%M') if ticket.get('Ticket_Time') else None,
                    'Ticket_User_Asign': ticket.get('Ticket_User_Asign'),
                    'Ticket_Name_Asign': ticket.get('Ticket_Name_Asign'),
                    'Ticket_Status': ticket.get('Ticket_Status'),
                }
                list_ticket.append(ticket_data)
        ########################### COMPANY DATA###########################
        companys = Company.objects.filter(Company_Status = True)
        list_companys = [{'Company_ID': company.Company_ID, 'Company_Name': company.Company_Name} for company in companys]
        ########################### TGROUP DATA###########################
        tgroups = TGroup.objects.filter(TGroup_Status = True)
        list_tgroups = [{'TGroup_ID': tgroup.TGroup_ID, 'TGroup_Name': tgroup.TGroup_Name} for tgroup in tgroups]
        ########################### SUPPORT DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]
        ########################### COMPANY USER DATA###########################
        # id_user = userinfo['ID_user']
        comp_user = Users.objects.filter(ID_user = userinfo['ID_user'])
        list_user_company = [{'Company_ID': uc.Company_ID} for uc in comp_user]
        context = { 
            'data': list_ticket,
            'companys': list_companys,
            'tgroups': list_tgroups,
            'users': list_users,
            'users_company': list_user_company}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
        # return HttpResponse(data, content_type='application/json')
    else:
        return redirect('/')

def load_Ticket(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # cookie_facebook_data   = GetCookie(request, 'cookie_facebook_data')
    # cookie_google_data     = GetCookie(request, 'cookie_google_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListTicket.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA TICKET

#FUNCTION TRANSFER TITLE
def convert_title_to_slug(title):
    # Chuyển đổi tiêu đề sang dạng không dấu
    slug = slugify(title)
    # Thay thế khoảng trắng bằng dấu "-"
    slug = re.sub(r'\s+', '-', slug)
    return slug
#FUNCTION TRANSFER TITLE

#FUNCTION UPDATE STATUS TICKET 
@csrf_exempt
def Ticket_Status_Update(request):
    if request.method == 'POST':
        ticket_id = request.POST.get('ticketid')
        new_status = request.POST.get('new_status')
        try:
            ticket =  Ticket.objects.get(Ticket_ID = ticket_id)
            if(ticket):
                ticket.Ticket_Status = new_status
                ticket.save()
                # Status Done -  Send Mail
                if new_status == '0': 
                    FullNamesession =  request.session['UserInfo']['FullName'] 
                    send_email_Done(request,FullNamesession,ticket)
                return JsonResponse({'success': True, 'message': 'Cập nhật status thành công!',}) 
            else:
                return JsonResponse({'success': False, 'message': 'ID ' + ticket_id +' không tồn tại'})
        except Ticket.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ID ' + ticket_id +' không tồn tại',}) 
    else:
        return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
#FUNCTION UPDATE STATUS TICKET 

#FUNCTION UPDATE ASSIGN TICKET 
@csrf_exempt
def Ticket_Assign_Update(request):
    try:
        if request.method == 'POST':
            ticket_id = request.POST.get('ticketid')
            new_assign = request.POST.get('new_assign')
            try:
                ticket =  Ticket.objects.get(Ticket_ID = ticket_id)
                if(ticket):
                    user = Users.objects.get(ID_user = new_assign)
                    if user:
                        ticket.Ticket_User_Asign = user.ID_user
                        ticket.Ticket_Name_Asign = user.FullName
                        ticket.save()
                        return JsonResponse({'success': True, 'message': 'Cập nhật status thành công!','FullName': user.FullName}) 
                    else:
                        return JsonResponse({'success': False, 'message': 'User Không Tồn Tại!',}) 
                else:
                    return JsonResponse({'success': False, 'message': 'ID ' + ticket_id +' không tồn tại'})
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + ticket_id +' không tồn tại',}) 
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ASSIGN TICKET 

#FUNCTION DELETE TICKET 
@csrf_exempt
def delete_ticket(request):
    try:
        if request.method == 'POST':
            ticketid = request.POST.get('ticketid')
            try:
                ticket = Ticket.objects.get(Ticket_ID = ticketid)
                if(ticket):
                    # comment = Comment.objects.filter(Ticket_ID=ticketid).first()
                    # if comment:
                    #     comment.delete()
                    # attach =  Attachment.objects.filter(Ticket_ID=ticketid).first()
                    # if attach:
                    #     attach.delete()
                    # timage = TImage.objects.filter(Ticket_ID=ticketid).first()
                    # if timage:
                    #     timage.delete()
                    ticket.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'ID ' + ticketid +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + ticketid +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE TICKET 

#FUNCTION CREATE TICKET 
@csrf_exempt
def Create_Ticket(request):
    try:
        if request.method == 'POST':
            title = request.POST.get('title')
            detail = request.POST.get('detail')  
            type = request.POST.get('type')
            companyID = int(request.POST.get('company'))
            groupID = int(request.POST.get('group'))
            support = request.POST.get('support')
            supportName = request.POST.get('supportName')       
            userID =  request.session['UserInfo']['ID_user']
            userName =  request.session['UserInfo']['FullName']

            if support:
                assign_User = support
                assign_Name = supportName
            else:    
                suggest_data = suggest_ticket(request,title, groupID)
                assign_User = suggest_data['ticket_assigned_user_id']
                assign_Name = suggest_data['ticket_assigned_username']
            
            company = Company.objects.get( Company_ID = companyID)
            group = TGroup.objects.get( TGroup_ID = groupID)
            user = Users.objects.get( ID_user = userID)

            slug_title =  convert_title_to_slug(title)

            ticket = Ticket.objects.create(
                Ticket_Title = title, 
                Ticket_Desc  = detail, 
                Ticket_Type  = type, 
                Company_ID   = company,
                TGroup_ID    = group, 
                ID_User      = user,
                Ticket_User_Name = userName,
                Ticket_Date  = datetime.datetime.now().date(),
                Ticket_Time  = datetime.datetime.now().time(),
                Ticket_User_Asign = assign_User,
                Ticket_Name_Asign = assign_Name,
                Ticket_Status = 1,
            )
            ticket.save()
            TicketID = ticket.Ticket_ID
            email_assign = Users.objects.get(ID_user = assign_User)
            
            send_email(email_assign.Mail,TicketID,ticket.Ticket_Title,slug_title,ticket.Ticket_User_Name,ticket.Ticket_Date.strftime('%d/%m/%Y'),ticket.Ticket_Time.strftime('%H:%M'), ticket.Ticket_Name_Asign,request)
            # if TicketID:
                
            #     attach_file = Attachment.objects.create(

            #     )
            # return JsonResponse(response, status=201)
            return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Được Gửi Thành Công!',
                    'Ticket_ID': TicketID,
                    'Ticket_Title': ticket.Ticket_Title,
                    'Ticket_Title_Slug': slug_title,
                    'Ticket_Desc': ticket.Ticket_Desc,
                    'Ticket_Type': ticket.Ticket_Type,
                    'Company_Name': company.Company_Name,
                    'Group_Name': group.TGroup_Name,
                    'Email'     : email_assign.Mail,
                    'Ticket_User_Name': ticket.Ticket_User_Name,
                    'Ticket_Date': ticket.Ticket_Date.strftime('%d/%m/%Y'),
                    'Ticket_Time': ticket.Ticket_Time.strftime('%H:%M'),
                    'Ticket_Name_Asign': ticket.Ticket_Name_Asign,
                    'Ticket_Status': ticket.Ticket_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE TICKET 

#FUNCTION UPDATE DATA TICKET 
@csrf_exempt
def Data_Update_Ticket(request):
       try:
        if request.method == 'POST':
            id_ticket = request.POST.get('id_ticket')
            ticket = Ticket.objects.filter(Ticket_ID = id_ticket)
            list_ticket = [{'Ticket_ID': p.Ticket_ID, 
                     'Ticket_Title': p.Ticket_Title, 
                     'Ticket_Desc': p.Ticket_Desc, 
                     'Company_ID': p.Company_ID.Company_ID, 
                     'TGroup_ID': p.TGroup_ID.TGroup_ID, 
                     'ID_User': p.ID_User.ID_user, 
                     'Ticket_Type': p.Ticket_Type,
                     'Ticket_Status': p.Ticket_Status,} for p in ticket]
            attachs = Attachment.objects.filter(Ticket_ID = id_ticket, Attachment_Status =  True)
            list_attachs = [{'Attachment': attach.Attachment_Url} for attach in attachs]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Tickets': list_ticket,
                'Attachs': list_attachs}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Ticket ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA TICKET 

#FUNCTION UPDATE TICKET 
@csrf_exempt
def Update_Ticket(request):
    try:
        if request.method == 'POST':
            ticketID = request.POST.get('ticketID')       
            status = int(request.POST.get('status'))
            title = request.POST.get('title')
            detail = request.POST.get('detail')  
            type = request.POST.get('type')
            companyID = int(request.POST.get('company'))
            groupID = int(request.POST.get('group'))
            support = request.POST.get('support')
            supportName = request.POST.get('supportName')       
            # userID =  request.session['UserInfo']['ID_user']
            # userName =  request.session['UserInfo']['FullName']
            
            company = Company.objects.get( Company_ID = companyID)
            group = TGroup.objects.get( TGroup_ID = groupID)
            # user = Users.objects.get( ID_user = userID)

            # Cập nhật đối tượng Product trong CSDL
            tickets = Ticket.objects.get(Ticket_ID = ticketID)
            if tickets:
                tickets.Ticket_Title = title
                tickets.Ticket_Desc  = detail
                tickets.Ticket_Type  = type
                tickets.Company_ID   = company
                tickets.TGroup_ID    = group
                # tickets.Ticket_User_Name = userName,
                tickets.Ticket_User_Asign = support
                tickets.Ticket_Name_Asign = supportName
                tickets.Ticket_Status = status
                tickets.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Ticket_ID':        ticketID,
                    'Ticket_Title':     tickets.Ticket_Title,
                    'Ticket_Desc':      tickets.Ticket_Desc,
                    'Ticket_Type':      tickets.Ticket_Type,
                    'Company_ID':       company.Company_ID,
                    'Company_Name':     company.Company_Name,
                    'Group_ID':         group.TGroup_ID, 
                    'Group_Name':       group.TGroup_Name, 
                    'Ticket_User_Name': tickets.Ticket_User_Name,
                    'Ticket_Name_Asign': tickets.Ticket_Name_Asign,
                    'Ticket_Status':    tickets.Ticket_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Yêu Cầu Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE TICKET 

#FUNCTION UPDATE ATTACHMENT TICKET
@csrf_exempt
def upload_files(request):
    try:
        # if request.method == 'POST':
        if request.method == 'POST' and request.FILES:
            files = request.FILES.getlist('files')
            ticketID = request.POST.get('ticketID')
            if files:
                fs = FileSystemStorage(location='my_project/static/Asset/Attachment-Upload/')
                userID =  request.session['UserInfo']['ID_user']
                userName =  request.session['UserInfo']['FullName']
                user = Users.objects.get(ID_user = userID)
                ticket_data = Ticket.objects.get(Ticket_ID = ticketID)
                for file in files:
                    date = datetime.datetime.now().date()
                    time = datetime.datetime.now().time()
                    file_name = ticketID + '_' + date.strftime('%d%m%Y') + '_' + time.strftime('%H%M') + '_'+userID+'_' + file.name
                    fs.save(file_name, file)

                    attach = Attachment.objects.create(
                        Ticket_ID = ticket_data,
                        ID_user   = user,
                        Attachment_Url = file_name,
                        Attachment_User_Name = userName,
                        Attachment_Date = date,
                        Attachment_Time = time,
                        Attachment_Status = 1,
                    )
                    attach.save()
                return JsonResponse({'success': True,'message': 'Yêu Cầu Thành Công!'})
            else:
                return JsonResponse({'success': False,'message': 'Upload file lỗi'})
        else:
            return JsonResponse({'success': False,'message': 'Không tìm thấy file'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ATTACHMENT TICKET

#FUNCTION SUGGEST TICKET
def suggest_ticket(request,Title_ticket, Group_ticket):
    user = Assign_User.objects.filter(TGroup_ID = Group_ticket, Assign_User_Status = True).values('ID_user')
    if user: 
        user_assign  = Users.objects.filter(User_Type__lt = 2, User_Status = True, ID_user__in = user.values('ID_user') ).values('ID_user','FullName')
    else:
        user_assign  = Users.objects.filter(User_Type__lt = 2, User_Status = True ).values('ID_user','FullName')
        
    words = Title_ticket.split()  # Tách chuỗi thành các từ riêng lẻ, 
    query = Q()  # Tạo một truy vấn rỗng
    for word in words:
        query |= Q(Ticket_Title__icontains=word)  # Sử dụng Q objects để tạo điều kiện tìm kiếm

    titles = (Ticket.objects
        .filter(Ticket_Status__lt = 2, Ticket_User_Asign__in = user_assign.values('ID_user'), TGroup_ID = Group_ticket)
        .filter(query)  # Áp dụng điều kiện tìm kiếm vào truy vấn
        .values_list('Ticket_Title','Ticket_User_Asign','Ticket_Name_Asign'))
        # .values_list('Ticket_Title','Ticket_User_Asign','Ticket_Name_Asign', flat=True))

    if titles:
        closest_data = find_closest_title(Title_ticket, titles)
        if closest_data['closest_User_ID']:            
            suggest_data = {
                'ticket_assigned_user_id': closest_data['closest_User_ID'],
                'ticket_assigned_username': closest_data['closest_User_Name']
            }
            return suggest_data
        else:
            data_user =[]
            for item in user_assign.values('ID_user'):
                a = item['ID_user']
                count_ticket = (Ticket.objects
                .filter(Ticket_Status = 1 , Ticket_User_Asign = item['ID_user'], TGroup_ID = Group_ticket)
                .values('Ticket_User_Asign', 'Ticket_Name_Asign')
                .annotate(total_assigned=Count('Ticket_ID'))
                .order_by('total_assigned')
                .first())
                if count_ticket:
                    data = {'ID': count_ticket['Ticket_User_Asign'],'Name': count_ticket['Ticket_Name_Asign'],'total_assigned':count_ticket['total_assigned']}
                    data_user.append(data)
                else:
                    name_user = Users.objects.filter(ID_user = item['ID_user']).values('ID_user','FullName').first()
                    data = {'ID': name_user['ID_user'] ,'Name': name_user['FullName'],'total_assigned': 0 }
                    data_user.append(data)
            data_user_sorted = sorted(data_user, key=lambda x: x['total_assigned'])
            first_row = data_user_sorted[0] if data_user_sorted else None
            suggest_data = {
                'ticket_assigned_user_id': first_row['ID'],
                'ticket_assigned_username': first_row['Name']
            }
            return suggest_data
        
            # count_ticket = (Ticket.objects
            #     .filter(Ticket_Status__lt = 2 , Ticket_User_Asign__in = user_assign.values('ID_user'), TGroup_ID = Group_ticket)
            #     .values('Ticket_User_Asign', 'Ticket_Name_Asign')
            #     .annotate(total_assigned=Count('Ticket_ID'))
            #     .order_by('total_assigned')
            #     .first())         
            # suggest_data = {
            #     'ticket_assigned_user_id': count_ticket['Ticket_User_Asign'],
            #     'ticket_assigned_username': count_ticket['Ticket_Name_Asign']
            # }
            # return suggest_data
             
    else:
        data_user =[]
        for item in user_assign.values('ID_user'):
            a = item['ID_user']
            count_ticket = (Ticket.objects
            .filter(Ticket_Status = 1 , Ticket_User_Asign = item['ID_user'], TGroup_ID = Group_ticket)
            .values('Ticket_User_Asign', 'Ticket_Name_Asign')
            .annotate(total_assigned=Count('Ticket_ID'))
            .order_by('total_assigned')
            .first())
            if count_ticket:
                data = {'ID': count_ticket['Ticket_User_Asign'],'Name': count_ticket['Ticket_Name_Asign'],'total_assigned':count_ticket['total_assigned']}
                data_user.append(data)
            else:
                name_user = Users.objects.filter(ID_user = item['ID_user']).values('ID_user','FullName').first()
                data = {'ID': name_user['ID_user'] ,'Name': name_user['FullName'],'total_assigned': 0 }
                data_user.append(data)
        data_user_sorted = sorted(data_user, key=lambda x: x['total_assigned'])
        first_row = data_user_sorted[0] if data_user_sorted else None
        suggest_data = {
            'ticket_assigned_user_id': first_row['ID'],
            'ticket_assigned_username': first_row['Name']
        }
        return suggest_data

        # count_ticket = (Ticket.objects
        #         .filter(Ticket_Status__lt = 2 , Ticket_User_Asign__in = user_assign.values('ID_user'), TGroup_ID = Group_ticket)
        #         .values('Ticket_User_Asign', 'Ticket_Name_Asign')
        #         .annotate(total_assigned=Count('Ticket_ID'))
        #         .order_by('total_assigned')
        #         .first()) 
        # if  count_ticket:         
        #     suggest_data = {
        #             'ticket_assigned_user_id': count_ticket['Ticket_User_Asign'],
        #             'ticket_assigned_username': count_ticket['Ticket_Name_Asign']
        #         }
        #     return suggest_data
        # else:
        #     count_ticket = (Ticket.objects
        #         .filter(Ticket_Status__lt = 2 , Ticket_User_Asign__in = user_assign.values('ID_user'))
        #         .values('Ticket_User_Asign', 'Ticket_Name_Asign')
        #         .annotate(total_assigned=Count('Ticket_ID'))
        #         .order_by('total_assigned')
        #         .first())
        #     suggest_data = {
        #             'ticket_assigned_user_id': count_ticket['Ticket_User_Asign'],
        #             'ticket_assigned_username': count_ticket['Ticket_Name_Asign']
        #         }
        #     return suggest_data

def find_closest_title(input_str, titles):
    closest_title = None
    closest_distance = float('inf')
    closest_User_ID = ''
    closest_User_Name = ''
    
    if titles:
        for title in titles:
            distance = compute_similarity(input_str, title[0])
            if distance < 0.3:
                if distance < closest_distance:
                    closest_distance = distance
                    closest_title = title[0]
                    closest_User_ID = title[1]
                    closest_User_Name = title[2]
            if distance < 0.1:
                closest_distance = distance
                closest_title = title[0]
                closest_User_ID = title[1]
                closest_User_Name = title[2]
                closest_data = {
                    'closest_title': closest_title,
                    'closest_User_ID': closest_User_ID,
                    'closest_User_Name': closest_User_Name,
                }
                return closest_data
    else:
        # Xử lý trường hợp danh sách titles rỗng
        closest_title = ''
    
    closest_data = {
        'closest_title': closest_title,
        'closest_User_ID': closest_User_ID,
        'closest_User_Name': closest_User_Name,
    }
    return closest_data

def compute_similarity(input_str, title):
    similarity_ratio = similar(input_str, title)
    
    return 1 - similarity_ratio

def similar(a, b):
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio()

#FUNCTION SUGGEST TICKET


################################################### PAGE TICKET DATA #######################  

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE COMPANY DATA #######################
#FUNCTION LOAD AND PROCESS DATA COMPANY
def load_Company_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        companys = Company.objects.all().order_by('-Company_ID')
        list_company = []
        for company in companys:
            company_data = {
                'Company_ID':           company.Company_ID,
                'Company_Name':         company.Company_Name,
                'Company_User_Name':    company.Company_User_Name,
                'Company_Date':         company.Company_Date.strftime('%d/%m/%Y'),
                'Company_Time':          company.Company_Time.strftime('%H:%M'),
                'Company_Status':       company.Company_Status,
            }
            list_company.append(company_data)       
        ########################### USERS DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]


        context = { 
            'data': list_company,
            'users': list_users,}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('')

def load_company(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListCompany.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA COMPANY

#FUNCTION CREATE TICKET 
@csrf_exempt
def Create_Company(request):
    try:
        if request.method == 'POST':
            Company_Name = request.POST.get('Company_Name')

            userID =  request.session['UserInfo']['ID_user']
            user = Users.objects.get(ID_user = userID)
               
            company = Company.objects.create(
                Company_Name = Company_Name, 
                ID_user  = user, 
                Company_User_Name  = user.FullName, 
                Company_Date  = datetime.datetime.now().date(),
                Company_Time  = datetime.datetime.now().time(),
                Company_Status = 1,
            )
            company.save()
            companyID = company.Company_ID
            return JsonResponse({
                    'success': True,
                    'message': 'Công Ty Tạo Thành Công!',
                    'Company_ID':           companyID,
                    'Company_Name':         company.Company_Name,
                    'ID_user':              user.ID_user,
                    'Company_User_Name':    user.FullName,
                    'Company_Date':         company.Company_Date.strftime('%d/%m/%Y'),
                    'Company_Time':         company.Company_Time.strftime('%H:%M'),
                    'Company_Status':       company.Company_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE TICKET 

#FUNCTION DELETE TICKET 
@csrf_exempt
def Delete_Company(request):
    try:
        if request.method == 'POST':
            CompanyID = request.POST.get('CompanyID')
            try:
                company = Company.objects.get(Company_ID = CompanyID)
                if(company):
                    company.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'Company ID ' + CompanyID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + CompanyID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE TICKET 

#FUNCTION UPDATE TICKET 
@csrf_exempt
def Update_Company(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            companyID = request.POST.get('CompanyID')       
            CompanyName = request.POST.get('CompanyName')       
            company = Company.objects.get(Company_ID = companyID)
            if company:
                if CompanyName:
                    company.Company_Name = CompanyName
                    company.Company_Status = status
                    company.save()
                else:
                    company.Company_Name = company.Company_Name
                    company.Company_Status = status
                    company.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Company_ID':        company.Company_ID,
                    'Company_Name':      company.Company_Name,                   
                    'Company_Status':    company.Company_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Công Ty Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE TICKET 

#FUNCTION UPDATE DATA COMPANY 
@csrf_exempt
def Data_Update_Company(request):
       try:
        if request.method == 'POST':
            companyID = request.POST.get('companyID')
            company = Company.objects.filter(Company_ID = companyID)
            list_company = [{
                     'Company_ID':     p.Company_ID, 
                     'Company_Name':   p.Company_Name, 
                     'Company_Status': p.Company_Status,} for p in company]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Companys': list_company,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Company ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA COMPANY 

################################################### PAGE COMPANY DATA #######################  

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE GROUP DATA #########################
#FUNCTION LOAD AND PROCESS DATA GROUP
def load_Group_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        groups = TGroup.objects.all().order_by('-TGroup_ID')
        list_group = []
        for group in groups:
            group_data = {
                'TGroup_ID':           group.TGroup_ID,
                'TGroup_Name':         group.TGroup_Name,
                'TGroup_User_Name':    group.TGroup_User_Name,
                'TGroup_Date':         group.TGroup_Date.strftime('%d/%m/%Y'),
                'TGroup_Time':         group.TGroup_Time.strftime('%H:%M'),
                'TGroup_Status':       group.TGroup_Status,
            }
            list_group.append(group_data)       
        ########################### USERS DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]


        context = { 
            'data': list_group,
            'users': list_users,}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_group(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListGroup.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA GROUP

#FUNCTION CREATE GROUP 
@csrf_exempt
def Create_Group(request):
    try:
        if request.method == 'POST':
            Group_Name = request.POST.get('Group_Name')

            userID =  request.session['UserInfo']['ID_user']
            user = Users.objects.get(ID_user = userID)
               
            group = TGroup.objects.create(
                TGroup_Name = Group_Name, 
                ID_user  = user, 
                TGroup_User_Name  = user.FullName, 
                TGroup_Date  = datetime.datetime.now().date(),
                TGroup_Time  = datetime.datetime.now().time(),
                TGroup_Status = 1,
            )
            group.save()
            GroupID = group.TGroup_ID
            return JsonResponse({
                    'success': True,
                    'message': 'Công Ty Tạo Thành Công!',
                    'TGroup_ID':           GroupID,
                    'TGroup_Name':         group.TGroup_Name,
                    'ID_user':              user.ID_user,
                    'TGroup_User_Name':    user.FullName,
                    'TGroup_Date':         group.TGroup_Date.strftime('%d/%m/%Y'),
                    'TGroup_Time':         group.TGroup_Time.strftime('%H:%M'),
                    'TGroup_Status':       group.TGroup_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE GROUP 

#FUNCTION DELETE GROUP 
@csrf_exempt
def Delete_Group(request):
    try:
        if request.method == 'POST':
            GroupID = request.POST.get('GroupID')
            try:
                group = TGroup.objects.get(TGroup_ID = GroupID)
                if(group):
                    group.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'Group ID ' + GroupID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + GroupID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE GROUP 

#FUNCTION UPDATE GROUP 
@csrf_exempt
def Update_Group(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            GroupID = request.POST.get('GroupID')       
            groupName = request.POST.get('GroupName')       
            group = TGroup.objects.get(TGroup_ID = GroupID)
            if group:
                if groupName:
                    group.TGroup_Name = groupName
                    group.TGroup_Status = status
                    group.save()
                else:
                    group.TGroup_Name = group.TGroup_Name
                    group.TGroup_Status = status
                    group.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'TGroup_ID':        group.TGroup_ID,
                    'TGroup_Name':      group.TGroup_Name,                   
                    'TGroup_Status':    group.TGroup_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Nhóm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE GROUP 

#FUNCTION UPDATE DATA GROUP 
@csrf_exempt
def Data_Update_Group(request):
       try:
        if request.method == 'POST':
            GroupID = request.POST.get('GroupID')
            group = TGroup.objects.filter(TGroup_ID = GroupID)
            list_group = [{
                     'TGroup_ID':     p.TGroup_ID, 
                     'TGroup_Name':   p.TGroup_Name, 
                     'TGroup_Status': p.TGroup_Status,} for p in group]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Groups': list_group,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA GROUP 

################################################### PAGE GROUP DATA #########################

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE ATTACHMENT DATA #########################
#FUNCTION LOAD AND PROCESS DATA ATTACHMENT
def load_Attachment_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        attachs = Attachment.objects.all().order_by('-Attachment_ID')
        list_attach = []
        for attach in attachs:
            attach_data = {
                'Attachment_ID':           attach.Attachment_ID,
                'Ticket_ID':               attach.Ticket_ID.Ticket_ID,
                'Attachment_Name':         attach.Attachment_Url,
                'Attachment_User_Name':    attach.Attachment_User_Name,
                'Attachment_Date':         attach.Attachment_Date.strftime('%d/%m/%Y'),
                'Attachment_Time':         attach.Attachment_Time.strftime('%H:%M'),
                'Attachment_Status':       attach.Attachment_Status,
            }
            list_attach.append(attach_data)       

        context = { 
            'data': list_attach,
        }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('')

def load_attachment(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListFiles.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA ATTACHMENT

#FUNCTION DELETE ATTACHMENT 
@csrf_exempt
def Delete_Attachment(request):
    try:
        if request.method == 'POST':
            AttachID = request.POST.get('AttachID')
            try:
                attach = Attachment.objects.get(Attachment_ID = AttachID)
                if(attach):
                    relative_path = 'Asset/Attachment-Upload/' + attach.Attachment_Url
                    absolute_path = os.path.join(settings.BASE_DIR, settings.STATICFILES_DIRS[0], relative_path)
                    if absolute_path:  # Kiểm tra xem file có tồn tại hay không
                        os.remove(absolute_path)  # Xóa file
                    attach.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!' })
                else:
                    return JsonResponse({'success': False, 'message': 'Group ID ' + AttachID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + AttachID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE ATTACHMENT 

#FUNCTION UPDATE ATTACHMENT 
@csrf_exempt
def Update_Attachment(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            AttachID = request.POST.get('AttachID')       
            attachName = request.POST.get('AttachName')       
            attachment = Attachment.objects.get(Attachment_ID = AttachID)
            if attachment:
                if attachName:
                    relative_path = 'Asset/Attachment-Upload/' + attachment.Attachment_Url
                    absolute_path = os.path.join(settings.BASE_DIR, settings.STATICFILES_DIRS[0], relative_path)
                    if absolute_path:  # Kiểm tra xem file có tồn tại hay không
                        file_dir, old_filename = os.path.split(absolute_path)
                        new_file_path = os.path.join(file_dir, attachName)
                        os.rename(absolute_path, new_file_path)
                    attachment.Attachment_Url   = attachName
                    attachment.Attachment_Status = status
                    attachment.save()
                else:
                    attachment.Attachment_Url = attachment.Attachment_Url
                    attachment.Attachment_Status = status
                    attachment.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Attachment_ID':        attachment.Attachment_ID,
                    'Attachment_Name':      attachment.Attachment_Url,                   
                    'Attachment_Status':    attachment.Attachment_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Đính Kèm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ATTACHMENT 

#FUNCTION UPDATE DATA ATTACHMENT 
@csrf_exempt
def Data_Update_Attachment(request):
       try:
        if request.method == 'POST':
            AttachID = request.POST.get('AttachID')
            attachment = Attachment.objects.filter(Attachment_ID = AttachID)
            list_attach = [{
                     'Attachment_ID':     p.Attachment_ID, 
                     'Attachment_Name':   p.Attachment_Url, 
                     'Attachment_Status': p.Attachment_Status,} for p in attachment]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Attachment': list_attach,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA ATTACHMENT 

################################################### PAGE ATTACHMENT DATA #########################

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE ASSIGN DATA #######################  
#FUNCTION LOAD AND PROCESS DATA ASSIGN
def load_Assign_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        assigns = Assign_User.objects.all().order_by('-Assign_ID')
        list_assign = []
        for assign in assigns:
            assign_data = {
                'Assign_ID':                assign.Assign_ID,
                'ID_user':                  assign.ID_user.ID_user,
                'UserName':                 assign.ID_user.FullName,
                'TGroup_ID':                assign.TGroup_ID.TGroup_Name,
                'Assign_User_ID':           assign.Assign_User_ID,
                'Assign_User_Name':         assign.Assign_User_Name ,
                'Assign_User_Date':         assign.Assign_User_Date.strftime('%d/%m/%Y'),
                'Assign_User_Time':         assign.Assign_User_Time.strftime('%H:%M'),
                'Assign_User_Status':       assign.Assign_User_Status,
            }
            list_assign.append(assign_data)       
        ########################### TGROUP DATA###########################
        tgroups = TGroup.objects.filter(TGroup_Status = True)
        list_tgroups = [{'TGroup_ID': tgroup.TGroup_ID, 'TGroup_Name': tgroup.TGroup_Name} for tgroup in tgroups]

        context = { 
            'data': list_assign,
            'group': list_tgroups,
        }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_assign(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListAssign.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA ASSIGN

#FUNCTION CREATE ASSIGN 
@csrf_exempt
def Create_Assign(request):
    try:
        if request.method == 'POST':           
            assigns = []
            group = request.POST.get('group')
            users = request.POST.get('users')
            userID =  request.session['UserInfo']['ID_user']
            Name =  request.session['UserInfo']['FullName']
                  
            tgroup  =TGroup.objects.get(TGroup_ID = group)
            assign_users =json.loads(users)
            for u in assign_users:
                user = Users.objects.get(ID_user = u)
                check_assign = Assign_User.objects.filter(ID_user = u, TGroup_ID = group).first()
                if check_assign:
                    if (check_assign.Assign_User_Status == True): 
                        Status_Create = 'Exits'
                    else:
                        check_assign.Assign_User_Status = True
                        check_assign.save()
                        Status_Create = 'Update'
                    assign_item = {
                            'Assign_ID':                check_assign.Assign_ID,
                            'ID_user':                  check_assign.ID_user.ID_user,
                            'UserName':                 check_assign.ID_user.FullName,
                            'TGroup_ID':                check_assign.TGroup_ID.TGroup_ID,
                            'TGroup_Name':              check_assign.TGroup_ID.TGroup_Name,
                            'Assign_User_ID':           check_assign.Assign_User_ID,
                            'Assign_User_Name':         check_assign.Assign_User_Name,
                            'Assign_User_Date':         check_assign.Assign_User_Date.strftime('%d/%m/%Y'),
                            'Assign_User_Time':         check_assign.Assign_User_Time.strftime('%H:%M'),
                            'Assign_User_Status':       check_assign.Assign_User_Status,
                            'Status_Create':            Status_Create
                        }
                    assigns.append(assign_item)
                else:
                    assign = Assign_User.objects.create(
                        ID_user  = user, 
                        TGroup_ID  = tgroup, 
                        Assign_User_ID = userID,
                        Assign_User_Name  = Name, 
                        Assign_User_Date  = datetime.datetime.now().date(),
                        Assign_User_Time  = datetime.datetime.now().time(),
                        Assign_User_Status = 1,
                    )
                    assign.save()
                    assignID = assign.Assign_ID
                    assign_item = {
                        'Assign_ID':                assignID,
                        'ID_user':                  assign.ID_user.ID_user,
                        'UserName':                 assign.ID_user.FullName,
                        'TGroup_ID':                assign.TGroup_ID.TGroup_ID,
                        'TGroup_Name':              assign.TGroup_ID.TGroup_Name,
                        'Assign_User_ID':           assign.Assign_User_ID,
                        'Assign_User_Name':         assign.Assign_User_Name,
                        'Assign_User_Date':         assign.Assign_User_Date.strftime('%d/%m/%Y'),
                        'Assign_User_Time':         assign.Assign_User_Time.strftime('%H:%M'),
                        'Assign_User_Status':       assign.Assign_User_Status,
                        'Status_Create':            'Insert'
                    }
                    assigns.append(assign_item)
            context = {
                'success': True,
                'message': 'Công Ty Tạo Thành Công!',
                'assigns': assigns
            }
            return JsonResponse(context)
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE ASSIGN 

#FUNCTION DELETE ASSIGN 
@csrf_exempt
def Delete_Assign(request):
    try:
        if request.method == 'POST':
            AssignID = request.POST.get('AssignID')
            try:
                assign = Assign_User.objects.get(Assign_ID = AssignID)
                if(assign):
                    assign.delete()
                    return JsonResponse({'success': True,'message': 'Xóa thành công!' })
                else:
                    return JsonResponse({'success': False, 'message': 'Assign ID ' + AssignID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + AssignID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE ASSIGN 

#FUNCTION UPDATE ATTACHMENT 
@csrf_exempt
def Update_Assign(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            AsignID = request.POST.get('AsignID')         
            assign = Assign_User.objects.get(Assign_ID = AsignID)
            if assign:
                assign.Assign_User_Status = status
                assign.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Assign_ID':             assign.Assign_ID,                
                    'Assign_User_Status':    assign.Assign_User_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Phân Công Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ATTACHMENT 

#FUNCTION AUTOCOMPLETE
@csrf_exempt
def user_autocomplete(request):
    if request.method == 'POST':
        term = request.POST.get('term')
        group = request.POST.get('group')
        assign = Assign_User.objects.filter(TGroup_ID = group , Assign_User_Status = True).values('ID_user')
        users = Users.objects.filter(FullName__icontains=term, User_Type__lt = 2, User_Status =  True).exclude(ID_user__in=assign).values_list('ID_user', 'FullName')
        result = [f'{user[0]} - {user[1]}' for user in users]
        return JsonResponse(result, safe=False)
#FUNCTION AUTOCOMPLETE

################################################### PAGE ASSIGN DATA #######################  
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE MENU DATA #######################  
#FUNCTION LOAD AND PROCESS DATA MENU
def load_Menu_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        menus = Menu.objects.all().order_by('-Menu_ID')
        list_menu = []
        for menu in menus:
            menu_data = {
                'Menu_ID':                  menu.Menu_ID,
                'Menu_Name':                menu.Menu_Name,
                'Menu_Adress':              menu.Menu_Adress,
                'Menu_Icon':                menu.Menu_Icon,
                'Menu_Level':               menu.Menu_Level,
                'Menu_CreateID':            menu.Menu_CreateID,
                'Menu_CreateBy':            menu.Menu_CreateBy ,
                'Menu_Date':                menu.Menu_Date.strftime('%d/%m/%Y'),
                'Menu_Time':                menu.Menu_Time.strftime('%H:%M'),
                'Menu_Status':              menu.Menu_Status,
            }
            list_menu.append(menu_data)       
        context = { 
            'data': list_menu,
        }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_menu(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_Menu.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA MENU

#FUNCTION CREATE MENU 
@csrf_exempt
def Create_Menu(request):
    try:
        if request.method == 'POST':           
            Menu_Name = request.POST.get('Menu_Name')
            Menu_Adress = request.POST.get('Menu_Adress')
            Menu_Icon = request.POST.get('Menu_Icon')
            Menu_Level = request.POST.get('Menu_Level')
            Menu_status = request.POST.get('Menu_status')

            userID =  request.session['UserInfo']['ID_user']
            Name =  request.session['UserInfo']['FullName']

            menu = Menu.objects.create(
                        Menu_Name  = Menu_Name, 
                        Menu_Adress  = Menu_Adress, 
                        Menu_Icon = Menu_Icon,
                        Menu_Level  = Menu_Level, 
                        Menu_CreateID  = userID, 
                        Menu_CreateBy  = Name, 
                        Menu_Date  = datetime.datetime.now().date(),
                        Menu_Time  = datetime.datetime.now().time(),
                        Menu_Status = Menu_status,
                    )
            menu.save()    
            Menu_ID = menu.Menu_ID
            return JsonResponse({
                    'success': True,
                    'message': 'Menu Tạo Thành Công!',
                    'Menu_ID':             Menu_ID,
                    'Menu_Name':           menu.Menu_Name,
                    'Menu_Adress':         menu.Menu_Adress,
                    'Menu_Icon':           menu.Menu_Icon,
                    'Menu_Level':          menu.Menu_Level,
                    'Menu_CreateID':       menu.Menu_CreateID,
                    'Menu_CreateBy':       menu.Menu_CreateBy,
                    'Menu_Date':           menu.Menu_Date.strftime('%d/%m/%Y'),
                    'Menu_Time':           menu.Menu_Time.strftime('%H:%M'),
                    'Menu_Status':         menu.Menu_Status,
                })          
            # context = {
            #     'success': True,
            #     'message': 'Menu Tạo Thành Công!',
            #     'menu': menu
            # }
            # return JsonResponse(context)
        else:
            # response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE ASSIGN 

#FUNCTION DELETE MENU 
@csrf_exempt
def Delete_Menu(request):
    try:
        if request.method == 'POST':
            MenuID = request.POST.get('MenuID')
            try:
                menu = Menu.objects.get(Menu_ID = MenuID)
                if(menu):
                    menu.delete()
                    return JsonResponse({'success': True,'message': 'Xóa thành công!' })
                else:
                    return JsonResponse({'success': False, 'message': 'Menu ID ' + MenuID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + MenuID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE USERS 

#FUNCTION UPDATE GROUP 
@csrf_exempt
def Update_Menu(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            MenuID = request.POST.get('MenuID')       
            MenuName = request.POST.get('MenuName')       
            MenuAdress = request.POST.get('MenuAdress')       
            MenuIcon = request.POST.get('MenuIcon')       
            MenuLevel = request.POST.get('MenuLevel')       
            menu = Menu.objects.get(Menu_ID = MenuID)
            if menu:
                if MenuName:
                    menu.Menu_Name = MenuName
                    menu.Menu_Adress = MenuAdress
                    menu.Menu_Icon   = MenuIcon
                    menu.Menu_Level  = MenuLevel
                    menu.Menu_Status = status
                    menu.save()
                else:
                    menu.Menu_Name   = menu.Menu_Name
                    menu.Menu_Adress = menu.Menu_Adress
                    menu.Menu_Icon   = menu.Menu_Icon
                    menu.Menu_Level  = menu.Menu_Level
                    menu.Menu_Status = status
                    menu.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Menu_ID':             menu.Menu_ID,
                    'Menu_Name':           menu.Menu_Name,
                    'Menu_Adress':         menu.Menu_Adress,
                    'Menu_Icon':           menu.Menu_Icon,
                    'Menu_Level':          menu.Menu_Level,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Nhóm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE GROUP 

#FUNCTION UPDATE DATA GROUP 
@csrf_exempt
def Data_Update_Menu(request):
       try:
        if request.method == 'POST':
            MenuID = request.POST.get('MenuID')
            menu = Menu.objects.filter(Menu_ID = MenuID)
            if menu:
                list_menu = [{
                        'Menu_ID':             p.Menu_ID,
                        'Menu_Name':           p.Menu_Name,
                        'Menu_Adress':         p.Menu_Adress,
                        'Menu_Icon':           p.Menu_Icon,
                        'Menu_Level':          p.Menu_Level,
                        'Menu_CreateID':       p.Menu_CreateID,
                        'Menu_Status':         p.Menu_Status,
                        } for p in menu]
                context = { 
                    'success': True,
                    'message': 'Lấy Data Thành Công',
                    'Menus': list_menu,}
                data = json.dumps(context, ensure_ascii=False)
                return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA GROUP 

################################################### PAGE MENU DATA ####################### 
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE USERS DATA #########################
#FUNCTION LOAD AND PROCESS DATA USERS
def load_User_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### Check Role###########################    
        # type = userinfo['Acc_type']
        # status = check_role(userinfo['ID_user'], type)
        # if status == False:
        #     # return redirect('/dashboard/')
        #     return HttpResponseRedirect('/dashboard/')
        # else:
            ########################### User DATA###########################    
            user = Users.objects.all().order_by('-ID_user')       
            
            list_user = []
            for u in user:
                cop = Company.objects.get(Company_ID = u.Company_ID)
                user_data = {
                    'Avatar':                   u.Avatar if u.Avatar else 'Null',
                    'img'   :                   u.displayName[0] + u.FullName[0],
                    'ID_user':                  u.ID_user,
                    'Company_Name':             cop.Company_Name,
                    'Mail':                     u.Mail,
                    'FullName':                 u.FullName,
                    'User_Type':                u.User_Type,
                    'Acc_Type':                 u.Acc_Type,
                    'Jobtitle':                 u.Jobtitle if u.Jobtitle else 'No Data',
                    'Birthday':                 u.Birthday.strftime('%d/%m/%Y'),
                    'Address':                  u.Address if u.Address else 'No Data',
                    'Phone':                    u.Phone if u.Phone else 'No Data',
                    'ID_Create':                u.ID_Create if u.ID_Create else 'No Data',
                    'Name_Create':              u.Name_Create if u.Name_Create else 'No Data',
                    'Date_Create':              u.Date_Create.strftime('%d/%m/%Y'),
                    'Time_Create':              u.Time_Create.strftime('%H:%M'),
                    'User_Type':                u.User_Type,
                    'User_Status':              u.User_Status,
                    # 'Online_Status': 'Online' if u.is_online else 'Offline',
                }
                list_user.append(user_data)       

            context = { 
                'data': list_user,
            }
            return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')
    
def load_User(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListUser.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA USERS

@csrf_exempt
def Company_Json(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            ########################### COMPANY DATA###########################    
            companys = Company.objects.filter(~Q(Company_Name = 'No Data') , Company_Status = True)
            list_company = []
            for company in companys:
                company_data = {
                    'Company_ID':           company.Company_ID,
                    'Company_Name':         company.Company_Name,
                }
                list_company.append(company_data)       
            context = { 
                'success': True, 
                'data': list_company,}
            return JsonResponse(context)
            
            # return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE USERS 
@csrf_exempt
def Create_User(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('pass')
            fullname = request.POST.get('fullname')
            company = request.POST.get('company')
            phone = request.POST.get('phone')
            address = request.POST.get('address')
            birthday = request.POST.get('birthday')
            jobtitle = request.POST.get('jobtitle')
            role = request.POST.get('role')
            status_user = request.POST.get('status')

            IDsession =  request.session['UserInfo']['ID_user']            
            Namesession =  request.session['UserInfo']['FullName'] 

            UserID = Check_IDUser()
            user = Users.objects.create(
                ID_user         = UserID,
                Mail            = email, 
                Password        = password, 
                FullName        = fullname, 
                Company_ID      = company, 
                displayName     = fullname, 
                Birthday        = birthday if birthday else datetime.date(1990, 1, 1), 
                User_Type       = role, 
                Acc_Type        = 'System',
                Address         = address, 
                Jobtitle        = jobtitle, 
                Phone           = phone if phone else 0, 
                ID_Create       = IDsession, 
                Name_Create     = Namesession, 
                Date_Create     = datetime.datetime.now().date(),
                Time_Create     = datetime.datetime.now().time(),
                User_Status     = status_user,
            )
            user.save()
            id_user = user.ID_user
            if user.User_Type == '2':
                login_role(id_user)
            elif user.User_Type == '1':
                login_role_mod(id_user)
            cop = Company.objects.get(Company_ID = company)
            return JsonResponse({
                    'success'           : True,
                    'message'           : 'Người dùng được tạo thành công!',
                    'ID_user'           : UserID,
                    'Avatar'            : user.Avatar,
                    'Mail'              : user.Mail,
                    'FullName'          : user.FullName,
                    'Company_Name'      : cop.Company_Name,
                    'Birthday'          : user.Birthday,
                    'User_Type'         : user.User_Type,
                    'Acc_Type'          : user.Acc_Type,
                    'Address'           : user.Address,
                    'Jobtitle'          : user.Jobtitle,
                    'Phone'             : user.Phone,
                    'ID_Create'         : user.ID_Create,
                    'Name_Create'       : user.Name_Create,
                    'Date_Create'       : user.Date_Create.strftime('%d/%m/%Y'),
                    'Time_Create'       : user.Time_Create.strftime('%H:%M'),
                    'User_Status'       : user.User_Status,
                    'img'               : user.displayName[0] +user.FullName[0],
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE USERS 

def login_role_mod(IDUser):
    try:
        user_create = Users.objects.get(ID_user = 'U000000001')
        assign = Users.objects.get(ID_user = IDUser)
        role_mod_exclude=[
            'ZRG',
            'ZRL',
            'ZAU',
        ]
        # Tạo điều kiện Q động cho mỗi giá trị trong role_mod_exclude
        conditions = [Q(Role_Name__icontains=value) for value in role_mod_exclude]
        # Kết hợp các điều kiện bằng toán tử OR
        combined_condition = reduce(lambda x, y: x | y, conditions)
        # Loại bỏ các bản ghi thỏa mãn điều kiện
        list_role = Role_Single.objects.exclude(combined_condition)
        # list_role = Role_Single.objects.exclude(Role_Name__in=role_mod_exclude)
        for role in list_role:
            auth_role = Authorization_User(
                ID_user = assign,
                Role_ID = role,
                Authorization_From = datetime.datetime.now().date(),
                Authorization_To = datetime.date(9999, 12, 31),
                Authorization_CreateID = user_create.ID_user,
                Authorization_CreateBy = user_create.FullName,
                Authorization_Date = datetime.datetime.now().date(),
                Authorization_Time = datetime.datetime.now().time(),
                Authorization_Status = 1
            )
            auth_role.save()
        return True
    except Exception as ex:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
            })
#FUNCTION DELETE USERS 
@csrf_exempt
def Delete_User(request):
    try:
        if request.method == 'POST':
            UserID = request.POST.get('UserID')
            try:
                user = Users.objects.get(ID_user = UserID)
                if(user):
                    user.delete()
                    return JsonResponse({'success': True,'message': 'Xóa thành công!' })
                else:
                    return JsonResponse({'success': False, 'message': 'User ID ' + UserID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + UserID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE USERS 

#FUNCTION UPDATE USERS 
@csrf_exempt
def Update_User(request):
    try:
        if request.method == 'POST':
            action = request.POST.get('Action')       
            status = request.POST.get('status')       
            UserID = request.POST.get('UserID')         
            UserType = request.POST.get('UserType')   

            FullName = request.POST.get('FullName')         
            Company_ID = request.POST.get('Company_ID')         
            Phone = request.POST.get('Phone')         
            Address = request.POST.get('Address')         
            Birthday = request.POST.get('Birthday')         
            Jobtitle = request.POST.get('Jobtitle')    

            user = Users.objects.filter(ID_user = UserID).first()
            if user:
                if(action == 'status'):
                    user.User_Status = status
                    user.save()
                elif (action == 'role'):
                    user.User_Type = UserType
                    user.save()
                elif (action == 'update'):
                    user.FullName   = FullName
                    user.Company_ID   = Company_ID
                    user.Phone      = Phone
                    user.Address    = Address
                    user.Birthday   = Birthday if Birthday else datetime.date(1990, 1, 1)
                    user.Jobtitle   = Jobtitle
                    user.User_Type  = UserType
                    user.User_Status = status
                    user.save()
                
                
                if Company_ID:
                    comp = Company.objects.get(Company_ID = Company_ID)
                else:
                    comp = Company.objects.get(Company_ID = user.Company_ID)

                if isinstance(user.Birthday, datetime.date):
                    data_birthday = user.Birthday.strftime('%d/%m/%Y')
                else:
                    data_birthday = datetime.datetime.strptime(user.Birthday, '%Y-%m-%d').strftime('%d/%m/%Y')

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'ID_user':             user.ID_user,                
                    'User_Status':         user.User_Status,
                    'FullName':            user.FullName,
                    'Company_Name':        comp.Company_Name,
                    # 'Birthday':            datetime.datetime.strptime(user.Birthday, '%Y-%m-%d').strftime('%d/%m/%Y'),
                    'Birthday':            data_birthday,
                    'Address':             user.Address,
                    'Jobtitle':            user.Jobtitle,
                    'Phone':               user.Phone,
                    'User_Type':           user.User_Type,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Phân Công Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE USERS 

#FUNCTION UPDATE PROFILE 
def Update_Profile(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            id = userinfo['ID_user']
            data_user =  Users.objects.get(ID_user = id)
            user_profile = {
                'ID_user'       : data_user.ID_user,
                'Mail'          : data_user.Mail,
                'FullName'      : data_user.FullName,
                'Phone'         : data_user.Phone,
                'Address'       : data_user.Address,
                'Company_ID'    : data_user.Company_ID,
                'Birthday'      : data_user.Birthday.strftime('%Y-%m-%d'),
                'Jobtitle'      : data_user.Jobtitle,
                'User_Type'     : data_user.User_Type,
                'User_Status'   : data_user.User_Status,
            }

            comp = Company.objects.all()
            company = ({
                'Company_ID'  : c.Company_ID,
                'Company_Name': c.Company_Name,
            } for c in comp)

            context = {
                'user_profile': user_profile,
                'company'     : company,
            }

            return render(request, 'Ticket_Profile.html', context)
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE PROFILE 

#FUNCTION UPDATE DATA USERS 
@csrf_exempt
def Data_Update_User(request):
       try:
        if request.method == 'POST':
            UserID = request.POST.get('UserID')
            user = Users.objects.filter(ID_user = UserID).first()
            list_user = {
                     'ID_user':           user.ID_user, 
                     'Mail':              user.Mail, 
                     'FullName':          user.FullName,
                     'Company_ID':        user.Company_ID,
                     'Phone':             user.Phone,
                     'Address':           user.Address,
                     'Birthday':          user.Birthday.strftime('%Y-%m-%d'),
                     'Jobtitle':          user.Jobtitle,
                     'User_Type':         user.User_Type,
                     'User_Status':       user.User_Status,
                     }
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Users': list_user,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA USERS 

#FUNCTION CHECK EMAIL USERS 
@csrf_exempt
def Check_Email(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            try:
                user = Users.objects.filter(Mail = email).first()
                if(user):
                    if user.User_Status == False:
                         return JsonResponse({'success': False,'message': 'UserID:' + user.ID_user +'\n - Mail: ' + user.Mail + ' đã tồn tại và chưa được kích hoạt' })
                    else:
                        return JsonResponse({'success': False,'message': 'UserID:' + user.ID_user +'\n - Mail: ' + user.Mail + ' đã tồn tại' })
                else:
                    return JsonResponse({'success': True, 'message': 'Tạo Người Dùng'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Users.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK EMAIL USERS 

@csrf_exempt
def Check_Company(request):
    try:
        if request.method == 'POST':
            CompanyID = request.POST.get('CompanyID')
            comp = Company.objects.get(Company_ID = CompanyID)
            if comp:
                if 'No Data' in comp.Company_Name:
                    return JsonResponse({
                        'success': False
                    })
                else:
                   return JsonResponse({
                        'success': True
                    })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })

def generate_random_string():
    # Sử dụng module string để lấy tất cả các ký tự chữ và số
    characters = string.ascii_letters + string.digits
    # Sử dụng hàm random.choice để chọn ngẫu nhiên 10 ký tự từ danh sách
    random_string = ''.join(random.choice(characters) for i in range(20))
    return random_string

def send_email_reset_pass(User, New_pass):
    try:
        subject = '[SDP] Reset password Account' 
        data = {
        'ID_user': User.ID_user,
        'FullName': User.FullName,
        'Password': New_pass,       
        }
        message = render_to_string('Email_Reser_pass.html', context=data)
        
        from_email = settings.MICROSOFT_EMAIL
        recipient_list = [User.Mail]
        # cc_list = ['cc@example.com']
        send_mail(
            subject=subject,
            message=message,
            from_email=from_email,
            recipient_list=recipient_list,
            # cc=cc_list,
            fail_silently=False,  # Thông báo lỗi nếu có vấn đề khi gửi email
            html_message=message,  # Sử dụng nội dung HTML
        )
        return True
    except Exception as ex:
        return False

@csrf_exempt
def Reset_pass_User(request):
    try:
        if request.method == 'POST':
            IDUser = request.POST.get('ID_User')
            New_Pass = request.POST.get('New_Pass')
            user = Users.objects.get(ID_user = IDUser)
            if user:
                if New_Pass:
                    pass_new = New_Pass
                else:
                    pass_new = generate_random_string()
                if pass_new:
                    user.Password = pass_new
                    user.save()
                    send_email_reset_pass(user, pass_new)
                    return JsonResponse({
                        'success': True,
                        'message': 'Reset Password thành công',
                    })            
                else:
                   return JsonResponse({
                        'success': False,
                        'message': 'Lỗi Không xin được password random',
                    })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })

@csrf_exempt
def Check_Status_User(request):
    try:
        if request.method == 'POST':
            IDUser = request.session.get('UserInfo')['ID_user']
            user = Users.objects.get(ID_user = IDUser)
            if user:
                if user.Password:
                    return JsonResponse({
                        'success': True,
                    })            
                else:
                   return JsonResponse({
                        'success': False,
                    })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
@csrf_exempt
def Update_Pass_By_User(request):
    try:
        if request.method == 'POST':
            IDUser = request.session.get('UserInfo')['ID_user']
            Status = request.POST.get('Status')
            OldPass = request.POST.get('OldPass')
            NewPass = request.POST.get('NewPass')
            user = Users.objects.get(ID_user = IDUser)
            if user:
                if Status == 'True':
                    if user.Password == OldPass:
                        user.Password = NewPass
                        user.save()
                        return JsonResponse({
                            'success': True,
                            'message': 'Cập nhật mật khẩu thành công.'
                        })            
                    else:
                        return JsonResponse({
                                'success': False,
                                'message': 'Mật Khẩu cũ chưa đúng, vui lòng kiểm tra lại.'
                            })
                else:
                    user.Password = NewPass
                    user.save()
                    return JsonResponse({
                        'success': True,
                        'message': 'Cập nhật mật khẩu thành công.'
                    })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })

def Update_Pass_User(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_Change_Pass.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA USERS
################################################### PAGE USERS DATA #########################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE IMPORT - EXPORT DATA #########################
@csrf_exempt
def import_excel_user(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        date_birth = data['Birthday']
                        # birth = datetime.datetime(date_birth).strptime('%Y-%m-%d')
                        birth = datetime.datetime.strptime(date_birth, '%d/%m/%Y')
                        new_format_str = birth.strftime('%Y-%m-%d')
                        check_id = Users.objects.filter(Mail = data['Mail'] )
                        if not check_id.exists():
                            UserID = Check_IDUser()
                            company = Company.objects.get(Company_Name = data['Company'])
                            user = Users.objects.get(ID_user = 'U000000001')
                            users = Users.objects.create(
                                    ID_user         = UserID,
                                    Mail            = data['Mail'], 
                                    FullName        = data['Full Name'], 
                                    Company_ID      = company.Company_ID, 
                                    displayName     = data['Display Name'], 
                                    Birthday        = new_format_str if new_format_str else datetime.date(1990, 1, 1), 
                                    User_Type       = 2, 
                                    Acc_Type        = 'System',
                                    Address         = data['Address'], 
                                    Jobtitle        = data['Jobtitle'], 
                                    Phone           = data['Phone'] if data['Phone'] else 0, 
                                    ID_Create       = user.ID_user, 
                                    Name_Create     = user.FullName, 
                                    Date_Create     = datetime.datetime.now().date(),
                                    Time_Create     = datetime.datetime.now().time(),
                                    User_Status     = True,
                            )
                            users.save()
                            id_user = users.ID_user
                            if users.User_Type == 2:
                                login_role(id_user)
                            elif users.User_Type == 1:
                                login_role_mod(id_user)
                        else:
                            id_list_exits.append(data['Mail'])
                        # user_data = Users.objects.filter(ID_user__in = id_list_exits)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
def export_excel_user(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        # data_json = request.GET.get('data')
        data_json = request.body.decode('utf-8')
        json_user = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_user:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 20,
                'B': 35,
                'C': 20,
                'D': 35,
                'E': 20,
                'F': 10,
                'G': 30,
                'H': 12,
                'I': 30,
                'J': 12,
                'K': 20,
                'L': 25,
                'M': 12,
                'N': 12,
                'O': 10,
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['User_ID', 'Email', 'Full Name', 'Company', 'Role', 'User Type',
                      'Job Title', 'Birth Day', 'Address', 'Phone', 'Created ID',
                      'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_user:
                if dt['User_Type'] == 0:
                    user_type = 'Administrator'
                elif dt['User_Type'] == 1:
                    user_type = 'Mod'
                else:
                    user_type = 'User Member'
                list_data = [
                    dt['ID_user'],
                    dt['Mail'],
                    dt['FullName'],
                    dt['Company_Name'],
                    user_type,
                    dt['Acc_Type'],
                    dt['Jobtitle'],
                    dt['Birthday'],
                    dt['Address'],
                    dt['Phone'],
                    dt['ID_Create'],
                    dt['Name_Create'],
                    dt['Date_Create'],
                    dt['Time_Create'],
                    dt['User_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=User_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def export_excel_ticket(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        # data_json = request.POST.get('data')
        data_json = request.body.decode('utf-8')
        json_ticket = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_ticket:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 12,
                'B': 40,
                'C': 12,
                'D': 30,
                'E': 10,
                'F': 12,
                'G': 20,
                'H': 10,
                'I': 12,
                'J': 20,
                'K': 12,
                'L': 12,
                'M': 10,
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Ticket_ID', 'Title', 'Company ID', 'Company', 'Group', 
                      'User Support', 'User Support Name',
                      'Ticket Type', 'Created ID','Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_ticket:
                if dt['Ticket_Status'] == 0:
                    status = 'Hoàn Thành'
                elif dt['Ticket_Status'] == 1:
                    status = 'Đang làm'
                elif dt['Ticket_Status'] == 2:
                    status = 'Đang Treo'
                else:
                    status = 'Hủy'

                if dt['Ticket_Type'] == 0:
                    ticket_type = 'Sự Cố'
                else:
                    ticket_type = 'Hỗ Trợ'
                list_data = [
                    dt['Ticket_ID'],
                    dt['Ticket_Title'],
                    dt['Company_ID'],
                    dt['Company_Name'],
                    dt['Group_Name'],
                    dt['Ticket_User_Asign'],
                    dt['Ticket_Name_Asign'],
                    ticket_type,
                    dt['ID_user'],
                    dt['Ticket_User_Name'],
                    dt['Ticket_Date'],
                    dt['Ticket_Time'],
                    status,
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Ticket_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def export_excel_company(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        # data_json = request.POST.get('data')
        data_json = request.body.decode('utf-8')
        json_company = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_company:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 12,
                'B': 35,
                'C': 20,
                'D': 12,
                'E': 12,
                'F': 12,
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Company_ID', 'Company_Name',
                    #   'Created ID',
                      'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_company:
                list_data = [
                    dt['Company_ID'],
                    dt['Company_Name'],
                    # dt['ID_user'],
                    dt['Company_User_Name'],
                    dt['Company_Date'],
                    dt['Company_Time'],
                     dt['Company_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Company_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def import_excel_company(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        check_company = Company.objects.filter(Company_Name = data['Company_Name'])
                        if not check_company.exists():
                            user = Users.objects.get(ID_user = 'U000000001')
                            company = Company.objects.create(                                
                                    Company_Name    = data['Company_Name'],
                                    ID_user         = user, 
                                    Company_User_Name     = user.FullName, 
                                    Company_Date     = datetime.datetime.now().date(),
                                    Company_Time     = datetime.datetime.now().time(),
                                    Company_Status     = True,
                            )
                            company.save()
                        else:
                            id_list_exits.append( data['Company_Name'])
                        # user_data = Users.objects.filter(ID_user__in = id_list_exits)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
def export_excel_assgin(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        data_json = request.body.decode('utf-8')
        json_assign = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_assign:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 10,
                'B': 15,
                'C': 25,
                'D': 10,
                'E': 20,
                'F': 15,
                'G': 20,
                'H': 12, 
                'I': 12, 
                'J': 12,
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Assign ID', 'ID user', 'FullName',
                     'Group ID','Group Name',
                     'Created ID', 'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_assign:
                group = TGroup.objects.get(TGroup_Name = dt['TGroup_ID'])
                list_data = [
                    dt['Assign_ID'],
                    dt['ID_user'],
                    dt['UserName'],
                    group.TGroup_ID,
                    group.TGroup_Name,
                    dt['Assign_User_ID'],
                    dt['Assign_User_Name'],
                    dt['Assign_User_Date'],
                    dt['Assign_User_Time'],
                    dt['Assign_User_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Assign_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def import_excel_assign(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        assign_user = Users.objects.filter(ID_user = data['ID user'])
                        if not assign_user.exists():
                            error_user = 'MNV: ' + str(data['ID user']) + ' - Mã nhân viên không tồn tại\n'
                            id_list_exits.append(error_user)
                            continue

                        if  data['Group ID']:
                            # gr_id = data['Group ID']
                            group_id = TGroup.objects.filter(TGroup_ID = data['Group ID'])
                            if not group_id.exists():
                                error_user = 'MN: ' + str(data['Group ID']) + ' - Mã Group không tồn tại\n'
                                id_list_exits.append(error_user)
                                continue
                            assign = Assign_User.objects.filter(ID_user = data['ID user'], TGroup_ID = group_id.first().TGroup_ID)
                        elif  data['Group Name']:
                            GroupID = TGroup.objects.filter(TGroup_Name = data['Group Name']).first()
                            if GroupID:
                                # gr_id = GroupID.TGroup_ID
                                group_id = TGroup.objects.filter(TGroup_ID = GroupID.TGroup_ID)
                                if not group_id.exists():
                                    error_user = 'MN: ' + str(data['Group ID']) + ' - Mã Group không tồn tại\n'
                                    id_list_exits.append(error_user)
                                    continue
                                assign = Assign_User.objects.filter(ID_user = data['ID user'], TGroup_ID =  group_id.first().TGroup_ID)
                            else:
                                error_group = data['Group Name'] + ' - Group Name không tồn tại\n'
                                id_list_exits.append(error_group)
                                continue
                        else:
                            context = { 
                                'success': False,
                                'message': 'Chưa nhập đủ thông tin (Group ID hoặc Group Name)'
                                }
                            return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                    
                        if not assign.exists():
                            user = Users.objects.get(ID_user = 'U000000001')
                            assg = Assign_User.objects.create(                                
                                    ID_user             = assign_user.first(),
                                    TGroup_ID           = group_id.first(), 
                                    Assign_User_ID      = user.ID_user, 
                                    Assign_User_Name    = user.FullName, 
                                    Assign_User_Date     = datetime.datetime.now().date(),
                                    Assign_User_Time     = datetime.datetime.now().time(),
                                    Assign_User_Status     = True,
                            )
                            assg.save()
                        else:
                            if  data['Group ID']:
                                error_data = 'MNV: ' + str(data['ID user']) + ' - MN: ' + str(data['Group ID']) + ' Đã tồn tại\n'
                            elif  data['Group ID']:
                                error_data = 'MNV: ' + str(data['ID user']) + ' - TN: ' + data['Group Name'] + ' Đã tồn tại\n'
                            else:
                                error_data = 'MNV: ' + str(data['ID user']) + ' Đã tồn tại\n'
                            id_list_exits.append(error_data)
                        # user_data = Users.objects.filter(ID_user__in = id_list_exits)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
def export_excel_grouprole(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        data_json = request.body.decode('utf-8')
        json_grouprole = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_grouprole:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 10,
                'B': 35,
                'C': 10,
                'D': 20,
                'E': 25,
                'F': 15,
                'G': 20,
                'H': 12, 
                'I': 12, 
                'J': 12, 
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Group ID', 'Group Name', 'Menu ID', 'Menu Name', 'Link',
                     'Created ID', 'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_grouprole:
                list_data = [
                    dt['Role_Group_ID'],
                    dt['Role_Group_Name'],
                    dt['Menu_ID'],
                    dt['Menu_Name'],
                    dt['Role_Group_Address'],
                    dt['ID_user'],
                    dt['Role_Group_CreateBy'],
                    dt['Role_Group_Date'],
                    dt['Role_Group_Time'],
                    dt['Role_Group_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=GroupRole_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def import_excel_grouprole(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        menu = Menu.objects.filter(Menu_ID = data['Menu ID'])
                        if not menu.exists():
                            error_menu = 'MNV: ' + str(data['Menu ID']) + ' - Mã Menu không tồn tại\n'
                            id_list_exits.append(error_menu)
                            continue
                        check_group = Role_Group.objects.filter(Role_Group_Name = data['Group Name'], Menu_ID = data['Menu ID'])
                        if not check_group.exists():
                            user = Users.objects.get(ID_user = 'U000000001')                                                    
                            grouprole = Role_Group.objects.create(                                
                                    ID_user             = user,
                                    Menu_ID             = menu.first(), 
                                    Role_Group_Name     = data['Group Name'], 
                                    Role_Group_Address  = data['Link'], 
                                    Role_Group_CreateBy = user.FullName, 
                                    Role_Group_Date     = datetime.datetime.now().date(),
                                    Role_Group_Time     = datetime.datetime.now().time(),
                                    Role_Group_Status     = True,
                            )
                            grouprole.save()
                        else:
                            error_data = 'Group Name: ' + data['Group Name'] + ' - Menu: ' + str(data['Menu ID']) + ' Đã tồn tại\n'
                            id_list_exits.append(error_data)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
def export_excel_role(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        data_json = request.body.decode('utf-8')
        json_grouprole = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_grouprole:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 10,
                'B': 55,
                'C': 10,
                'D': 40,
                'E': 12,
                'F': 20,
                'G': 20,
                'H': 12, 
                'I': 12, 
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Role ID', 'Role Name', 'Group ID', 'Group Name',
                     'Created ID', 'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_grouprole:
                list_data = [
                    dt['Role_ID'],
                    dt['Role_Name'],
                    dt['Role_Group_ID'],
                    dt['Role_Group_Name'],
                    dt['ID_user'],
                    dt['Role_CreateBy'],
                    dt['Role_Date'],
                    dt['Role_Time'],
                    dt['Role_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Role_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
def import_excel_role(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        group = Role_Group.objects.filter(Role_Group_ID = data['Group ID'])
                        if not group.exists():
                            error_menu = 'Group: ' + str(data['Group ID']) + ' - Mã Group không tồn tại\n'
                            id_list_exits.append(error_menu)
                            continue
                        check_role = Role_Single.objects.filter(Role_Name = data['Role Name'], Role_Group_ID = data['Group ID'])
                        if not check_role.exists():
                            user = Users.objects.get(ID_user = 'U000000001')                                                    
                            role = Role_Single.objects.create(                                
                                    ID_user             = user,
                                    Role_Group_ID       = group.first(), 
                                    Role_Name           = data['Role Name'], 
                                    Role_CreateBy       = user.FullName, 
                                    Role_Date           = datetime.datetime.now().date(),
                                    Role_Time           = datetime.datetime.now().time(),
                                    Role_Status         = True,
                            )
                            role.save()
                        else:
                            error_data = 'Role Name: ' + data['Role Name'] + ' - Group: ' + str(data['Group ID']) + ' Đã tồn tại\n'
                            id_list_exits.append(error_data)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
def export_excel_authorize(request):
    try:
        # Nhận dữ liệu từ tham số truy vấn data
        data_json = request.body.decode('utf-8')
        json_grouprole = json.loads(data_json)
        data = []
        # Tạo một workbook và một worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if json_grouprole:
            # Thiết lập độ rộng cho các cột
            column_widths = {
                'A': 15,
                'B': 12,
                'C': 25,
                'D': 10,
                'E': 55,
                'F': 18,
                'G': 15,
                'H': 12, 
                'I': 20, 
                'J': 12, 
                'K': 15, 
                'L': 12, 
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # In đậm header
            header = ['Authorization ID', 'User ID', 'FullName', 'Role ID', 'Role Name', 'Authorization From','Authorization To', 
                      'Created ID', 'Created Name', 'Created Date', 'Created Time', 'Status']
            for col_num, header_value in enumerate(header, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_value
                cell.font = openpyxl.styles.Font(bold=True)

            for dt in json_grouprole:
                list_data = [
                    dt['Authorization_ID'],
                    dt['ID_user'],
                    dt['FullName'],
                    dt['Role_ID'],
                    dt['Role_Name'],
                    dt['Authorization_From'],
                    dt['Authorization_To'],
                    dt['Authorization_CreateID'],
                    dt['Authorization_CreateBy'],
                    dt['Authorization_Date'],
                    dt['Authorization_Time'],
                    dt['Authorization_Status'],
                ]
                data.append(list_data)

            if data:
                # Thêm dữ liệu vào worksheet
                for row_num, row_data in enumerate(data, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            top=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin')
                        )

                # Tạo tên file Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Authorize_data_list.xlsx'

                # Lưu workbook vào HttpResponse
                workbook.save(response)
                return response
            else:
                return JsonResponse({'success': False, 'message': 'No data'})
        else:
            return JsonResponse({'success': False, 'message': 'No data'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def load_Data_Authorize_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### USERS DATA###########################   
        # users = Users.objects.all().filter(User_Status = True).order_by('-ID_user')
        auth = Authorization_User.objects.all().order_by('-Authorization_ID')
        list_auth = []
        for au in auth:
            auth_data = {
                'Authorization_ID':         au.Authorization_ID,
                'ID_user':                  au.ID_user.ID_user,
                'FullName':                 au.ID_user.FullName,
                'Role_ID':                  au.Role_ID.Role_ID,
                'Role_Name':                au.Role_ID.Role_Name,
                'Authorization_From':       au.Authorization_From,
                'Authorization_To':         au.Authorization_To,
                'Authorization_CreateID':   au.Authorization_CreateID,
                'Authorization_CreateBy':   au.Authorization_CreateBy,
                'Authorization_Date':       au.Authorization_Date,
                'Authorization_Date':       au.Authorization_Date.strftime('%d/%m/%Y'),
                'Authorization_Time':       au.Authorization_Time.strftime('%H:%M'),
                'Authorization_Status':     au.Authorization_Status
            }
            list_auth.append(auth_data)      
        # ########################### GROUP ROLE DATA###########################
        # groups = Role_Group.objects.all().filter(Role_Group_Status = True)      
        # list_group = [{'Role_Group_ID': group.Role_Group_ID, 'Role_Group_Name': group.Role_Group_Name} for group in groups]
        # ########################### ROLE DATA###########################
        # roles = Role_Single.objects.all().filter(Role_Status =  True)       
        # list_role = [{'Role_ID': role.Role_ID, 'Role_Name': role.Role_Name, 'Role_Group_ID': role.Role_Group_ID.Role_Group_ID } for role in roles]

        context = { 
            'data': list_auth,
            # 'groups': list_group,
            # 'roles': list_role,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('')

@csrf_exempt
def import_excel_authorize(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Đọc tệp Excel và lấy sheet đầu tiên
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                # combobox_column_name = 'Company'

                # Lấy danh sách tên cột (header)
                headers = [cell.value for cell in sheet[1]]

                # Khởi tạo danh sách để lưu dữ liệu từ các dòng
                data_list = []

                # Bắt đầu từ dòng thứ 2 (sau header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_dict = {}
                    for i, value in enumerate(row):
                        # Sử dụng tên cột từ header để làm key cho dữ liệu
                        header = headers[i]
                        data_dict[header] = value
                    data_list.append(data_dict)

                # Thực hiện xử lý dữ liệu ở đây, ví dụ: lưu vào cơ sở dữ liệu
                if data_list:
                    id_list_exits = []
                    for data in data_list:
                        UID = Users.objects.filter(ID_user = data['User ID'])
                        if not UID.exists():
                            error_user = 'User ID: ' + str(data['User ID']) + ' - Mã NV không tồn tại\n'
                            id_list_exits.append(error_user)
                            continue

                        RID = Role_Single.objects.filter(Role_ID = data['Role ID'])
                        if not RID.exists():
                            error_role = 'Role ID: ' + str(data['Role ID']) + ' - Mã Role không tồn tại\n'
                            id_list_exits.append(error_role)
                            continue

                        if data['Authorization From'] and data['Authorization To']:
                            to_date = datetime.datetime.strptime(data['Authorization To'], '%Y-%m-%d').date() 
                            from_date = datetime.datetime.strptime(data['Authorization From'], '%Y-%m-%d').date() 
                            if to_date > from_date:
                                today_date = datetime.datetime.now().date()
                                if from_date < today_date:
                                    error_date = 'Authorization From phải lớn hơn ngày hiện tại\n'
                                    id_list_exits.append(error_date)
                                    continue
                            else:
                                error_date = 'Authorization To phải lớn hơn  Authorization From\n'
                                id_list_exits.append(error_date)
                                continue
                        else:
                            error_date = 'Authorization To và  Authorization From không được trống\n'
                            id_list_exits.append(error_date)
                            continue

                        check_auth = Authorization_User.objects.filter(ID_user = data['User ID'], Role_ID = data['Role ID'])
                        if not check_auth.exists():
                            user = Users.objects.get(ID_user = 'U000000001')                                                    
                            auth = Authorization_User.objects.create(                                
                                    ID_user                     = UID.first(),
                                    Role_ID                     = RID.first(), 
                                    Authorization_From          = from_date, 
                                    Authorization_To            = to_date, 
                                    Authorization_CreateID      = user.ID_user, 
                                    Authorization_CreateBy      = user.FullName, 
                                    Authorization_Date          = datetime.datetime.now().date(),
                                    Authorization_Time          = datetime.datetime.now().time(),
                                    Authorization_Status        = True,
                            )
                            auth.save()
                        else:
                            error_data = 'User ID: ' + str(data['User ID']) + ' - Role ID: ' + str(data['Role ID']) + ' Đã tồn tại\n'
                            id_list_exits.append(error_data)
                    if id_list_exits:
                         context = { 
                            'success': True,
                            'list_data': id_list_exits,
                        }
                    else:
                        context = { 
                            'success': True,
                        }
                    return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
                # return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Chỉ cho phép tệp Excel .xlsx'})
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

################################################### PAGE IMPORT - EXPORT DATA #########################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE VIEW TICKET DATA #########################
#FUNCTION LOAD AND PROCESS DATA TICKET DETAIL
def load_Ticket_Detail_Json(request,title, ticketID):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ticket = get_object_or_404(Ticket, Ticket_ID=ticketID)
        if ticket:
            ########################### User DATA###########################    
            user = Users.objects.filter(ID_user = ticket.ID_User.ID_user).first()
            ########################### Assign DATA###########################    
            assign = Users.objects.filter(ID_user = ticket.Ticket_User_Asign).first()
            ########################### Company DATA###########################    
            company = Company.objects.filter(Company_ID = ticket.Company_ID.Company_ID).first()
            ########################### Comment DATA###########################    
            comment = Comment.objects.filter(Ticket_ID = ticket.Ticket_ID, Comment_Status = True)
            ########################### attach DATA###########################    
            attach = Attachment.objects.filter(Ticket_ID = ticket.Ticket_ID, Attachment_Status = True).order_by('-Attachment_ID')
            context = {
                'ticket': ticket,
                'Title': ticket.Ticket_Title,
                'user': user,
                'assign': assign,
                'company': company,
                'comments': comment,
                'attachs': attach,
            }
            return render(request, 'Ticket_ViewTicket.html', context)
        else:
            return redirect('/danh-sach-yeu-cau/')
        # return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA TICKET DETAIL

#FUNCTION UPDATE STATUS TICKET 
@csrf_exempt
def Update_Status_Ticket(request):
    try:
        if request.method == 'POST':
            ticketID = request.POST.get('ticketID')       
            status = request.POST.get('status')       

            ticket = Ticket.objects.filter(Ticket_ID = ticketID).first()
            if ticket:
                ticket.Ticket_Status = status
                ticket.save()
                if status == '0': 
                    FullNamesession =  request.session['UserInfo']['FullName']              
                    send_email_Done(request,FullNamesession,ticket)

                return JsonResponse({
                    'success': True,
                    'message': 'Cập Nhật Status Thành Công!',
                })

            else:
                return JsonResponse({
                'success': False,
                'message': 'Yêu Cầu Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def send_email_Done(request,Fullname,ticket):
    try:
        # full_host = request.get_full_host() 
        slug_title =  convert_title_to_slug(ticket.Ticket_Title)  
        full_host = request.build_absolute_uri('/') 
        subject = str(ticket.Ticket_ID) + ' - ' + ticket.Ticket_Title
        data = {
        'TicketID': ticket.Ticket_ID,
        'AssignName': ticket.Ticket_Name_Asign,
        'UserDone': Fullname,
        'Title': ticket.Ticket_Title,
        'CreateDate': str(ticket.Ticket_Date) + ' ' + ticket.Ticket_Time.strftime('%H:%M'),
        'CreateName': ticket.Ticket_User_Name,
        'Link': full_host+'chi-tiet-yeu-cau/'+str(ticket.Ticket_ID)+'/'+ slug_title +'/',
        }
        message = render_to_string('Email_Ticket_Done.html', context=data)
        user = Users.objects.filter(ID_user = ticket.ID_User.ID_user).first()
        if user:
            recipient_list = [user.Mail]
            # message = 'This is a test email.'
            # from_email = 'giang.llt@bamboocap.com.vn'
            from_email = settings.MICROSOFT_EMAIL
            
            # cc_list = ['cc@example.com']
            send_mail(
                subject=subject,
                message=message,
                from_email=from_email,
                recipient_list=recipient_list,
                # cc=cc_list,
                fail_silently=False,  # Thông báo lỗi nếu có vấn đề khi gửi email
                html_message=message,  # Sử dụng nội dung HTML
            )
            return True
    except Exception as ex:
        return False

#FUNCTION UPDATE STATUS TICKET 

#FUNCTION UPDATE UNREAD COMMENT
def update_unread(request,ticketID,title,ReadCommentID):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        userID = userinfo['ID_user']
        comment = ReadComment.objects.get(ReadComment_ID = ReadCommentID,ID_user = userID)
        if comment:           
            comment.ReadComment_Isread = True
            comment.save()
        link = '/chi-tiet-yeu-cau/'+str(ticketID)+'/'+title
        return redirect(link)
    else:
        return redirect('')
#FUNCTION UPDATE UNREAD COMMENT

#FUNCTION UPDATE UNREAD ALl COMMENT
@csrf_exempt
def update_unread_all(request):
    try:
        if request.method == 'POST':
            List_Unread = request.POST.getlist('data[]')
            if List_Unread:
                for u in List_Unread:
                    data_ID = json.loads(u)
                    comment = ReadComment.objects.get(ReadComment_ID = data_ID['ReadCommentID'])
                    if comment:           
                        comment.ReadComment_Isread = True
                        comment.save()
            return JsonResponse({
                    'success': True,
                    })
        else:
            return JsonResponse({
                    'success': False,
                    'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE UNREAD ALl COMMENT

#FUNCTION CREATE COMMENT 
@csrf_exempt
def Create_Comment(request):
    try:
        if request.method == 'POST':
            comment = request.POST.get('comment')
            ticketid = request.POST.get('ticketid')

            userID =  request.session['UserInfo']['ID_user']
            user = Users.objects.get(ID_user = userID)

            ticket = Ticket.objects.get(Ticket_ID = ticketid)
               
            comment = Comment.objects.create(
                Ticket_ID = ticket, 
                ID_user  = user, 
                Comment_Desc  = comment, 
                Comment_User_Name = user.FullName,
                Comment_Date  = datetime.datetime.now().date(),
                Comment_Time  = datetime.datetime.now().time(),
                Comment_Status = 1,
            )
            comment.save()
            CommnetID = comment.Comment_ID
            Update_Unread_Comment(CommnetID,ticketid,userID)
            return JsonResponse({
                    'success': True,
                    'message': 'Bình Luận Thành Công!',
                    'Comment_ID':          CommnetID,
                    'Ticket_ID':           comment.Ticket_ID.Ticket_ID,
                    'Comment_Desc':        comment.Comment_Desc,
                    'ID_user':             comment.ID_user.ID_user,
                    'Comment_User_Name':   comment.Comment_User_Name,
                    'Comment_Date':        comment.Comment_Date.strftime('%d/%m/%Y'),
                    'Comment_Time':        comment.Comment_Time.strftime('%H:%M'),
                    'Comment_Status':      comment.Comment_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE COMMENT 

#FUNCTION CREATE UNREAD COMMENT 
def Update_Unread_Comment(CommnetID,TicketID,userID):
    try:
        list_user = []
        usr = Ticket.objects.get(Ticket_ID = TicketID)
        if usr:
            list_user.append({'UserID': usr.ID_User.ID_user})
            list_user.append({'UserID': usr.Ticket_User_Asign})
        User = Comment.objects.filter(Ticket_ID = TicketID).values('ID_user').distinct()
        if User:
            for u in User:
                user_exists = any(user_dict['UserID'] == u['ID_user'] for user_dict in list_user)
            #    if u['ID_user'] not in [user_dict['UserID'] for user_dict in list_user] or userID != u['ID_user']:
                if user_exists == False:
                    list_user.append({'UserID': u['ID_user']})

        unique_list = []
        [unique_list.append(item) for item in list_user if item not in unique_list]
        if unique_list:
            ID_Comment = Comment.objects.get(Comment_ID = CommnetID)
            for l in unique_list: 
                if l['UserID'] != userID:
                    ID = Users.objects.get(ID_user = l['UserID'])             
                    Rcomment = ReadComment.objects.create(
                        Comment_ID = ID_Comment,
                        ID_user = ID,
                        ReadComment_Isread =  False
                    )
                    Rcomment.save()
            return True
        else:
            return False
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE UNREAD COMMENT 

#FUNCTION UPDATE COMMENT 
@csrf_exempt
def Update_Comment(request):
    try:
        if request.method == 'POST':
            CommentID = request.POST.get('CommentID')       
            Desc = request.POST.get('Desc')       

            comment = Comment.objects.filter(Comment_ID = CommentID).first()
            if comment:
                if Desc:
                    comment.Comment_Desc = Desc
                    comment.save()

                    return JsonResponse({
                        'success': True,
                        'message': 'Cập Nhật Commnet Thành Công!',
                        'Desc': Desc,
                    })
                else:
                    comment.Comment_Status = False
                    comment.save()

                    return JsonResponse({
                        'success': True,
                        'message': 'Xóa Commnet Thành Công!',
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Comment Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE COMMENT

#FUNCTION UPDATE COMMENT 
@csrf_exempt
def Update_Ticket_Desc(request):
    try:
        if request.method == 'POST':
            TicketID = request.POST.get('TicketID')       
            Desc = request.POST.get('Desc')       

            ticket = Ticket.objects.filter(Ticket_ID = TicketID).first()
            if ticket:
                ticket.Ticket_Desc = Desc
                ticket.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Cập Nhật Yêu Cầu Thành Công!',
                    'Desc': Desc,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Yêu Cầu Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE COMMENT

#FUNCTION CHECK FILE NAME
@csrf_exempt
def Get_File_Name(request):
    try:
        if request.method == 'POST':
            name = request.POST.get('name')
            ticketID = request.POST.get('ticketID')
            try:
                file_name = Attachment.objects.filter(Ticket_ID = ticketID, Attachment_Url__contains = name).first()
                if(file_name):
                    return JsonResponse({'success': True,'message': 'Thành Công', 'name': file_name.Attachment_Url })
                else:
                    return JsonResponse({'success': False, 'message': 'Tên File Không Tồn Tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Users.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK FILE NAME 
################################################### PAGE VIEW TICKET DATA #########################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

################################################### PAGE COMMENT DATA #######################
#FUNCTION LOAD AND PROCESS DATA COMMENT
def load_Comment_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMMENT DATA###########################    
        comments = Comment.objects.all().order_by('-Comment_ID')
        list_comment = []
        for comment in comments:
            comment_data = {
                'Comment_ID':           comment.Comment_ID,
                'Ticket_ID':            comment.Ticket_ID.Ticket_ID,
                'ID_user':              comment.ID_user.ID_user,
                'Comment_Desc':         comment.Comment_Desc,
                'Comment_User_Name':    comment.Comment_User_Name,
                'Comment_Date':         comment.Comment_Date.strftime('%d/%m/%Y'),
                'Comment_Time':         comment.Comment_Time.strftime('%H:%M'),
                'Comment_Status':       comment.Comment_Status,
            }
            list_comment.append(comment_data)       
        ########################### USERS DATA###########################
        # users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        # list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]
        context = { 
            'data': list_comment,
            # 'users': list_users,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_comment(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_ListComment.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA COMMENT

# #FUNCTION CREATE TICKET 
# @csrf_exempt
# def Create_Company(request):
#     try:
#         if request.method == 'POST':
#             Company_Name = request.POST.get('Company_Name')

#             userID =  request.session['UserInfo']['ID_user']
#             user = Users.objects.get(ID_user = userID)
               
#             company = Company.objects.create(
#                 Company_Name = Company_Name, 
#                 ID_user  = user, 
#                 Company_User_Name  = user.FullName, 
#                 Company_Date  = datetime.datetime.now().date(),
#                 Company_Time  = datetime.datetime.now().time(),
#                 Company_Status = 1,
#             )
#             company.save()
#             companyID = company.Company_ID
#             return JsonResponse({
#                     'success': True,
#                     'message': 'Công Ty Tạo Thành Công!',
#                     'Company_ID':           companyID,
#                     'Company_Name':         company.Company_Name,
#                     'ID_user':              user.ID_user,
#                     'Company_User_Name':    user.FullName,
#                     'Company_Date':         company.Company_Date.strftime('%d/%m/%Y'),
#                     'Company_Time':         company.Company_Time.strftime('%H:%M'),
#                     'Company_Status':       company.Company_Status,
#                 })
#         else:
#             response = {'success': False, 'message': 'Invalid request method'}
#             # return JsonResponse(response, status=405)
#             return JsonResponse({
#                 'success': False,
#                 'message': 'Invalid request method'
#                 })
#     except Exception as ex:
#         return JsonResponse({
#             'success': False,
#             'message': f'Lỗi: {str(ex)}',
#         })
# #FUNCTION CREATE TICKET 

# #FUNCTION DELETE TICKET 
# @csrf_exempt
# def Delete_Company(request):
#     try:
#         if request.method == 'POST':
#             CompanyID = request.POST.get('CompanyID')
#             try:
#                 company = Company.objects.get(Company_ID = CompanyID)
#                 if(company):
#                     company.delete()
#                     return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
#                 else:
#                     return JsonResponse({'success': False, 'message': 'Company ID ' + CompanyID +' Không tôn tại'})
#             except Exception as ex:
#                 return JsonResponse({
#                     'success': False,
#                     'message': f'Lỗi: {str(ex)}',
#                 })
#             except Ticket.DoesNotExist:
#                 return JsonResponse({'success': False, 'message': 'ID ' + CompanyID +' Không tôn tại'})
#         else:
#             return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
#     except Exception as ex:
#         return JsonResponse({
#             'success': False,
#             'message': f'Lỗi: {str(ex)}',
#         })
# #FUNCTION DELETE TICKET 

# #FUNCTION UPDATE COMMENT 
@csrf_exempt
def Update_Comment(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            CommentID = request.POST.get('CommentID') 
            comment = Comment.objects.get(Comment_ID = CommentID)
            if comment:
                if str(comment.Comment_Status) == status:
                     return JsonResponse({
                    'success': False,
                    'message': 'Status đang ở trạng thái: "Không Kích Hoạt"!',         
                })
                else:
                    comment.Comment_Status = status
                    comment.save()

                    return JsonResponse({
                        'success': True,
                        'message': 'Yêu Cầu Thành Công!',         
                        'Comment_Status':    comment.Comment_Status,
                        'Comment_ID'    :    comment.Comment_ID,
                    })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Comment Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
# #FUNCTION UPDATE COMMENT 

# #FUNCTION UPDATE COMMENT 
@csrf_exempt
def Update_Comment_Desc(request):
    try:
        if request.method == 'POST':    
            CommentID = request.POST.get('CommentID') 
            CommentDesc = request.POST.get('CommentDesc') 

            comment = Comment.objects.get(Comment_ID = CommentID)
            if comment:
                comment.Comment_Desc = CommentDesc
                comment.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',         
                    'Comment_Desc'  :    comment.Comment_Desc,
                    'Comment_ID'    :    comment.Comment_ID,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Comment Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
# #FUNCTION UPDATE COMMENT 

# #FUNCTION UPDATE DATA COMMENT 
@csrf_exempt
def Data_Update_Comment(request):
       try:
        if request.method == 'POST':
            CommentID = request.POST.get('CommentID')
            comment = Comment.objects.filter(Comment_ID = CommentID).first()
            list_comment = ({
                     'Comment_ID':     comment.Comment_ID, 
                     'Comment_Desc':   comment.Comment_Desc})
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Comment': list_comment,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Comment ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
# #FUNCTION UPDATE DATA COMMENT 

################################################### PAGE COMMENT DATA ##########################  
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE GROUP ROLE DATA #######################
#FUNCTION LOAD AND PROCESS DATA GROUP ROLE
def load_Group_Role_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        grouproles = Role_Group.objects.all().order_by('-Role_Group_ID')
        list_group_role = []
        for group in grouproles:
            group_role_data = {
                'Role_Group_ID':           group.Role_Group_ID,
                'Role_Group_Name':         group.Role_Group_Name,
                'Menu_ID':                 group.Menu_ID.Menu_ID ,
                'Menu_Name':               group.Menu_ID.Menu_Name if group.Menu_ID.Menu_Name else "No Data",
                'Role_Group_Address':      group.Role_Group_Address if group.Role_Group_Address else "No Data",
                'ID_user':                 group.ID_user.ID_user,
                'Role_Group_CreateBy':     group.Role_Group_CreateBy,
                'Role_Group_Date':         group.Role_Group_Date.strftime('%d/%m/%Y'),
                'Role_Group_Time':         group.Role_Group_Time.strftime('%H:%M'),
                'Role_Group_Status':       group.Role_Group_Status,
            }
            list_group_role.append(group_role_data)       
        ########################### USERS DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]


        context = { 
            'data': list_group_role,
            'users': list_users,}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_group_role(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_RoleGroup.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA GROUP ROLE

#FUNCTION CREATE GROUP ROLE
@csrf_exempt
def Create_Group_Role(request):
    try:
        if request.method == 'POST':
            Group_Name = request.POST.get('Group_Name')
            MenuID = request.POST.get('MenuID')
            Menu_Add = request.POST.get('Menu_Add')

            userID =  request.session['UserInfo']['ID_user']
            user = Users.objects.get(ID_user = userID)

            menu = Menu.objects.get(Menu_ID = MenuID)
               
            group = Role_Group.objects.create(
                Role_Group_Name = Group_Name, 
                ID_user  = user, 
                Menu_ID = menu,
                Role_Group_Address = Menu_Add,
                Role_Group_CreateBy  = user.FullName, 
                Role_Group_Date  = datetime.datetime.now().date(),
                Role_Group_Time  = datetime.datetime.now().time(),
                Role_Group_Status = 1,
            )
            group.save()
            GroupID = group.Role_Group_ID
            return JsonResponse({
                    'success': True,
                    'message': 'Nhóm Quyền Tạo Thành Công!',
                    'Role_Group_ID':           GroupID,
                    'Role_Group_Name':         group.Role_Group_Name,
                    'ID_user':                 user.ID_user,
                    'Menu_Name':               group.Menu_ID.Menu_Name,
                    'Role_Group_Address':      group.Role_Group_Address,
                    'Role_Group_CreateBy':     user.FullName,
                    'Role_Group_Date':         group.Role_Group_Date.strftime('%d/%m/%Y'),
                    'Role_Group_Time':         group.Role_Group_Time.strftime('%H:%M'),
                    'Role_Group_Status':       group.Role_Group_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE GROUP ROLE

#FUNCTION DELETE GROUP ROLE
@csrf_exempt
def Delete_Group_Role(request):
    try:
        if request.method == 'POST':
            GroupRoleID = request.POST.get('GroupRoleID')
            try:
                group = Role_Group.objects.get(Role_Group_ID = GroupRoleID)
                if(group):
                    group.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'Group ID ' + GroupRoleID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + GroupRoleID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE GROUP ROLE

#FUNCTION UPDATE GROUP ROLE
@csrf_exempt
def Update_Group_Role(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            GroupID = request.POST.get('GroupID')       
            groupName = request.POST.get('GroupName')   
            menu = request.POST.get('Menu')   
            Address = request.POST.get('Address')   

            group = Role_Group.objects.get(Role_Group_ID = GroupID)
            if group:             
                if groupName or menu or Address: 
                    menuID = Menu.objects.get(Menu_ID = menu)                
                    group.Role_Group_Name = groupName
                    group.Menu_ID = menuID
                    group.Role_Group_Address = Address
                    group.Role_Group_Status = status
                    group.save()
                else:
                    menuID = Menu.objects.get(Menu_ID = group.Menu_ID.Menu_ID)
                    group.Role_Group_Name = group.Role_Group_Name
                    group.Menu_ID = menuID
                    group.Role_Group_Address = group.Role_Group_Address
                    group.Role_Group_Status = status
                    group.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Role_Group_ID':        group.Role_Group_ID,
                    'Role_Group_Name':      group.Role_Group_Name,                   
                    'Menu_Name':              group.Menu_ID.Menu_Name,                   
                    'Role_Group_Address':      group.Role_Group_Address,                   
                    'Role_Group_Status':    group.Role_Group_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Nhóm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE GROUP ROLE

#FUNCTION UPDATE DATA GROUP ROLE
@csrf_exempt
def Data_Update_Group_Role(request):
       try:
        if request.method == 'POST':
            GroupID = request.POST.get('GroupID')
            group = Role_Group.objects.filter(Role_Group_ID = GroupID)
            list_group = [{
                     'Role_Group_ID':     p.Role_Group_ID, 
                     'Role_Group_Name':   p.Role_Group_Name, 
                     'Menu_ID':           p.Menu_ID.Menu_ID, 
                     'Role_Group_Address':   p.Role_Group_Address, 
                     'Role_Group_Status': p.Role_Group_Status,} for p in group]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Groups': list_group,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA GROUP ROLE

#FUNCTION LOAD MENU
@csrf_exempt
def Load_Data_Menu(request):
    try:
        if request.method == 'POST':          
            menus = Menu.objects.all().order_by('Menu_Level')
            list_menu = []
            for menu in menus:
                menu_data = {
                    'Menu_ID':           menu.Menu_ID,
                    'Menu_Name':         menu.Menu_Name,
                }
                list_menu.append(menu_data)
            context = { 
                'success': True,
                'Menus': list_menu,
            }
            return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION LOAD MENU

################################################### PAGE GROUP ROLE DATA #######################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE ROLE DATA #######################
#FUNCTION LOAD AND PROCESS DATA ROLE
def load_Role_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### COMPANY DATA###########################    
        roles = Role_Single.objects.all().order_by('-Role_ID')
        list_role = []
        for group in roles:
            role_data = {
                'Role_ID':           group.Role_ID,
                'Role_Name':         group.Role_Name,
                'Role_Group_ID':     group.Role_Group_ID.Role_Group_ID,
                'Role_Group_Name':   group.Role_Group_ID.Role_Group_Name,
                'ID_user':           group.ID_user.ID_user,
                'Role_CreateBy':     group.Role_CreateBy,
                'Role_Date':         group.Role_Date.strftime('%d/%m/%Y'),
                'Role_Time':         group.Role_Time.strftime('%H:%M'),
                'Role_Status':       group.Role_Status,
            }
            list_role.append(role_data)       
        ########################### USERS DATA###########################
        users = Users.objects.filter(User_Status = True, User_Type__lt = 2)
        list_users = [{'ID_user': user.ID_user, 'FullName': user.FullName} for user in users]

        ########################### GROUP ROLE DATA###########################
        groups = Role_Group.objects.all()
        list_group = [{'Role_Group_ID': group.Role_Group_ID, 'Role_Group_Name': group.Role_Group_Name} for group in groups]

        context = { 
            'data': list_role,
            'users': list_users,
            'groups': list_group,}
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_role(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_Role.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA ROLE

#FUNCTION CREATE ROLE
@csrf_exempt
def Create_Role(request):
    try:
        if request.method == 'POST':
            RoleName = request.POST.get('Role_Name')
            Group_ID = request.POST.get('Group_ID')

            userID =  request.session['UserInfo']['ID_user']
            user = Users.objects.get(ID_user = userID)
            group = Role_Group.objects.get(Role_Group_ID = Group_ID)
               
            role = Role_Single.objects.create(
                Role_Name = RoleName, 
                ID_user  = user, 
                Role_Group_ID =  group,
                Role_CreateBy  = user.FullName, 
                Role_Date  = datetime.datetime.now().date(),
                Role_Time  = datetime.datetime.now().time(),
                Role_Status = 1,
            )
            group.save()
            RoleID = role.Role_ID
            return JsonResponse({
                    'success': True,
                    'message': 'Công Ty Tạo Thành Công!',
                    'Role_ID':           RoleID,
                    'ID_user':                 user.ID_user,
                    'Role_Name':               role.Role_Name,
                    'Role_Group_ID':           role.Role_Group_ID.Role_Group_ID,
                    'Role_Group_Name':         role.Role_Group_ID.Role_Group_Name,
                    'Role_CreateBy':           user.FullName,
                    'Role_Date':               role.Role_Date.strftime('%d/%m/%Y'),
                    'Role_Time':               role.Role_Time.strftime('%H:%M'),
                    'Role_Status':             role.Role_Status,
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            # return JsonResponse(response, status=405)
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE ROLE

#FUNCTION DELETE ROLE
@csrf_exempt
def Delete_Role(request):
    try:
        if request.method == 'POST':
            RoleID = request.POST.get('RoleID')
            try:
                role = Role_Single.objects.get(Role_ID = RoleID)
                if(role):
                    role.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'Group ID ' + RoleID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + RoleID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE ROLE

#FUNCTION UPDATE ROLE
@csrf_exempt
def Update_Role(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            RoleID = request.POST.get('RoleID')       
            RoleName = request.POST.get('RoleName')       
            GroupID = request.POST.get('GroupID')       
            role = Role_Single.objects.get(Role_ID = RoleID)
            if role:
                if RoleName or GroupID:             
                    role.Role_Name = RoleName if RoleName else role.Role_Name
                    if GroupID: 
                        groupid = Role_Group.objects.filter(Role_Group_ID = GroupID).first()
                        role.Role_Group_ID = groupid
                    role.Role_Status = status
                    role.save()
                else:
                    role.Role_Name = role.Role_Name
                    role.Role_Status = status
                    role.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Role_ID':        role.Role_ID,
                    'Role_Name':      role.Role_Name,                   
                    'Role_Group_ID':  role.Role_Group_ID.Role_Group_ID,                   
                    'Role_Group_Name':  role.Role_Group_ID.Role_Group_Name,                   
                    'Role_Status':    role.Role_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Nhóm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ROLE

#FUNCTION UPDATE DATA ROLE
@csrf_exempt
def Data_Update_Role(request):
       try:
        if request.method == 'POST':
            RoleID = request.POST.get('RoleID')
            role = Role_Single.objects.filter(Role_ID = RoleID)
            list_group = [{
                     'Role_ID':         p.Role_ID, 
                     'Role_Name':       p.Role_Name, 
                     'Role_Group_ID':   p.Role_Group_ID.Role_Group_ID, 
                     'Role_Status':     p.Role_Status,} for p in role]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Groups': list_group,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA ROLE

################################################### PAGE ROLE DATA #######################

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE ROLE DATA #######################
#FUNCTION LOAD AND PROCESS DATA AUHTORIZE
def load_Authorize_Json(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        ########################### USERS DATA###########################   
        users = Users.objects.all().filter(User_Status = True).order_by('-ID_user')
        list_user = []
        for user in users:
            user_data = {
                'ID_user':          user.ID_user,
                'FullName':         user.FullName,
            }
            list_user.append(user_data)      
        ########################### GROUP ROLE DATA###########################
        groups = Role_Group.objects.all().filter(Role_Group_Status = True)      
        list_group = [{'Role_Group_ID': group.Role_Group_ID, 'Role_Group_Name': group.Role_Group_Name} for group in groups]
        ########################### ROLE DATA###########################
        roles = Role_Single.objects.all().filter(Role_Status =  True)       
        list_role = [{'Role_ID': role.Role_ID, 'Role_Name': role.Role_Name, 'Role_Group_ID': role.Role_Group_ID.Role_Group_ID } for role in roles]

        context = { 
            'users': list_user,
            'groups': list_group,
            'roles': list_role,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('/')

def load_authorize(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        return render(request, 'Ticket_Authorize.html')
    else:
        return redirect('/')
#FUNCTION LOAD AND PROCESS DATA AUHTORIZE

#FUNCTION LOAD DATA AUHTORIZE USER
@csrf_exempt
def load_Authorize_data(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        currentDate = datetime.datetime.now()
        user = request.POST.get('UserID')
        ########################### GROUP ROLE DATA###########################
        groups = Role_Group.objects.all().filter(Role_Group_Status = True)       
        list_group = [{'Role_Group_ID': group.Role_Group_ID, 'Role_Group_Name': group.Role_Group_Name} for group in groups]
        ########################### ROLE DATA###########################
        roles = Role_Single.objects.all().filter(Role_Status =  True)        
        list_role = [{'Role_ID': role.Role_ID, 'Role_Name': role.Role_Name, 'Role_Group_ID': role.Role_Group_ID.Role_Group_ID } for role in roles]

        for role in list_role:
            try:
                # auth = Authorization_User.objects.filter(ID_user=user, Role_ID=role['Role_ID'],Authorization_From__lte=currentDate,Authorization_To__gte=currentDate, Authorization_Status = True).order_by('-Authorization_From','-Authorization_To').first()
                auth = Authorization_User.objects.filter(ID_user=user, Role_ID=role['Role_ID'],Authorization_To__gte=currentDate, Authorization_Status = True).order_by('Authorization_From').first()
                auth_to = Authorization_User.objects.filter(ID_user=user, Role_ID=role['Role_ID'],Authorization_To__gte=currentDate, Authorization_Status = True).order_by('-Authorization_To').first()
                if auth:
                    if auth.Authorization_Status == True:
                        role['isStatus'] = 'checked'
                        role['DateFrom'] = auth.Authorization_From.strftime('%d/%m/%Y'),
                        role['DateTo']   = auth_to.Authorization_To.strftime('%d/%m/%Y'),
                    else:
                        role['isStatus'] = ''
                        role['DateFrom'] = ''
                        role['DateTo']   = ''
                else:
                    role['isStatus'] = ''
                    role['DateFrom'] = ''
                    role['DateTo']   = ''
            except Authorization_User.DoesNotExist:
                role['isStatus'] = ''
                role['DateFrom'] = ''
                role['DateTo']   = ''

        context = { 
            'groups': list_group,
            'roles': list_role,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    else:
        return redirect('')
#FUNCTION LOAD DATA AUHTORIZE USER

#FUNCTION CREATE ROLE
@csrf_exempt
def Create_Authorize(request):
    try:
        if request.method == 'POST':
            Date_Current = datetime.datetime.now().date()
            Date_From = request.POST.get('Date_From')
            Date_To = request.POST.get('Date_To')
            List_Users = request.POST.getlist('List_Users[]')
            List_Roles = request.POST.getlist('List_Roles[]')

            userID =  request.session['UserInfo']['ID_user']
            user_create = Users.objects.get(ID_user = userID)

            if Date_From == "": 
                Date_From = Date_Current
            else: 
                Date_From = datetime.datetime.strptime(Date_From, '%Y-%m-%d').date()

            if Date_To == "": 
                Date_To = datetime.datetime.strptime('9999-12-31', '%Y-%m-%d').date() 
            else: 
                Date_To = datetime.datetime.strptime(Date_To, '%Y-%m-%d').date() 

            if Date_From < Date_Current:
                return JsonResponse({
                    'success': False,
                    'message': 'Ngày Bắt Đầu Phải Lớn Hơn Hoặc Bằng Ngày Hiện Tại.',
                })
            elif Date_To <= Date_Current:
                return JsonResponse({
                    'success': False,
                    'message': 'Ngày Kết Thúc Phải Lón Hơn Ngày Hiện Tại.',
                })
            elif Date_From > Date_To:
                return JsonResponse({
                    'success': False,
                    'message': 'Ngày Bắt Đầu Phải Nhỏ Hơn Ngày Kết Thúc.',
                })

            for u in List_Users:
                data_user = json.loads(u)
                user = Users.objects.get(ID_user = data_user['UserID'])
                if user:
                    for r in List_Roles:
                        data_role = json.loads(r)
                        role = Authorization_User.objects.filter(
                                ID_user=user.ID_user, 
                                Role_ID=data_role['RoleID'],
                                Authorization_From__lte=Date_From,
                                Authorization_To__gte=Date_From,
                                Authorization_Status = True).order_by('-Authorization_ID').first()
                        if data_role['Status'] == 'True':   #checkbox is checked                                                     
                            if role: #authorize is exits
                                role_single = Role_Single.objects.get(Role_ID = role.Role_ID.Role_ID)
                                # Update_role_data(Date_From, Date_To ,role.Authorization_ID,'True')
                                Update_role_data(Date_From, Date_To ,role_single,user,'True')
                                # Insert_role_data(Date_From, Date_To, role.Role_ID, role.ID_user, user_create.ID_user,user_create.FullName)
                                Insert_role_data(Date_From, Date_To, role_single, user, user_create.ID_user,user_create.FullName)
                            else: #authorize is not exits
                                role_single = Role_Single.objects.get(Role_ID = data_role['RoleID'])
                                # Insert_role_data(Date_From, Date_To, data_role['RoleID'], user.ID_user, user_create.ID_user,user_create.FullName)
                                Insert_role_data(Date_From, Date_To, role_single , user, user_create.ID_user,user_create.FullName)
                        else: #checkbox is uncheck
                            role_single = Role_Single.objects.get(Role_ID = data_role['RoleID'])
                            if role: #authorize is exits
                                Update_unrole_data(Date_From,user,role_single)
                            else:
                                Update_unrole_data(Date_From,user,role_single)                                            
                            
            return JsonResponse({
                    'success': True,
                    'message': 'Phân Quyền Thành Công!',
                })
        else:
            response = {'success': False, 'message': 'Invalid request method'}
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def Update_role_data(Date_From,Date_To, Role, User, Status):
    try:
        date_last = Date_From - datetime.timedelta(days=1)
        # role = Authorization_User.objects.get(Authorization_ID = Authorization_ID)
        update = False
        role = Authorization_User.objects.filter(
            Authorization_Status = True, 
            Authorization_To__gte=Date_From,
            ID_user=User.ID_user, 
            Role_ID = Role.Role_ID).order_by('Authorization_To')
        if role:
            for r in role:
                if r.Authorization_To <  Date_To or r.Authorization_From <  Date_From:
                    update = True
                    auth_id = r.Authorization_ID
                else:
                    update = False
        else: 
            update = True

        if update == True:
            Date_3112 = datetime.datetime.strptime('9999-12-31', '%Y-%m-%d').date()
            auth = Authorization_User.objects.get(Authorization_ID = auth_id)
            if auth.Authorization_From < Date_From or auth.Authorization_To == Date_3112:
                auth.Authorization_To = date_last
                if Status == 'False':
                    auth.Authorization_Status = Status
                auth.save()
            else:
                auth.Authorization_To = auth.Authorization_From
                if Status == 'False':
                    auth.Authorization_Status = Status
                auth.save()
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })

def Update_unrole_data(Date_From, user, Role):
    try:
        role = Authorization_User.objects.filter(
            ID_user = user.ID_user,
            Authorization_Status = True, 
            Authorization_To__gte=Date_From,
            Role_ID = Role.Role_ID)
        if role:
            for r in role:
                r.Authorization_Status = 'False'
                r.save()
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })

def Insert_role_data(Date_From, Date_To, Role, User, ID_Create ,FullName_Create):
    try:
        date_create = datetime.datetime.now().date()
        time_create = datetime.datetime.now().time()
        insert = 0
        role = Authorization_User.objects.filter(
            Authorization_Status = True, 
            Authorization_To__gte=Date_From,
            ID_user=User.ID_user,
            Role_ID = Role.Role_ID).order_by('Authorization_To')
        if role:
            for r in role:
                if r.Authorization_To <  Date_To or r.Authorization_From <  Date_From:
                    insert = 1
                elif r.Authorization_From > Date_To:
                    date_last = r.Authorization_From - datetime.timedelta(days=1)
                    insert = 2
                    Date_To_= date_last
                else:
                    insert = 0
        else: 
            insert = 1

        if insert == 1:
            auth = Authorization_User.objects.create(
                Role_ID = Role,
                ID_user = User,
                Authorization_From = Date_From,
                Authorization_To = Date_To,
                Authorization_CreateID = ID_Create,
                Authorization_CreateBy = FullName_Create,
                Authorization_Date = date_create,
                Authorization_Time = time_create,
                Authorization_Status = True
            )
            auth.save()
        elif insert == 2:
            auth = Authorization_User.objects.create(
                Role_ID = Role,
                ID_user = User,
                Authorization_From = Date_From,
                Authorization_To = Date_To_,
                Authorization_CreateID = ID_Create,
                Authorization_CreateBy = FullName_Create,
                Authorization_Date = date_create,
                Authorization_Time = time_create,
                Authorization_Status = True
            )
            auth.save()
        
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CREATE ROLE

#FUNCTION DELETE ROLE
@csrf_exempt
def Delete_Role(request):
    try:
        if request.method == 'POST':
            RoleID = request.POST.get('RoleID')
            try:
                role = Role_Single.objects.get(Role_ID = RoleID)
                if(role):
                    role.delete()
                    return JsonResponse({'success': True,'message': 'Xóa yêu cầu thành công!',})
                else:
                    return JsonResponse({'success': False, 'message': 'Group ID ' + RoleID +' Không tôn tại'})
            except Exception as ex:
                return JsonResponse({
                    'success': False,
                    'message': f'Lỗi: {str(ex)}',
                })
            except Ticket.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ID ' + RoleID +' Không tôn tại'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION DELETE ROLE

#FUNCTION UPDATE ROLE
@csrf_exempt
def Update_Role(request):
    try:
        if request.method == 'POST':
            status = request.POST.get('status')       
            RoleID = request.POST.get('RoleID')       
            RoleName = request.POST.get('RoleName')       
            GroupID = request.POST.get('GroupID')       
            role = Role_Single.objects.get(Role_ID = RoleID)
            if role:
                if RoleName or GroupID:             
                    role.Role_Name = RoleName if RoleName else role.Role_Name
                    if GroupID: 
                        groupid = Role_Group.objects.filter(Role_Group_ID = GroupID).first()
                        role.Role_Group_ID = groupid
                    role.Role_Status = status
                    role.save()
                else:
                    role.Role_Name = role.Role_Name
                    role.Role_Status = status
                    role.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Yêu Cầu Thành Công!',
                    'Role_ID':        role.Role_ID,
                    'Role_Name':      role.Role_Name,                   
                    'Role_Group_ID':  role.Role_Group_ID.Role_Group_ID,                   
                    'Role_Group_Name':  role.Role_Group_ID.Role_Group_Name,                   
                    'Role_Status':    role.Role_Status,
                })
            else:
                return JsonResponse({
                'success': False,
                'message': 'Mã Nhóm Không Tồn Tại',
                })

        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request method'
                })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE ROLE

#FUNCTION UPDATE DATA ROLE
@csrf_exempt
def Data_Update_Role(request):
       try:
        if request.method == 'POST':
            RoleID = request.POST.get('RoleID')
            role = Role_Single.objects.filter(Role_ID = RoleID)
            list_group = [{
                     'Role_ID':         p.Role_ID, 
                     'Role_Name':       p.Role_Name, 
                     'Role_Group_ID':   p.Role_Group_ID.Role_Group_ID, 
                     'Role_Status':     p.Role_Status,} for p in role]
            context = { 
                'success': True,
                'message': 'Lấy Data Thành Công',
                'Groups': list_group,}
            data = json.dumps(context, ensure_ascii=False)
            return HttpResponse(data, content_type='application/json')
        else:
           return JsonResponse({'success': False,'message': 'Group ID Không Tồn Tại.'}) 
       except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION UPDATE DATA ROLE

################################################### PAGE ROLE DATA #######################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE AUTHORIZE ROLE #######################
#FUNCTION CHECK AUTHORIZE USERS 
@csrf_exempt
def Auth_Role_User(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZUM_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZUM_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZUM_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZUM_Del','Role': 'Delete','Status': 'False'},
                {'TCode': 'ZUM_Admin','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 2, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_user(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 2, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE USERS 

#FUNCTION CHECK AUTHORIZE TICKET 
@csrf_exempt
def Auth_Role_Ticket(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZTC_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZTC_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZTC_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZTC_Del','Role': 'Delete','Status': 'False'},
                {'TCode': 'ZTC_Admin','Role': 'Admin View','Status': 'False'},
                {'TCode': 'ZTC_LoadMail','Role': 'Load Mail','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 3, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_ticket(request):
    try:
        userinfo = request.session.get('UserInfo')
        isAdmin = False
        Tcode = 'ZTC_Admin'
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 3, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         for item in auth:
                            name = item.Role_ID.Role_Name
                            if Tcode in name:
                                isAdmin = True

                         return JsonResponse({
                            'success': True,
                            'isAdmin': isAdmin
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                        })     
            else:
                 isAdmin = True
                 return JsonResponse({
                        'success': True,
                        'isAdmin': isAdmin
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE TICKET 

#FUNCTION CHECK AUTHORIZE ATTACH FILE
@csrf_exempt
def Auth_Role_Attach(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZAF_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZAF_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZAF_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZAF_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 4, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_attach(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 4, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE ATTACH FILE 

#FUNCTION CHECK AUTHORIZE COMMENT
@csrf_exempt
def Auth_Role_Comment(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZCM_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZCM_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZCM_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 5, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_comment(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 5, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE COMMENT

#FUNCTION CHECK AUTHORIZE COMPANY
@csrf_exempt
def Auth_Role_Company(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZCP_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZCP_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZCP_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZCP_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 6, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_company(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 6, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE COMPANY 

#FUNCTION CHECK AUTHORIZE GROUP
@csrf_exempt
def Auth_Role_Group(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZGR_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZGR_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZGR_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZGR_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 7, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_group(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 7, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE GROUP 

#FUNCTION CHECK AUTHORIZE ASSIGN
@csrf_exempt
def Auth_Role_Assign(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZAS_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZAS_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZAS_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZAS_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 8, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_assign(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 8, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE ASSIGN 

#FUNCTION CHECK AUTHORIZE GROUP ROLE
@csrf_exempt
def Auth_Role_GroupRole(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZRG_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZRG_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZRG_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZRG_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 10, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_grouprole(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 10, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE GROUP ROLE

#FUNCTION CHECK AUTHORIZE ROLE
@csrf_exempt
def Auth_Role(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZRL_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZRL_Edit','Role': 'Edit','Status': 'False'},
                {'TCode': 'ZRL_Add','Role': 'Add','Status': 'False'},
                {'TCode': 'ZRL_Del','Role': 'Delete','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 11, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_role(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 11, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE ROLE

#FUNCTION CHECK AUTHORIZE
@csrf_exempt
def Auth_Authorize(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZAU_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZAU_Add','Role': 'Add','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 14, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_authorize(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 14, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE

#FUNCTION CHECK AUTHORIZE
@csrf_exempt
def Auth_Menu(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 2:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZMN_View','Role': 'View','Status': 'False'},
                {'TCode': 'ZMN_Edit','Role': 'View','Status': 'False'},
                {'TCode': 'ZMN_Add','Role': 'View','Status': 'False'},
                {'TCode': 'ZMN_Del','Role': 'View','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 19, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_menu(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 19, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK AUTHORIZE

#FUNCTION CHECK DOCUMENT TEMPLATE
def check_template(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Dash_Role_Data = [
                {'TCode': 'ZTP_View','Role': 'View','Status': 'False'},
                ]
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 17,Role_Name__contains= Dash_Role_Data[0]['TCode'], Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                        return True
                    else:
                        return False 
                else:
                    return False  
            else:
                return True
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })

def check_document(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Dash_Role_Data = [
                {'TCode': 'ZDC_View','Role': 'View','Status': 'False'},
                ]
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 18,Role_Name__contains= Dash_Role_Data[0]['TCode'], Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                        return True
                    else:
                        return False 
                else:
                    return False  
            else:
                return True
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK DOCUMENT TEMPLATE

#FUNCTION CHECK GITHUB
@csrf_exempt
def Auth_Role_Github(request):
    try:
        if request.method == 'POST':
            userinfo =  request.session['UserInfo']
            ID_user  =  userinfo['ID_user']
            if int(userinfo['Acc_type']) < 1:
                IsAdmin = True
            else:
                IsAdmin = False
            
            current_date = datetime.datetime.now()
            Dash_Role_Data = [
                {'TCode': 'ZGH_LoadGithub','Role': 'Load','Status': 'False'},
            ]
            Sing_Role = Role_Single.objects.filter(Role_Group_ID = 21, Role_Status = True)
            if Sing_Role:
                for s in Sing_Role:
                    auth = Authorization_User.objects.filter(
                        ID_user =  ID_user, 
                        Role_ID = s.Role_ID,
                        Authorization_From__lte=current_date,
                        Authorization_To__gte=current_date, 
                        Authorization_Status = True).order_by('-Authorization_To').first() 
                    if auth:
                        for item_role in Dash_Role_Data:
                            tcode = item_role['TCode']
                            if tcode in auth.Role_ID.Role_Name:
                                item_role['Status'] = 'True'
            context = { 
                'success': True,
                'IsAdmin': IsAdmin,
                'Roles': Dash_Role_Data,
            }
        return HttpResponse(json.dumps(context, default=date_handler, ensure_ascii=False), content_type='application/json')
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(ex)}',
        })
    
def check_Github(request):
    try:
        userinfo = request.session.get('UserInfo')
        if userinfo:
            idUser = userinfo['ID_user']
            typeUser = userinfo['Acc_type']
            if typeUser > 0:
                current_date = datetime.datetime.now()
                Sing_Role = Role_Single.objects.filter(Role_Group_ID = 21, Role_Status = True).values('Role_ID')
                if Sing_Role:
                    auth = Authorization_User.objects.filter(
                                ID_user =  idUser, 
                                Role_ID__in = Sing_Role.values('Role_ID'),
                                Authorization_From__lte=current_date,
                                Authorization_To__gte=current_date, 
                                Authorization_Status = True) 
                    if auth:
                         return JsonResponse({
                            'success': True,
                            # 'Status':  True,
                        })
                    else:
                        return JsonResponse({
                            'success': False,
                            # 'Status':  False,
                        })     
            else:
                 return JsonResponse({
                        'success': True,
                        # 'Status':  True,
                })
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })
#FUNCTION CHECK GITHUB 
################################################### PAGE AUTHORIZE ROLE #######################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
################################################### PAGE TEMPLATE #######################
def Template(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
    if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
        status = check_template(request)
        if status == True:
            return render(request, 'Ticket_Template.html')
        else:
            return redirect('/dashboard/')
    else:
        return redirect('/')

def Document(request):
    try:
        cookie_system_data     = GetCookie(request, 'cookie_system_data')
        cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
        # if(cookie_microsoft_data or cookie_system_data or request.session['UserInfo']):
        if cookie_microsoft_data or cookie_system_data or 'UserInfo' in request.session:
            status = check_document(request)
            if status == True:
                return render(request, 'Ticket_Document.html')
            else:
                return redirect('/dashboard/')
        else:
            return redirect('/')
    except Exception as ex:
        return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(ex)}',
        })

    
################################################### PAGE TEMPLATE #######################
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------#

############################################ PAGE TICKET HELPDESK - END ############################################################







############################################ OTHER - START ############################################################
def Admin(request):
    #load template in folder teamplates
    template =  loader.get_template('Master_page.html')
    return HttpResponse(template.render())

def Trangchu(request):
    #load template in folder teamplates
    template =  loader.get_template('trangchu.html')
    return HttpResponse(template.render())

def Home(request):
    #load template in folder teamplates
    template =  loader.get_template('Home_Admin.html')
    return HttpResponse(template.render())

# def Page_data(request):
#     #load template in folder teamplates
#     data =  T001.objects.using('default').all()
#     return render(request, 'Page_data.html', {'data': data})

#function Product
def Product_data(request):
    #load template in folder teamplates
    # data =  Product.objects.all().select_related('category')
    # data =  Product.objects.prefetch_related('id_cate').filter(price__gte=20000, price__lte=1000000)
    SearchMaterial = request.GET.get('SearchMaterial')
    SearchCategory = request.GET.get('SearchCategory')
    # if SearchMaterial != "" or SearchMaterial != None or SearchCategory != "" or SearchCategory != None:
    if (SearchMaterial is not None and SearchMaterial != "" ) and (SearchCategory is not None and SearchCategory != ""):
        data = Product.objects.prefetch_related('id_cate').filter(name__icontains=SearchMaterial,category__icontains=SearchCategory)
    elif (SearchMaterial is not None and SearchMaterial != "" ):
        data = Product.objects.prefetch_related('id_cate').filter(name__icontains=SearchMaterial)
    elif (SearchCategory is not None and SearchCategory != ""):
        data = Product.objects.prefetch_related('id_cate').filter(category__icontains=SearchCategory)
    else:
        data =  Product.objects.prefetch_related('id_cate')
    combobox = Category.objects.using('default').all()
    paginator = Paginator(data, 4)  # hiển thị 10 đối tượng trên mỗi trang
    page_number = request.GET.get('page') # get current page - default page 1
    page_obj = paginator.get_page(page_number)

    return render(request, 'Product.html',  {'page_obj': page_obj, 'combobox': combobox})

def load_data(request):
    userinfo = request.session.get('UserInfo')
    if userinfo:
        #load template in folder teamplates
        products = Product.objects.prefetch_related('id_cate').order_by('-id')
        list = [{'id': p.id, 'name': p.name, 'category': p.category, 'price': int(p.price), 'id_cate': p.id_cate.pk} for p in products]
        data = json.dumps(list, ensure_ascii=False)
        userinfo = request.session.get('UserInfo')
        return HttpResponse(data, content_type='application/json')
    else:
        return redirect('/login/')
    # return JsonResponse({'json_data': json_data})
    # json_data = json.dumps(list, ensure_ascii=False)
    # template = loader.get_template('ajax.html')
    # context = {'list': list}
    # data = template.render(context, request)

def load_Product(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    cookie_facebook_data   = GetCookie(request, 'cookie_facebook_data')
    cookie_google_data     = GetCookie(request, 'cookie_google_data')
    if(cookie_microsoft_data or cookie_google_data or cookie_facebook_data or cookie_system_data):
        return render(request, 'ajax.html')
    else:
        return redirect('/login/')

@csrf_exempt
def delete_Product(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        try:
            product = Product.objects.get(id=product_id)
            product.delete()
            return JsonResponse({'success': True})
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'ID ' + product_id +' does not exist'})
    else:
        return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ'})

@csrf_exempt
def add_product(request):
    if request.method == 'POST':
        # print(request.POST)
        # data = json.loads(request.data)
        name = request.POST.get('name')
        id_cate = request.POST.get('category')
        price = request.POST.get('price')
        category = Category.objects.get(pk=id_cate)

        # name = data['name']
        # id_cate = data['category']
        # category = Category.objects.get(pk=id_cate)
        # price = data['price']
        product = Product.objects.create(name=name, id_cate=category, category=category.name, price=price)
        product.save()
        # return JsonResponse(response, status=201)
        return JsonResponse({
                'success': True,
                'message': 'Product added successfully!',
                'name': name,
                'category': category.name,
                'price': price
            })
    else:
        response = {'success': False, 'message': 'Invalid request method'}
        # return JsonResponse(response, status=405)
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
            })

def Data_Category(request):
      data =  Category.objects.using('default').all()
      List = [{'id': c.id, 'category': c.name} for c in data]
      cate = json.dumps(List, ensure_ascii=False)
      return HttpResponse(cate, content_type='application/json')

@csrf_exempt
def Data_Update_Product(request):
       if request.method == 'POST':
        id = request.POST.get('id_product')
        product = Product.objects.prefetch_related('id_cate').filter(id = id)
        list = [{'id': p.id, 'name': p.name, 'category': p.category, 'price': int(p.price), 'id_cate': p.id_cate.pk} for p in product]
        data = json.dumps(list, ensure_ascii=False)
        return HttpResponse(data, content_type='application/json')

@csrf_exempt
def update_product(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        price = request.POST.get('price')

        # Cập nhật đối tượng Product trong CSDL
        product = Product.objects.get(id=id)
        categrogy = Category.objects.get(id=category_id)
        product.name = name
        product.id_cate = categrogy
        product.category = categrogy.name
        product.price = price
        product.save()
        # list = [{
        #     'success': True,
        #     'message': 'Product updated successfully!',
        #     'id': p.id,
        #     'name': p.name,
        #     'category': p.category,
        #     'price': int(p.price),
        #     'id_cate': p.id_cate.pk} for p in product]
        # data = json.dumps(list, ensure_ascii=False)
        # return HttpResponse(data, content_type='application/json')
        return JsonResponse({
            'success': True,
            'message': 'Product updated successfully!',
            'id': product.id,
            'name': product.name,
            'category': product.category,
            'price': int(product.price),
            'id_cate': product.id_cate.pk
            })

    else:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
            })
#function Product


#function Login
@csrf_exempt
def login_system_1(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('pass')
            remember = request.POST.get('checkbox')
            user = Users.objects.filter(~Q(User_Type__lt= 2), Mail = email, Password = password, User_Status = True,) # 0 - administrator , 1 - Mod
            if(user):
                 # Lưu dữ liệu vào session
                 for user in user:
                    request.session['UserInfo'] = {
                        'ID_user': user.ID_user,
                        'Mail': user.Mail,
                        'FullName': user.FullName,
                        'displayName': user.displayName,
                        'Avatar': user.Avatar,
                        'Photo': user.displayName[0] +user.FullName[0],
                    }
                    session_data = json.dumps(request.session['UserInfo'])

                    # Lưu chuỗi JSON vào cookie
                    # Lưu access_token vào cookie (lưu ý: cần kiểm tra tính bảo mật của cookie)
                    if(remember == 'true'):
                        url_redirect = '/danh-sach-test/'
                        response = HttpResponse('cookie success')
                        response = SetCookie(response , 'cookie_system_data', session_data, url_redirect)
                        response = DeleteCookie(response , 'cookie_microsoft_data')
                        response = DeleteCookie(response , 'cookie_google_data')
                        response = DeleteCookie(response , 'cookie_facebook_data')
                        return response                               
                 return JsonResponse({
                'success': True,
                'message': 'Login Successed',}) 
            else:
                return JsonResponse({
                'success': False,
                'message': 'Login Failed',})
    except Users.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'User does not exist',
        })
    except Exception as ex:
        return JsonResponse({
            'success': False,
            'message': f'Login Failed: {str(ex)}',
        })

# Page Login account office 365
def microsoft_login_token_1(request):
    code = request.GET.get('code')
    MICROSOFT_CLIENT_SECRET = settings.MICROSOFT_CLIENT_SECRET

    if code is None:
        # Redirect to the Microsoft login page
        url = 'https://login.microsoftonline.com/common/oauth2/v2.0/authorize'
        params = {
            'client_id': '5c17ff26-50a1-4003-bc31-f0545709c2f7',
            'response_type': 'code',
            'redirect_uri': 'https://localhost:8000/login/callback/',
            'scope': 'https://graph.microsoft.com/.default',
        }
        auth_url = url + '?' + urlencode(params)
        return redirect(auth_url)

    # Exchange the authorization code for an access token and a refresh token
    token_url = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'
    data = {
        'grant_type': 'authorization_code',
        'code': code,
        'redirect_uri': 'https://localhost:8000/login/callback/',
        'client_id': '5c17ff26-50a1-4003-bc31-f0545709c2f7',
        # 'client_secret': 'EeJ8Q~ip-6TA~p1C7Y9t24l81qig0lFv1t5CPdwO',
        'client_secret': MICROSOFT_CLIENT_SECRET,
    }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
    }
    response = requests.post(token_url, data=data, headers=headers)
    response_data = response.json()

    # Save the access token and refresh token in the session
    access_token = response_data['access_token']
    request.session['access_token'] = access_token
    if 'refresh_token' in response_data:
        refresh_token = response_data['refresh_token']
        request.session['refresh_token'] = refresh_token
    else:
        refresh_token = None  # or handle the error in some other way


    # Use the access token to get the user's profile
    user_url = 'https://graph.microsoft.com/v1.0/me'
    headers = {
        'Authorization': 'Bearer ' + access_token,
        'Accept': 'application/json',
    }
    response = requests.get(user_url, headers=headers)
    response_data = response.json()
    #get avatar
    photo_data = response_data.get('photo', {})
    photo_url = ''
    if 'small' in   photo_data:
        photo_url = photo_data['small']
    elif 'large' in photo_data:
            photo_url = photo_data['large']

    #check user exit
    email = response_data.get('mail')
    exit_IDuser = Users.objects.filter(Mail = email)
    if(exit_IDuser):
        for user in exit_IDuser:
            # Cậẻp nhật các thuộc tính của User
            user.Mail = email
            user.FullName = response_data.get('surname') + " " + response_data.get('givenName')
            user.displayName = response_data.get('displayName')
            if(response_data.get('jobTitle')):
                user.Jobtitle = response_data.get('jobTitle')
            if(response_data.get('mobilePhone')):
                user.Phone = response_data.get('mobilePhone')
            user.Avatar = photo_url

            user.save()
    else:
        #check data ID User -get new ID
        UserID = Check_IDUser()
        # Insert dữ liệu vào cơ sở dữ liệu
        user = Users(
            ID_user     =UserID,
            Mail        =email,
            Password    ='',
            FullName    =response_data.get('surname') + " " + response_data.get('givenName'),
            displayName =response_data.get('displayName'),
            Birthday    = response_data.get('birthday') if response_data.get('birthday') else datetime.date(1990, 1, 1),  # Chưa có thông tin về ngày sinh trong Microsoft Graph API
            Acc_Type    ='Microsoft',
            Address     ='',
            Jobtitle    =response_data.get('jobTitle') if response_data.get('jobTitle') else "",
            Phone       =response_data.get('mobilePhone') if response_data.get('jobTitle') else 0,
            Avatar      =photo_url,
            User_Type   = 2, # 2 - EndUser
            User_Status =True  #User_Status là True khi mới đăng ký        
        )
        user.save()


    # Lưu dữ liệu vào session
    request.session['UserInfo'] = {
        'ID_user': user.ID_user,
        'Mail': user.Mail,
        'FullName': user.FullName,
        'displayName': user.displayName,
        'Avatar': user.Avatar,
        'Photo': user.displayName[0] +user.FullName[0],
    }

    # Save the user's profile in a cookie
   # Chuyển đổi session data thành chuỗi JSON
    session_data = json.dumps(request.session['UserInfo'])

    # Lưu chuỗi JSON vào cookie
     # Lưu access_token vào cookie (lưu ý: cần kiểm tra tính bảo mật của cookie)
    url_redirect = '/danh-sach-test/'
    response = SetCookie(response , 'cookie_microsoft_data', session_data, url_redirect)
    response = DeleteCookie(response , 'cookie_google_data')
    response = DeleteCookie(response , 'cookie_facebook_data')  
    response = DeleteCookie(response , 'cookie_system_data')  
    return response

#Login account google
def google_login(request):
    # Lấy mã truy cập từ Google
    code = request.GET.get('code', None)
    if code:
        # Gửi yêu cầu lấy mã truy cập
        client_id = settings.GOOGLE_OAUTH2_CLIENT_ID
        client_secret = settings.GOOGLE_OAUTH2_CLIENT_SECRET
        # redirect_uri = settings.GOOGLE_OAUTH2_REDIRECT_URI
        redirect_uri = request.build_absolute_uri('/google/callback/') # URL callback
        url = 'https://accounts.google.com/o/oauth2/token'
        data = {
            'code': code,
            'client_id': client_id,
            'client_secret': client_secret,
            'redirect_uri': redirect_uri,
            'grant_type': 'authorization_code',
        }
        response = requests.post(url, data=data)
        if response.status_code == 200:
            # Lưu thông tin người dùng vào session và cookie
            json_data = response.json()
            access_token = json_data['access_token']

            # Lấy thông tin người dùng từ Google API
            user_info = get_user_info_google(access_token)  # Hàm lấy thông tin người dùng từ Google API
            email = user_info.get("email")
            User_exit =  check_User_exit(email)
            if(User_exit):
                 for user in User_exit:
                    if(user.User_Status == True):
                          # Cập nhật các thuộc tính của User
                        user.Mail = email
                        user.FullName = user_info.get('family_name') + " " + user_info.get('given_name')
                        user.displayName = user_info.get('displayName')
                        if(user_info.get('jobtitle')):
                            user.Jobtitle = user_info.get('jobtitle')
                        if(user_info.get('phone')):
                            user.Phone = user_info.get('phone')
                        user.Avatar = user_info.get('picture')
                        user.save()
                    else:
                        return HttpResponseRedirect('/login/')
            else:
                #check data ID User -get new ID
                UserID = Check_IDUser()
                # Insert dữ liệu vào cơ sở dữ liệu
                user = Users(
                    ID_user     = UserID,
                    Mail        = email,
                    Password    = '',
                    FullName    = user_info.get('family_name') + " " + user_info.get('given_name'),
                    displayName = user_info.get('displayName'),
                    Birthday    = user_info.get('birthday') if user_info.get('birthday') else datetime.date(1990, 1, 1),
                    Acc_Type    ='Google',
                    Address     ='',
                    Jobtitle    = user_info.get('jobtitle') if user_info.get('jobtitle') else "",
                    Phone       = user_info.get('phone') if user_info.get('phone') else 0,
                    Avatar      = user_info.get('picture'),
                    User_Status =True  #User_Status là True khi mới đăng ký
                )
                user.save()

            # Lưu dữ liệu vào session
            request.session['UserInfo'] = {
                'ID_user': user.ID_user,
                'Mail': user.Mail,
                'FullName': user.FullName,
                'displayName': user.displayName,
                'Avatar': user.Avatar,
                'Photo': user.displayName[0] +user.FullName[0],
            }

            # Lưu access_token vào session
            request.session['access_token'] = access_token
            session_data = json.dumps(request.session['UserInfo'])
    
            # Lưu access_token vào cookie (lưu ý: cần kiểm tra tính bảo mật của cookie)
            url_redirect = '/danh-sach-test/'
            # response = SetCookie(response , 'cookie_google_token', access_token, url_redirect)
            response = SetCookie(response , 'cookie_google_data', session_data, url_redirect)
            response = DeleteCookie(response , 'cookie_microsoft_data')
            response = DeleteCookie(response , 'cookie_facebook_data')
            response = DeleteCookie(response , 'cookie_system_data')
            return response
        else:
            return HttpResponseRedirect('/login/')
            # pass
    else:
        return HttpResponseRedirect('/login/')
        # pass

#Function login by facebook
@csrf_exempt
def facebook_login(request):
   # Lấy thông tin token từ request
    access_token = request.GET.get('access_token')

    # Gọi API Facebook để lấy thông tin người dùng
    response = requests.get(f'https://graph.facebook.com/v12.0/me?fields=id,name,email&access_token={access_token}')
    data = response.json()

    # Kiểm tra thông tin người dùng có hợp lệ không
    if 'email' not in data:
        return redirect('/login/')

    # Tìm hoặc tạo tài khoản người dùng
    user = authenticate(request, email=data['email'])
    if user is None:
        user = Users.objects.create_user(username=data['email'], email=data['email'], password=Users.objects.make_random_password())

    # Đăng nhập và tạo session cho người dùng
    login(request, user)
    return redirect('/')


def Login_page(request):
    cookie_system_data     = GetCookie(request, 'cookie_system_data')
    cookie_microsoft_data  = GetCookie(request, 'cookie_microsoft_data')
    # cookie_facebook_data   = GetCookie(request, 'cookie_facebook_data')
    # cookie_google_data     = GetCookie(request, 'cookie_google_data')
    if(cookie_microsoft_data):
        response_data = json.loads(cookie_microsoft_data)
        # Lưu dữ liệu vào session
        # a = request.session.get('User_Info').get('UserID')
        if 'UserInfo' in request.session:
            request.session['UserInfo'] = {
                'ID_user': response_data.get('ID_user'),
                'Mail': response_data.get('Mail'),
                'FullName': response_data.get('FullName'),
                'displayName': response_data.get('displayName'),
                'Avatar': response_data.get('Avatar'),
                'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
        }
        return redirect('/danh-sach-test/')
    # elif(cookie_google_data):
    #     response_data = json.loads(cookie_google_data)
    #     # Truy cập các giá trị trong response_data
    #     request.session['UserInfo'] = {
    #         'ID_user': response_data.get('ID_user'),
    #         'Mail': response_data.get('Mail'),
    #         'FullName': response_data.get('FullName'),
    #         'displayName': response_data.get('displayName'),
    #         'Avatar': response_data.get('Avatar'),
    #         'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
    #     }
    #     return redirect('/danh-sach-test/')
    # elif(cookie_facebook_data):
    #     response_data = json.loads(cookie_facebook_data)
    #     request.session['UserInfo'] = {
    #         'ID_user': response_data.get('ID_user'),
    #         'Mail': response_data.get('Mail'),
    #         'FullName': response_data.get('FullName'),
    #         'displayName': response_data.get('displayName'),
    #         'Avatar': response_data.get('Avatar'),
    #         'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
    #     }
    #     return redirect('/danh-sach-test/')
    elif(cookie_system_data):
        response_data = json.loads(cookie_system_data)
        if 'UserInfo' in request.session:
            request.session['UserInfo'] = {
                'ID_user': response_data.get('ID_user'),
                'Mail': response_data.get('Mail'),
                'FullName': response_data.get('FullName'),
                'displayName': response_data.get('displayName'),
                'Avatar': response_data.get('Avatar'),
                'Photo': response_data.get('displayName')[0] + response_data.get('FullName')[0],
        }
        return redirect('/danh-sach-test/')
    else:
        #load template in folder teamplates
        template =  loader.get_template('Login.html')
        return HttpResponse(template.render())
#function Login

# Hàm lấy thông tin người dùng từ Google API
def get_user_info_google(access_token):
    url = 'https://www.googleapis.com/oauth2/v1/userinfo'
    headers = {'Authorization': 'Bearer {}'.format(access_token)}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        # Parse dữ liệu json
        json_data = response.json()
        user_info = {
            'id': json_data.get('id'),
            'email': json_data.get('email'),
            'displayName': json_data.get('name'),
            'given_name': json_data.get('given_name'),
            'family_name': json_data.get('family_name'),
            'picture': json_data.get('picture'),
            'jobtitle': json_data.get('jobtitle'),
            'phone': json_data.get('phone'),
            'birthday': json_data.get('birthday')
        }
        return user_info
    else:
        # Xử lý lỗi khi không thể lấy thông tin người dùng từ Google API
        pass

# Hàm lấy thông tin người dùng từ Microsoft API
def get_user_info_microsoft(access_token):
    user_url = 'https://graph.microsoft.com/v1.0/me'
    headers = {
        'Authorization': 'Bearer ' + access_token,
        'Accept': 'application/json',
    }
    response = requests.get(user_url, headers=headers)
    if response.status_code == 200:
        # Parse dữ liệu json
        json_data = response.json()
        user_info = {
            'id': json_data.get('id'),
            'email': json_data.get('email'),
            'displayName': json_data.get('name'),
            'given_name': json_data.get('given_name'),
            'family_name': json_data.get('family_name'),
            'picture': json_data.get('picture'),
            'jobtitle': json_data.get('jobtitle'),
            'phone': json_data.get('phone'),
            'birthday': json_data.get('birthday')
        }
        return user_info
    else:
        # Xử lý lỗi khi không thể lấy thông tin người dùng từ Microsoft API
        pass

############################################ OTHER - END ############################################################

# def call_graph_api(request):    
#     mailid = ''
#     emails = []
#     authorization_code = request.GET.get('code')
#     full_host = request.build_absolute_uri('/') 

#     session_data = request.session.get('mail_id', {})
#     if session_data:
#         mailid = session_data.get('Mail_ID')
#         access_token = session_data.get('access_token')
#     else:
#         # mailid = request.session.get('mail_id')  
#         access_token = get_access_token(authorization_code,full_host)
    
#     if access_token == None :
#         return JsonResponse({'error': 'token is empty'}, status=400)
    
#     # Tính thời gian ngày hôm nay và ngày hôm trước
#     today = datetime.datetime.utcnow().date()
#     yesterday = today - timedelta(days=1)

#     # Đường dẫn thư mục trên máy chủ để lưu các tệp đính kèm
#     # attachment_dir = '/static/Asset/Attachment-Upload/'
#     attachment_dir = os.path.join(settings.BASE_DIR , 'static', 'Asset', 'Attachment-Upload')
#     attachment_dir_img = os.path.join(settings.BASE_DIR , 'static', 'Asset', 'Attachment-Image')
#     if(mailid):
#         response = requests.get(
#             # f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and subject eq \'{prefix}\' and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
#             f'https://graph.microsoft.com/v1.0/me/messages/{mailid}',
#             headers={'Authorization': f'Bearer {access_token}'}
#         )
#         if response.status_code == 200:
#          data_email = response.json()
#          emails.append(data_email)
#     else:
#         prefix = '[SDP]'
#         response = requests.get(
#             # f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and subject eq \'{prefix}\' and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
#             f'https://graph.microsoft.com/v1.0/me/messages?$filter=isRead eq false and contains(subject, \'{prefix}\') and (sentDateTime ge {yesterday.isoformat()} or receivedDateTime ge {yesterday.isoformat()}) and (sentDateTime lt {today.isoformat()} or receivedDateTime ge {today.isoformat()})',
#             headers={'Authorization': f'Bearer {access_token}'}
#         )
#         emails = response.json().get('value', [])
    
#     if emails:
#         for email in emails:
#             email_from = email['from']['emailAddress']['address']  # Địa chỉ email người gửi
#             user = Users.objects.filter(Mail = email_from, User_Status = True).first()
#             if user:
#                 email_id = email['id']  # ID của email
#                 email_subject = email['subject']  # Tiêu đề email          
#                 # Lấy danh sách người nhận
#                 recipients = email['toRecipients']
#                 recipient_addresses = [recipient['emailAddress']['address'] for recipient in recipients]

#                 # Lấy danh sách người được CC
#                 cc_recipients = email.get('ccRecipients', [])
#                 cc_addresses = [cc['emailAddress']['address'] for cc in cc_recipients]

#                 email_body = email['body']['content']  # Nội dung email

#                 #function create Ticket by mail 
#                 email_group_ticket = check_group_ticket(email_subject,email_body)           
#                 email_type_ticket = 1 #Type Hỗ Trợ
#                 email_company_ticket = user.Company_ID  

#                 # if 'inReplyTo' in email:
#                 #     # Lấy email ID của email gốc
#                 #     original_email_id = email['inReplyTo']['id']  
#                 Context = Create_Ticket_Mail(request,email_id, email_subject, email_body,email_type_ticket,email_company_ticket,email_group_ticket,email_from,recipients)
#                 if Context:
#                     email_attachments = email['hasAttachments']  # Danh sách đính kèm
#                     # Xử lý danh sách đính kèm
#                     # if email_attachments:
#                     # attachment_names = []
#                     attachment_url = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/attachments'
#                     attachment_response = requests.get(
#                         attachment_url,
#                         headers={'Authorization': f'Bearer {access_token}'}
#                     )               
#                     attachments = attachment_response.json().get('value', [])
#                     if attachments:
#                         #image
#                         soup = BeautifulSoup(email_body, 'html.parser')
#                         img_tags = soup.find_all('img')
#                         for img_tag in img_tags:
#                             src = img_tag.get('src', '')
#                             if src.startswith('cid:'):
#                                 content_id = src[4:]

#                                 for attachment in attachments:
#                                     if attachment['contentType'].startswith('image/') and attachment['contentId'] == content_id:
#                                         img_data = attachment['contentBytes']

#                                         img_filename = f'{content_id}.png'  # Tên tệp hình ảnh
#                                         img_path = os.path.join(attachment_dir_img, img_filename)  # Đường dẫn tới tệp hình ảnh trên máy chủ

#                                         with open(img_path, 'wb') as img_file:
#                                             img_file.write(base64.b64decode(img_data))
                                        
#                                         with open(img_path, 'rb') as img_file:
#                                             img_base64 = base64.b64encode(img_file.read()).decode('utf-8')
#                                             # img_src = f'data:{attachment["contentType"]};base64,{img_base64}'
#                                             img_src = f'/static/Asset/Attachment-Image/{content_id}.png'
#                                             img_tag['src'] = img_src
#                                         break
#                         ticket = Ticket.objects.get(Ticket_ID = Context['Ticket_ID']) 
#                         if ticket:
#                             ticket.Ticket_Desc = str(soup)    
#                             ticket.save()                                   
#                         #attachment file
#                         for attachment in attachments:
#                             if not attachment['contentType'].startswith('image/'):
#                                 ticketID = Context['Ticket_ID']
#                                 attachment_id   = attachment['id']
#                                 attachment_name = attachment['name']
#                                 # attachment_type = attachment['contentType']
#                                 # attachment_size = attachment['size']
#                                 # attachment_names.append(attachment_name)

#                                 attachment_item = f'https://graph.microsoft.com/v1.0/me/messages/{email_id}/attachments/{attachment_id}'
#                                 attachment_item_response = requests.get(
#                                     attachment_item,
#                                     headers={'Authorization': f'Bearer {access_token}'}
#                                 )
#                                 attachment_data = attachment_item_response.content
#                                 # Xây dựng đường dẫn lưu tệp trên máy chủ
#                                 current_datetime = datetime.datetime.now()
#                                 numeric_date = current_datetime.strftime('%d%m%Y')
#                                 numeric_time = current_datetime.strftime('%H%M')
#                                 attachment_name_full = str(ticketID) + '_' + numeric_date + '_' + numeric_time + '_' + attachment_name
#                                 attachment_path = os.path.join(attachment_dir, attachment_name_full)

#                                 # Lưu tệp đính kèm vào máy chủ
#                                 with open(attachment_path, 'wb') as attachment_file:
#                                     attachment_file.write(attachment_data)
                                
#                                 #create attachment data
#                                 create_attachment_mail(ticketID,attachment_name_full)

#                     # reply_email(request, access_token, email_id, email_from, Context['Ticket_ID'], Context['Slug_Title'])             

#                     # else:
#                     #     #gửi mail thống báo lỗi không thành công.
#                     #     mail = "Gửi mail thông báo"
#                     if(mailid):
#                         del request.session['mail_id']
#                     reply_email(request, access_token, email_id, email_from, Context['Ticket_ID'], Context['Slug_Title'])
#                     send_email_new_ticket(request,email_id, email_from,Context['Ticket_ID'],Context['Slug_Title'],Context['Email_Assign'])
#                     # read_email(access_token,email_id)  
#             else:
#                 if(mailid):
#                     del request.session['mail_id']
#                 email_id = email['id']  # ID của email
#                 subject = "[SDP] Thông báo lỗi - User Chưa có trên hệ thống"    
#                 reply_email_NoExitUser(request, access_token, email_id, email_from, subject)           
#         return JsonResponse({'success': 'Success'}, status=200)
#     else:
#         return JsonResponse({'error': 'No email data'}, status=400)
#     # time.sleep(5)
#     return JsonResponse({'success': 'Success'}, status=200)
#     # return redirect('/get-code/')

